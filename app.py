import os
import openai
from flask import Flask, request, jsonify, send_file,session,redirect
from flask_cors import CORS

from PyPDF2 import PdfReader
import pandas as pd

from azure.identity import DefaultAzureCredential
from azure.cosmos import CosmosClient, PartitionKey
from azure.storage.blob import BlobServiceClient

import logging
from datetime import datetime,timezone,timedelta,UTC
import uuid
import json
import io
import re
import fitz  # PyMuPDF
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import time
import threading

import zipfile
import lxml.etree as ET
import os
import io
import ast
from azure.cosmos.exceptions import CosmosResourceNotFoundError, CosmosHttpResponseError
import secrets
from flask_session import Session
from werkzeug.security import generate_password_hash, check_password_hash
import urllib.parse
from io import StringIO
from asgiref.wsgi import WsgiToAsgi
import asyncio
import requests
import pdfplumber
from openpyxl.utils import get_column_letter
from copy import copy
from difflib import SequenceMatcher
import jaconv
import regex as regcheck

# 日志格式定义 (时间格式，日志级别，消息)
log_format = '%(asctime)sZ: [%(levelname)s] %(message)s'

# 日志设定: 时间格式，日志级别，消息
logging.basicConfig(
    level=logging.INFO,  # 日志级别 (DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format=log_format,   # 日志格式
    handlers=[logging.StreamHandler()]
)

# Managed Identity Auth
credential = DefaultAzureCredential()
token_OPENAI = credential.get_token("https://cognitiveservices.azure.com/.default")
token_COSMOS = credential.get_token("https://cosmos.azure.com/.default")

# Flask app init
app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # 安全密钥
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # 会话有效期30分钟

# 🔹 Flask sesstion settings (save to file system)
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_COOKIE_SECURE"] = False 
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "None"
app.config["SESSION_COOKIE_NAME"] = "secure_session"  # session cookie name


Session(app)

# CORS(app, resources={r"/api/*": {"origins": "*"}})
CORS(app, supports_credentials=True, resources={
    r"*": {
        "origins": "*"  # need change to real domain
    }
})



# 模拟用户数据库
users = {
    "admin": {"password": "123"},
    "user": {"password": "123"}
}

#-----------------------------------------------------------------
# Azure OpenAI Setting
# openai.api_type = "azure"
# openai.api_key = os.getenv("AZURE_OPENAI_KEY")  # Get ENV API Key

# COSMOS_DB_KEY = os.getenv("COSMOS_DB_KEY")  # Cosmos DB Key
#-----------------------------------------------------------------


# AzureTokenCache class define
class AzureTokenCache:
    def __init__(self):
        self._lock = threading.Lock() # thredd safe lock
        self.credential = DefaultAzureCredential()
        self.scope = "https://cognitiveservices.azure.com/.default"
        
        self.cached_token = None
        self.token_expires = 0
        self.last_refreshed = 0
        
        self._refresh_token()
        self._start_refresh_thread()

    def get_token(self):
        with self._lock:
            # token 10 minute end before
            if time.time() >= self.token_expires - 600:  # 10 mintute befor end
                self._refresh_token()
            return self.cached_token

    def _acquire_new_token(self):
        """Get new token"""
        return self.credential.get_token(self.scope)

    def _refresh_token(self):
        """update token"""
        new_token = self._acquire_new_token()
        with self._lock:
            self.cached_token = new_token.token
            self.token_expires = new_token.expires_on
            self.last_refreshed = time.time()
        print(f"🔄Updated Token (END of at:,haha, {self._format_time(self.token_expires)})")

    def _start_refresh_thread(self):
        thread = threading.Thread(target=self._refresh_loop, daemon=True)
        thread.start()

    def _refresh_loop(self):
        while True:
            time.sleep(30)
            if time.time() >= self.token_expires - 600:
                self._refresh_token()

    def _format_time(self, timestamp):
        local_time = time.localtime(timestamp)
        adjusted_time = time.mktime(local_time) + (8 * 3600)  # 8小时
        return time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(adjusted_time))
# -------------------------------------------------------------------
token_cache = AzureTokenCache()
#---------

# token method
openai.api_type = "azure_ad"
openai.api_base = os.getenv("AZURE_OPENAI_ENDPOINT")  # Get Env
openai.api_version = os.getenv("AZURE_OPENAI_API_VERSION")  # API Version
deployment_id = os.getenv("AZURE_OPENAI_MODEL")  # Get Deploy Name(mini-ZZ)
_deployment_id = os.getenv("AZURE_OPENAI_MODEL_4")  # Get Deploy Name(mini-ZZ)

# Cosmos DB 连接 
COSMOS_DB_URI = os.getenv("COSMOS_DB_URI")
DATABASE_NAME = os.getenv("DATABASE_NAME")
CONTAINER_NAME = os.getenv("CONTAINER_NAME")  # debug not used

# Azure Storage
ACCOUNT_URL = os.getenv("ACCOUNT_URL")
STORAGE_CONTAINER_NAME = os.getenv("STORAGE_CONTAINER_NAME")

MAX_TOKENS=32768 # 16384 for _deployment_id
TEMPERATURE=0
SEED=42
PDF_DIR = ACCOUNT_URL + STORAGE_CONTAINER_NAME

# Cosmos DB
def get_db_connection(CONTAINER):
    # Cosmos DB 链接客户端
    client = CosmosClient(COSMOS_DB_URI, credential=credential)
    database = client.get_database_client(DATABASE_NAME)
    container = database.get_container_client(CONTAINER)
    print("Connected to Azure Cosmos DB SQL API")
    logging.info("Connected to Azure Cosmos DB SQL API")
    return container  # Cosmos DB

#-----------------------------------------------------------------
LOG_RECORD_CONTAINER_NAME = "log_record"
FILE_MONITOR_ITEM = "file_monitor_item"
TENBREND_CONTAINER_NAME = 'tenbrend_history'
PROXYINFO_CONTAINER_NAME = 'proxyInfo'
INTEGERATION_RURU_CONTAINER_NAME = 'integeration_ruru'
#-----------------------------------------------------------------
integeration_container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

# List proxy
@app.route('/api/proxyinfo', methods=['GET'])
def get_proxyinfos():
    # Cosmos DB 连接
    container = get_db_connection(PROXYINFO_CONTAINER_NAME)
    
    query = "SELECT * FROM c"
    users = list(container.query_items(query=query, enable_cross_partition_query=True))
    response = {
        "code": 200,
        "data": users
    }
    return jsonify(response), 200

# Create proxy
@app.route('/api/proxyinfo', methods=['POST'])
def create_proxyuser():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Missing username or password"}), 400

    # Cosmos DB 连接
    container = get_db_connection(PROXYINFO_CONTAINER_NAME)

    # 确认用户
    query = "SELECT * FROM c WHERE c.username = @username"
    params = [dict(name="@username", value=username)]
    existing_users = list(container.query_items(
        query=query, 
        parameters=params, 
        enable_cross_partition_query=True
    ))

    if existing_users:
        return jsonify({"error": "Username already exists"}), 409  # HTTP 409 Conflict

    user_item = {
        'id': str(uuid.uuid4()),
        'username': username,
        'password': password  # 密码 hashing
    }
    container.create_item(body=user_item)
    response = {
        "code": 200,
        "data": user_item
    }

    return jsonify(response), 201

# update proxy
@app.route('/api/proxyinfo', methods=['PUT'])
def update_proxyuser():
    try:
        data = request.get_json()
        new_username = data.get('username')
        new_password = data.get('password')

        if not all([new_username, new_password]):
            return jsonify({"error": "Required fields: proxyuserName and Password"}), 400

        container = get_db_connection(PROXYINFO_CONTAINER_NAME)

        try:
            query = f"SELECT * FROM c"
            existing_user = list(container.query_items(
                query=query,
                enable_cross_partition_query=True
            ))[0]
        except IndexError:
            return jsonify({"error": "Find error error"}), 404

        proxy_data = dict(username=new_username, password=new_password)
        if existing_user:
            existing_user.update(proxy_data)
            container.upsert_item(existing_user)
        else:
            proxy_data.update(id=str(uuid.uuid4()))
            container.upsert_item(proxy_data)

        return jsonify({
            "username": new_username,
            "code": 200
        }), 200

    except CosmosHttpResponseError as e:
        logging.error(f"Cosmos DB Error: {str(e)}")
        return jsonify({"error": "DB Error"}), 500
    except Exception as e:
        logging.error(f"server error: {str(e)}")
        return jsonify({"error": "server error"}), 500
    
USERINFO_CONTAINER_NAME = 'userInfo'
#----------------------User CRUD--------
@app.route('/api/users', methods=['GET'])
def get_users():
    # Cosmos DB 连接
    container = get_db_connection(USERINFO_CONTAINER_NAME)

    query = "SELECT * FROM c"
    users = list(container.query_items(query=query, enable_cross_partition_query=True))
    response = {
        "code": 200,
        "data": users
    }
    return jsonify(response), 200

@app.route('/api/users', methods=['POST'])
def create_user():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Missing username or password"}), 400

    container = get_db_connection(USERINFO_CONTAINER_NAME)

    # 确认用户
    query = "SELECT * FROM c WHERE c.username = @username"
    params = [dict(name="@username", value=username)]
    existing_users = list(container.query_items(
        query=query, 
        parameters=params, 
        enable_cross_partition_query=True
    ))

    if existing_users:
        return jsonify({"error": "Username already exists"}), 409  # HTTP 409 Conflict

    user_item = {
        'id': str(uuid.uuid4()),
        'username': username,
        'password': generate_password_hash(password)
    }
    container.create_item(body=user_item)
    response = {
        "code": 200,
        "data": user_item
    }

    return jsonify(response), 201

@app.route('/api/users/<user_id>', methods=['PUT'])
def update_user(user_id):
    try:
        data = request.get_json()
        new_username = data.get('username')
        new_password = data.get('password')

        if not all([new_username, new_password]):
            return jsonify({"error": "username and password need input"}), 400

        container = get_db_connection(USERINFO_CONTAINER_NAME)

        try:
            query = f"SELECT * FROM c WHERE c.id = '{user_id}'"
            existing_user = list(container.query_items(
                query=query,
                enable_cross_partition_query=True
            ))[0]
        except IndexError:
            return jsonify({"error": "Do not find user"}), 404

        if existing_user['username'] != new_username:
            dup_query = f"SELECT * FROM c WHERE c.username = '{new_username}'"
            if list(container.query_items(dup_query, enable_cross_partition_query=True)):
                return jsonify({"error": "username duplicate"}), 409

        updated_item = {
            "id": user_id,
            "username": new_username,
            "password": generate_password_hash(new_password),
            **{k: v for k, v in existing_user.items() if k not in ['username', 'password']}
        }

        container.delete_item(item=user_id, partition_key=existing_user['id'])
        container.create_item(body=updated_item)

        return jsonify({
            "id": updated_item['id'],
            "username": updated_item['username']
        }), 200

    except CosmosHttpResponseError as e:
        logging.error(f"Cosmos DB error: {str(e)}")
        return jsonify({"error": "db error"}), 500
    except Exception as e:
        logging.error(f"server error: {str(e)}")
        return jsonify({"error": "server error"}), 500
            

@app.route('/api/users/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    container = get_db_connection(USERINFO_CONTAINER_NAME)
    
    container.delete_item(item=user_id, partition_key=user_id)
    return jsonify({"message": "User deleted"}), 200

#--------------------------------------
@app.before_request
def check_session():
    # 检查会话有效期
    if 'user_id' in session:
        last_activity = session.get('last_activity')
        session.modified = True
        if last_activity and (datetime.now() - datetime.fromisoformat(last_activity)) > app.config['PERMANENT_SESSION_LIFETIME']:
            session.clear()
            return jsonify({"status": "error", "message": "Session expired"}), 401
        # 更新最后活动时间
        session['last_activity'] = datetime.now().isoformat()

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip().lower()
    password = data.get('password', '').strip()

    if not username or not password:
        return jsonify({"status": "error", "message": "ユーザー名またはパスワードが間違っています"}), 400

    container = get_db_connection(USERINFO_CONTAINER_NAME)
    
    query = "SELECT * FROM c WHERE c.username = @username"
    params = [dict(name="@username", value=username)]

    items = list(container.query_items(
        query=query,
        parameters=params,
        enable_cross_partition_query=True
    ))

    if not items:
        return jsonify({"success": "false", "message": "User not found"}), 404

    user = items[0]
    if not check_password_hash(user['password'], password):
        return jsonify({"success": "false", "message": "Invalid password"}), 401

    session.clear()
    session['user_id'] = user['id']
    session['username'] = username

    return jsonify({"success": "true", "message": "ログイン成功！"}), 200

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"status": "success", "message": "ログアウト"}), 200

@app.route('/api/protected', methods=['GET'])
def protected():
    if not session.get('session_id'):
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    
    return jsonify({
        "status": "success",
        "message": "Protected content",
        "secure_session": session.get('session_id')
    }), 200


CHECK_SESSION_COOKIE = "session_cookie"

@app.route('/api/session_cookie', methods=['GET'])
def get_session_cookie():
    try:
        container = get_db_connection(CHECK_SESSION_COOKIE)
        
        query = "SELECT * FROM c"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        for item in items:
            item['id'] = item['id']

        return jsonify(items), 200
        
    except CosmosResourceNotFoundError:
        logging.error("Monitoring status document not found")
        return jsonify({"error": "Status document not found"}), 404
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    

@app.route('/api/session_cookie', methods=['PUT'])
def update_session_cookie():
    try:
        container = get_db_connection(CHECK_SESSION_COOKIE)
        
        # secure_session
        # session_value = request.cookies.get('secure_session', 'none')
        session_value = request.json.get('status', 'off')
        
        status_item = {
            'id': 'session_cookie',
            'type': 'control',
            'session_value': session_value,
            "timestamp": datetime.utcnow().isoformat()
        }
        
        container.upsert_item(body=status_item)
        logging.info(f"Session value updated: {session_value}")
        return jsonify({
            'message': 'Session value updated',
            'session_value': session_value
        }), 200
        
    except CosmosResourceNotFoundError as e:
        logging.error(f"Cosmos DB error: {str(e)}")
        return jsonify({"error": "Database operation failed"}), 500
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
#-----------API--------------------
def remove_code_blocks(text):
    text = re.sub(r'```html', '', text)
    text = re.sub(r'```', '', text)
    return text.strip()

def remove_code_blocks_enhance(text):
    text = re.sub(r'```html\n?', '', text)  
    text = re.sub(r'```', '', text)
    text = re.sub(r'\n\n\*\*NG\*\*\n```', '', text)
    return text.strip()


@app.route('/api/dic_search_db', methods=['POST'])
def dic_search_db():
    try:
        data = request.json

        original = data.get('original')
        corrected = data.get('corrected')

        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        query = f"SELECT * FROM c WHERE c.original = '{original}' AND c.corrected = '{corrected}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        if items:
            results = [{"original": item["original"], "corrected": item["corrected"]} for item in items]
            return jsonify({"success": True, "data": results}), 200
        else:
            return jsonify({"success": False, "message": "No matching data found in DB."}), 404

    except Exception as e:
        logging.error(f"❌ Error occurred while searching DB: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/ask_gpt', methods=['POST'])
def ask_gpt():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        prompt = data.get("input", "")

        if not prompt:
            return jsonify({"success": False, "error": "No input provided"}), 400

        # db get map
        corrected_map = fetch_and_convert_to_dict()

        # 3. apply_corrections
        corrected = apply_corrections(prompt, corrected_map)


        prompt_result = f"""
        You are a professional Japanese text proofreading assistant. 
        Please carefully proofread the content of a Japanese report following the rules below. 
        This includes not only Japanese text but also English abbreviations (英略語), foreign terms (外来語),
        and specialized terminology (専門用語). Ensure that all language elements are reviewed according to the guidelines and corrected where necessary.:

        **Report Content to Proofread:**
        {corrected}

        **Proofreading Requirements:**
        1. **Check for typos and missing characters (誤字脱字がないこと):**
        - Ensure there are no **spelling errors** or **missing characters** in the report. 
        - あなたの役割は、日本語の誤字・脱字・表記ミスを修正し、不完全な単語や文章に適切な語を補完することです。  
        以下のルールに従い、入力されたテキストを校正してください。
            **Common Mistakes Examples (誤字・脱字の例)**:
            Example:
            - `リテール投家` → `リテール投資家` (誤字: 家 → 資)
            - `長国債` → `長期国債` (脱字: 期を追加)
            - `識された` → `意識された` (表記の統一)
            - `金緩和期待` → `金融緩和期待` (誤字: 金 → 金融)
            - `見方が動し` → `見方が変動し` (誤字: 動し → 動き)
            - `視する` → `重視する`  
            - `経成長` → `経済成長`  
            - `送配電備` → `送配電設備`  
            - `業見通し` → `業績見通し`  
            - `常増益` → `経常増益`  
            - `財政策` → `財政政策`  
            - `方` → `方針`  
            - `手Eコマース` → `大手Eコマース`  
            - `響しました` → `影響しました`  
            - `施され` → `実施された`  
            - `企業の合併・収` → `企業の合併・回収`  
            - `本とします` → `基本とします`  
            - `務状況` → `財務状況`
            - `内投資信託` → `国内投資信託`  
            - `持しました` → `維持しました`  
            - `マイナス因` → `マイナス要因`  
            - `反される` → `反映される`  
            - `替ヘッジ` → `為替ヘッジ`  
            - `比は` → `比率は`
            - `規緩和` → `規制緩和`
            - `景済指標` → `経済指標`
            - `剤` → `経済`
            - `昇するなどまちまちでした。` → `異なる動きとなりました。` (Ensure that the original text is not directly modified but follows this guideline.)
            - `積極姿勢とした` → `長めとした` (Ensure that the original text is not directly modified but follows this guideline.)
            - `消極姿勢とした` → `長めとした` (Ensure that the original text is not directly modified but follows this guideline.)
            - `（割安に）放置` → `割安感のある`
            - `限定的` → `他の適切な表現に修正 （修正理由: 効果や影響がプラスかマイナスか不明瞭なため）`
            - `利益確定の売り` → `～が出たとの見方 （修正理由: 断定的な表現では根拠が説明できないため）`
            - `利食い売り` → `～が出たとの見方 （修正理由: 断定的な表現では根拠が説明できないため）`
            - `必ず～` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
            - `～になる` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
            - `～である` → `根拠が明示されていないため使用不可 （修正理由: 将来の運用成績や経済指標・企業業績等について断定的な判断を示す表現はNG）`
  
            **Disambiguation Rule**:
            - 「沈静」＝自然に落ち着く (natural calming down; happens over time)
            - 「鎮静」＝人為的におさめる (intentional suppression; medically or artificially done)

            **Correction Policy**:
            1. Detect whether the context implies a natural or artificial calming.
            2. If the usage does not match the context, correct it using the appropriate word.
            3. Highlight the correction using the format below:
            `<span style="color:red;">Corrected Term</span> (<span>修正理由: 意味の誤用 <s style="background:yellow;color:red">Original Term</s> → Corrected Term</span>)`
            4. Do **not** modify the original sentence structure or paragraph formatting.
            5. Only apply the correction when the term is clearly misused.
            6. If the current usage is correct, do not change or annotate it.

            **Example**:
            - Input: 市場は徐々に鎮静していった。
            - Output: 市場は徐々に <span style="color:red;">沈静</span> (<span>修正理由: 意味の誤用 <s style="background:yellow;color:red">鎮静</s> → 沈静</span>) していった。


        - 表現の使用制限:
            Expression Usage Restrictions:
            Restricted Expressions:

            - 魅力的な
            - 投資妙味
            - 割高感
            - 割安感

            Usage Conditions:
            The above expressions can be used if evidence is provided.

            However, these expressions should not be used in contexts where the word "fund" (ファンド) or any related reference is mentioned. In any sentence or context where "fund" or "ファンド" appears, these expressions should be avoided.

            使用例:
            魅力的な: 根拠に基づいて使用することは可能ですが、ファンドについては使用しないようにしてください。
            投資妙味: 投資妙味があることを示す場合でも、ファンドに対する言及は避け、他の投資対象に適用するようにしてください。
            割高感: 割高感について述べる場合、ファンド以外の投資対象に対して適用してください。
            割安感: 割安感について言及する場合も、ファンドに対して使用することは不可です。

            ✅ 出力フォーマット:
            <span style="color:red;">魅力的な</span>
            (<span>修正理由: ファンドに対しての使用は不可。</span>)
            ✅ Exsample1:
            Input:ファンドは魅力的な投資先として紹介された。

            Output:
            ファンドは
            <span style="color:red;">魅力的な</span>
            (<span>修正理由: ファンドに対する使用は不可</span>)投資先として紹介された。
            ✅ Exsample2:
            Input:
            このファンド銘柄には投資妙味がある。
            
            Output:
            この銘柄には
            <span style="color:red;">投資妙味</span>
            (<span>修正理由: ファンドに対しての使用は不可。</span>)がある。


        - 数字やパーセント（％）を含む文章の誤りをチェックする
            文脈を理解し、数値・割合の前後の語句が適切か確認すること。
            例:
            月末時点（20日判定）での為替ヘッジのターゲット比は48％です。
            → 正しい表記: 月末時点（20日判定）での為替ヘッジのターゲット比率は48％です。
            市場の成長率は10%の見込みです。 ✅ (問題なし)
            インフレ率は2上昇しました。 ❌ (誤り: 2%上昇しました。 に修正)
            販売シェアは15の拡大が予想されます。 ❌ (誤り: 販売シェアは15%の拡大が予想されます。)
        - 校正ルール
            1. **「行って来い」の適切な置き換え**(Ensure that the original text is not directly modified but follows this guideline.)
            - 文章全体を分析し、「行って来い」が何を指しているのかを判断してください。
            - **価格・指数・レートなどが上昇した意味の場合** → 「行って来い」を「上昇した」に変換
            - **価格・指数・レートなどが下落した意味の場合** → 「行って来い」を「下落した」に変換

            2. **文脈を考慮した校正**
            - 修正の際、周辺の文脈を理解し、自然な形に調整してください。
            
            3. **「横ばい」の適切な置き換え**
            - 文章全体を分析し、「横ばい」の前後の文脈を考慮してください。
            - **期中の変動幅が小さい場合** → 「横ばい」を維持
            - **変動幅が大きく、結果的に同程度となった場合** → 「ほぼ変わらず」または「同程度となる」に変更
            - 周囲の文章に合わせて、より適切な表現を使用してください。

            4. **「出遅れ感」の適切な修正**
            - 対応方針:

            「出遅れ感」 は主観的な相場観が含まれるため、必ず「…と考えます。」に修正してください。

            修正方法: 文脈に応じて自然に「〜と考えます」形に修正します。

            修正後の表現: 文の流れに合わせて自然に表現を調整し、読みやすさを考慮します。

            修正理由: 相場観が含まれているため、主観的な表現を客観的な表現に変えることで文章の信頼性を向上させます。

            出力フォーマット:
            <span style="color:red;">出遅れ感</span>
            (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 「〜と考えます」に修正</span>)

            Exsample:
                Input: この銘柄には出遅れ感がある
                Output: この銘柄には
                <span style="color:red;">出遅れ感</span>
                (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 出遅れていると考えます</span>)
                がある

                Input: 一部のセクターには出遅れ感があると感じられる
                Output: 一部のセクターには
                <span style="color:red;">出遅れ感</span>
                (<span>修正理由: 主観的表現の修正 <s style="background:yellow;color:red">出遅れ感</s> → 出遅れていると考えます</span>)
                がある


            5. **「上昇要因」・「下落要因」の適切な説明追加**
            - **文脈を分析し、具体的な要因を追加してください**。
            - **「上昇要因」がある場合** → 上昇の理由（例: 企業決算の改善、政策の発表、需給バランスの変化など）を補足。
            - **「下落要因」がある場合** → 下落の理由（例: 景気後退懸念、金融引き締め、地政学リスクなど）を補足。
            - **修正後も文章の流れがスムーズになるように調整してください。

            6. **「予想」「心理」の適切な修正**
            - **「予想」 がある場合:**  
                - **誰の予想か明確でない場合** → 「市場予想」に修正  
            - **「心理」 がある場合:**  
                - **主語が曖昧な場合** → 「市場心理」に修正

    
        2. **Follow the "Fund Manager Comment Terminology Guide" (ファンドマネージャコメント用語集に沿った記載となっていること):**
        - **Consistent Terminology (表記の統一):**
            - Ensure the **writing format** of financial terms is **consistent throughout the report**.
            Example:
            - `相対に低かった` → `相対的に低かった` (文法修正)
            - `東証33業種分では` → `東証33業種分類では` (表記の統一)
        - **Common Mistakes and Corrections (誤記と修正例)**:
            Example:
            - `政府支の` → `政府支出の遅れ、1回はのを出、2回はのを削除` (誤字: の → 出)
            - `投資比率を維する` → `投資比率を維持する` (一致性不足 動産 → 不動産)
            - `（配当こみ）` → `（配当込み）` (表記の統一: こみ → 込み)
            - `いっぽうで` → `一方で` (表記の統一: いっぽう → 一方)
            - `ウクライナとロシアをめぐる` → `ウクライナとロシアを巡る` (表記の統一: めぐる → 巡る)
            - `東証33業種でみると` → `東証33業種で見ると` (文法修正: みる → 見る)
            - `ひき続き` → `引き続き` (表記の統一: ひき → 引き)
            - `電気機器、銀行業、保険業等` → `電気機器、銀行業、保険業など` (表記の統一: 等 → など)
            - `じき` → `次期`
            - `底いれ後` → `底入れ後`
            - `なかから` → `中から`
            - `おもな` → `主要な`
            - `はやく` → `早急に`
            - `かいつけした` → `買い付けした`
            - `など` → `等`
            - `のぞく` → `除く`
            - `くみいれ` → `組み入れ`
            - `じょうき` → `上記`
            - `とうファンド` → `当ファンド`

        - **Prohibited Words and Phrases (禁止（NG）ワード及び文章の注意事項):**
            - Check if any prohibited words or phrases are used in the report and correct them as per the guidelines.
        - **Replaceable and Recommended Terms/Expressions (置き換えが必要な用語/表現、置き換えを推奨する用語/表現):**
            - If you find terms or expressions that need to be replaced, revise them according to the provided rules.
            - ハト派／タカ派の表記（金融政策に関する）:
                -Exsample:
                - 金融緩和重視  → 金融引き締め重視
                - 金融緩和に前向き  → 金融引き締めに積極的
            - Exsample:
             - 織り込む  → 反映され
             - 相場  → 市場/価格
             - 連れ高  → 影響を受けて上昇
             - 伝播  → 広がる
             - トレンド  → 傾向
             - レンジ  → 範囲
        - **〇％を上回る（下回る）マイナスの表記:**
                - 〇％を上回る  → 〇％を超える
                - 〇％を下回る  → 縮小
                - 〇％を上回る  → 下回るマイナス幅
                - 〇％を下回る  → 縮小


        - **Use of Hiragana (ひらがなを表記するもの):**
            - Ensure the report follows the rules for hiragana notation, replacing content that does not conform to commonly used kanji.
        - **Kana Notation for Non-Standard Kanji (一部かな書き等で表記するもの):**
            - Ensure non-standard kanji are replaced with kana as the standard writing format.
        - **Correct Usage of Okurigana (一般的な送り仮名など):**
            - Ensure the correct usage of okurigana is applied.
        - **English Abbreviations, Loanwords, and Technical Terms (英略語、外来語、専門用語など):**
            - Check if English abbreviations, loanwords, and technical terms are expressed correctly.
        - **Identify and mark any 常用外漢字 (Hyōgai kanji):**
        - **Identify and mark any **常用外漢字 (Hyōgai kanji)** in the following text**
        - **常用外漢字** refers to characters **not included** in the [常用漢字表 (Jōyō Kanji List)](https://ja.wikipedia.org/wiki/常用漢字), which is Japan’s official list of commonly used kanji.
        - Refer to the [Wikipedia list of Hyōgai kanji](https://ja.wikipedia.org/wiki/常用漢字) to determine if a character falls into this category.
        - **For any detected 常用外漢字**, apply the following formatting:
        - **Highlight the incorrect character in red** (`<span style="color:red;">`).
        - **Strike through the incorrect character and provide the reason in yellow highlight.**

        ---

        ### **💡 Output Format (出力フォーマット)**
        - **Incorrect characters should be displayed in red (`<span style="color:red;">`)**.
        - **Corrected text should be marked with a strikethrough (`<s>`) and highlighted in yellow (`background:yellow`)** to show the correction.
        - Use the following structure:
            ```html
            <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>)
            ```
        - **Example Correction**:
            ```plaintext
            鴉 → <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>)
            ```
        - **For multiple Hyōgai kanji**, apply the same structure to each character.

        ---

        ### **✅ Example Input:**
        ```plaintext
        彼は鴉が空を飛ぶのを見た。

        ### **✅ Example output:**
        ```plaintext
        彼は <span style="color:red;">鴉</span> (<span>修正理由: 常用外漢字 <s style="background:yellow;color:red">鴉</s></span>) が空を飛ぶのを見た。

        - **Foreign Exchange Market Trend Analysis**
            In the foreign exchange market (`為替市場`), determine whether `"円だか"` should be revised to `"円高"` (Yen Appreciation) or `"円安"` (Yen Depreciation) based on the **context**.

            #### **** Criteria for Yen Appreciation (円高)**
            - **Yen appreciation (`円高`) occurs when the value of the yen increases relative to other currencies.**  
            The following situations indicate yen appreciation:
            1. **"多くの通貨が対円で下落した"** (Many currencies declined against the yen) → Change `円だか` to **円高**.
            2. **"ドル円が下落した"** (USD/JPY exchange rate declined) → Change `円だか` to **円高**.
            3. **"対米ドルで円の価値が上昇した"** (The yen appreciated against the US dollar) → Change `円だか` to **円高**.

            #### **** Criteria for Yen Depreciation (円安)**
            - **Yen depreciation (`円安`) occurs when the value of the yen declines relative to other currencies.**  
            The following situations indicate yen depreciation:
            1. **"多くの通貨が対円で上昇した"** (Many currencies rose against the yen) → Change `円だか` to **円安**.
            2. **"ドル円が上昇した"** (USD/JPY exchange rate increased) → Change `円だか` to **円安**.
            3. **"対米ドルで円の価値が下落した"** (The yen depreciated against the US dollar) → Change `円だか` to **円安**.


        3. **Replaceable and Recommended Terms/Expressions (推奨される表現の修正)**
        - Use the correct **kanji, hiragana, and katakana** combinations based on standard Japanese financial terms.
            Example:
            - `が好された輸送用機器など` → `が好感された輸送用機器など` (修正理由: 適切な表現)

        - **Task**: Header Date Format Validation & Correction  
        - **Target Area**: Date notation in parentheses following "今後運用方針 (Future Policy Decision Basis)"  
        ---
        ### Validation Requirements  
        1. **Full Format Compliance Check**:  
        - Must follow "YYYY年MM月DD日現在" (Year-Month-Day as of)  
        - **Year**: 4-digit number (e.g., 2024)  
        - **Month**: 2-digit (01-12, e.g., April → 4)  
        - **Day**: 2-digit (01-31, e.g., 5th → 5)  
        - **Suffix**: Must end with "現在" (as of)  

        2. **Common Error Pattern Detection**:  
        ❌ "1月0日" → Missing month leading zero + invalid day 0  
        ❌ "2024年4月1日" → 2024年4月1日
        ❌ "2024年12月" → Missing day value  
        ❌ "2024-04-05現在" → Incorrect separator usage (hyphen/slash)  
        ---
        ### Correction Protocol  
        1. **Leading Zero Enforcement**  
        - Add leading zeros to single-digit months/days (4月 → 4月, 5日 → 5日)  

        2. **Day 0 Handling**  
        - Replace day 0 with YYYYMMDD Date Format  
        - Example: 2024年4月0日 → 2024年4月00日

        3. **Separator Standardization**  
        - Convert hyphens/slashes to CJK characters:  
            `2024/04/05` → `2024年4月5日`  

        ---
        ### Output Format Specification  
        ```html
        <Correction Example>
        <span style="color:red;">（2024年4月0日現在）</span> 
        → 
        <span style="color:green;">（2024年04月00日現在）</span>
        修正理由:
        ①日付0をYYYYMMDD日付フォーマットに置換
        ---

        3. **Consistency with Report Data Section (レポートのデータ部との整合性確認):**
        - Ensure the textual description in the report is completely consistent with the data section, without any logical or content-related discrepancies.

        4. **Eliminate language fluency(単語間の不要なスペース削除):**
        - Ensure that there are no extra spaces.
            -Example:
            input:景気浮揚が意 識されたことで
            output:景気浮揚が意識されたことで
        
        5.  **Layout and Formatting Rules (レイアウトに関する統一):**
            - **文頭の「○」印と一文字目の間隔を統一:**

                - When a sentence starts with the symbol ○, make sure there is no space (half-width or full-width) between it and the first character. That is, use ○文字 instead of ○ 文字 or ○　文字.
                    - Any whitespace (half-width or full-width spaces) after ○ must be removed.
                    - This spacing rule must be applied consistently throughout the document.

            半角括弧を全角括弧に統一:
                - Convert all half-width parentheses () to full-width parentheses （）.
                - Example: (注) → （注）
                - Example input: 
                    ○ 世界の高配当株式指数(注)は月間では上昇しました。
                - Exsample output: 
                    <span style="color:red;">○世界</span> (<span>修正理由: 文頭の「○」印と一文字目の間隔を統一 <s style="background:yellow;color:red">○ 世界</s> → ○世界</span>)
                    <span style="color:red;">（注）</span> (<span>修正理由: 半角括弧を全角括弧に統一 <s style="background:yellow;color:red">(注)</s> → （注）</span>)


            - **文章の間隔の統一:**
                - If a sentence begins with "○", ensure that the spacing within the frame remains consistent.
            - **上位10銘柄 コメント欄について、枠内に適切に収まっているかチェック:**
                - If the stock commentary contains a large amount of text, confirm whether it fits within the designated frame. 
                - If the ranking changes in the following month, adjust the frame accordingly.
                - **Check point**
                    1. **文字数制限内に収まっているか？**
                    - 1枠あたりの最大文字数を超えていないか？
                    - 適切な行数で収まっているか？

                    2. **次月の順位変動に伴う枠調整の必要性**
                    - 順位が変更されると枠調整が必要なため、調整が必要な箇所を特定

                    3. **枠内に収まらない場合の修正提案**
                    - 必要に応じて、短縮表現や不要な情報の削除を提案
                    - 重要な情報を損なわずに適切にリライト

                    output Format:
                    - **コメントの枠超過チェック**
                    - (枠超過しているか: はい / いいえ)
                    - (超過している場合、オーバーした文字数)

                    - **順位変動による枠調整の必要性**
                    - (調整が必要なコメントリスト)

                    - **修正提案**
                    - (枠内に収めるための修正後のコメント)

            **Standardized Notation (表記の統一):**
            - **基準価額の騰落率:**
            When there are three decimal places, round off using the round-half-up method to the second decimal place. If there are only two decimal places, keep the value unchanged.
                Make modifications directly in this article and explain the reasons for the modifications.

                exsample:
                0.546％（×） → 0.55％（○）
                修正理由: 小数点以下の桁数の丸め（0.546％ → 0.55％）
                If the value is 0.00％, replace it with "前月末から変わらず" or "前月末と同程度" instead of stating "騰落率は変わらず".
                修正理由: 「騰落率は変わらず」という表記はNG。代わりに「基準価額（分配金再投資）は前月末から変わらず」や「前月末と同程度」と記載します。

                exsample:
                0.00％となり（×） → 前月末から変わらず（○）

                騰落率は変わらず（×） → 基準価額（分配金再投資）は前月末から変わらず（○）

                When comparing the performance of the fund with the benchmark (or reference index), the comparison must be made using rounded numbers.

                修正理由: 比較は丸めた数字で行なうこと。
                If the fund and benchmark (or reference index) have the same rate of return, use the phrase "騰落率は同程度となりました" instead of saying "騰落率は同じでした".
                修正理由: 同じという表現は避け、代わりに「同程度」と記載すること。

                exsample:
                「騰落率は同じでした」（×） → 「騰落率は同程度となりました」（○）
                If the fund's rate of return is greater than the benchmark's, use the phrase "上回りました" to indicate the fund outperformed the benchmark.
                修正理由: 上回った場合、「上回りました」と表記すること。

                exsample:
                騰落率は-1.435％（基金）と-2.221％（ベンチマーク）の場合、値の差は0.79％となるため、「上回りました」と記載します。

                If the fund's rate of return is lower than the benchmark's, use the phrase "下降しました" to indicate the fund underperformed the benchmark.
                修正理由: 下降した場合、「下降しました」と表記すること。

                exsample:
                騰落率は-1.435％（基金）と-0.221％（ベンチマーク）の場合、基金のパフォーマンスは「下降しました」と記載します。

            - **「今後の運用方針」作成日付のルール:**
                - 前月末（営業日）現在で作成。
                - 翌月初の日付になる場合は、作成した日付を入れる。
                - クライアント・サービス部へ送信する以降の日付（先日付）は入れない。
                - 「参考月」より後であり、「チェック期間」より前の日付のみ使用可。
                    - Example:
                    - OK: 参考月＝2024年2月 → 作成日が2024年2月28日（営業日） or 3月1日（翌月初）
                    - NG: 参考月＝2024年2月 → 作成日が3月5日（クライアント・サービス部送信後の先日付）

            - **％（パーセント）、カタカナ:**
                - **半角カタカナ → 全角カタカナ**（例:「ｶﾀｶﾅ」→「カタカナ」）
                - **半角記号 → 全角記号**（例:「%」→「％」、「@」→「＠」）
                    Example:
                        input: ﾍﾞﾝﾁﾏｰｸ (修正理由: 半角カタカナを全角カタカナに統一 ﾍﾞﾝﾁﾏｰｸ → ベンチマーク)に対して 
                        output: ベンチマーク (修正理由: 半角カタカナを全角カタカナに統一 ﾍﾞﾝﾁﾏｰｸ → ベンチマーク)に対して
                    Example:
                        input: ｶﾀｶﾅ 
                        output: カタカナ
                    Example:
                        input: %
                        output: ％ 
                    Example:
                        input: @
                        output: ＠ 

            - **数字、アルファベット、「＋」・「－」:**
                - **全角数字・アルファベット → 半角数字・アルファベット**（例:「１２３」→「123」、「ＡＢＣ」→「ABC」）
                - **全角「＋」「－」 → 半角「+」「-」**（例:「＋－」→「+-」
                    Example:
                        input: １２３ ＡＢＣ ｱｲｳ ＋－
                        output: 123 ABC アイウ +-

            - **スペースは変更なし**  

            - **「※」の使用:**
                - 「※」は可能であれば **上付き文字（superscript）※** に変換してください。
                - 出力形式の例:
                - 「重要事項※」 → 「重要事項<sup>※</sup>」

            - **（カッコ書き）:**
                - Parenthetical notes should only be included in their first occurrence in a comment.
                    For the following Japanese text, check if parentheses ("（ ）") are used appropriately.
                    If a parenthetical note appears more than once, remove the parentheses for subsequent occurrences.
                    The first occurrence should retain the parentheses, but any further appearances should have the parentheses removed.
                    Modification reason: Parentheses are redundant after the first mention, so the text is cleaned up for consistency and readability.

                **Check point**
                    1. **カッコ書きは、コメントの初出のみに記載されているか？**
                    - 同じカッコ書きが2回以上登場していないか？
                    - 初出ページ以降のコメントにカッコ書きが重複して記載されていないか？

                    2. **ディスクロのページ番号順に従ってルールを適用**
                    - シートの順番ではなく、実際のページ番号を基準にする。

                    3. **例外処理**
                    - 「一部例外ファンドあり」とあるため、例外的にカッコ書きが複数回登場するケースを考慮する。
                    - 例外として認められるケースを判断し、適切に指摘。

                    output Format:
                    - **カッコ書きの初出リスト**（どのページに最初に登場したか）
                    - **重複チェック結果**（どのページで二重記載されているか）
                    - **修正提案**（どのページのカッコ書きを削除すべきか）
                    - **例外ファンドが適用される場合、補足情報**

            - **会計期間の表記:**
                - The use of "～" is prohibited; always use "-".
                - Make modifications directly in this article and explain the reasons for the modifications.
                    - Example: 6～8月期（×） → 6-8月期（○）

                - 暦年を採用している国の年度表記:
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                カッコ書きで暦年の期間を明記する。
                - Example:
                    ブラジルの2021年度予算（×） → ブラジルの2021年度（2021年1月-12月）予算（○）

                - 決算期間は「●-●月期」に統一し、日付は省略する。
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                    - Example: 第1四半期（5月21日～8月20日）（×） → 5-8月期（○）
                    イレギュラーなケースも含め、原則「●-●月期」と表記。

            - **「TOPIX」または「東証株価指数」が含まれている場合、以下のルールを適用:**
                文中で使用する場合: 「TOPIX（東証株価指数）」と表記することを指示。
                「文中では『TOPIX（東証株価指数）』と表記してください。(Ensure that the original text is not directly modified but follows this guideline.)」
                ベンチマーク（BM）や参考指数として使用する場合: 「東証株価指数（TOPIX）（配当込み）」と表記することを指示。
                「BMや参考指数で使用する場合は、『東証株価指数（TOPIX）（配当込み）』と表記してください。(Ensure that the original text is not directly modified but follows this guideline.)」

            - **年をまたぐディスクロコメントの年度表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - When specifying the fiscal year in disclosure comments that span multiple years, always use "yyyy年度".
                - Similarly, for disclosures based on the January-March period, specify the corresponding year.
                - Example:
                    - For a disclosure with a December-end reference, released in January:
                    - 今年度（×） → 2021年度（○）
                    - 来年度（×） → 2022年度（○）
            - **Benchmark, Index, and Reference Index Name Formatting:(ベンチマーク・インデックス・参考指数の名称の表記)
                - Ensure Consistency in Index Terminology:
                    Read the context and identify terms related to "index" (指数) within the text. Ensure that these terms are unified and consistently referred to using the correct and standardized terminology.
                    It is important to carefully analyze each mention of "index" to make sure the terminology is consistent throughout the text.
                    Do not modify the original text directly. Instead, provide comments that explain the reasoning behind the proposed changes, especially when identifying inconsistencies or clarifications needed.
                Example Formatting Guidelines:

                    Incorrect format (×): "ISM非製造業景況"
                    Correct format (○): "ISM非製造業景況指数"
                    Incorrect format (×): "MSCIインド"
                    Correct format (○): "MSCIインド・インデックス"
                    If multiple terms are used to refer to the same index, they should be unified under the correct term. For example, if "MSCIインド指数" and "MSCIインド・インデックス" are used in different places, they should be unified as "MSCIインド・インデックス" in the final report to maintain consistency.
                Handling Multiple Terms Referring to the Same Index:

                    If it can be clearly determined that different terms refer to the same index (e.g., "MSCIインド指数" and "MSCIインド・インデックス"), do not modify them but mark them accordingly. These terms should be noted as referring to the same index.
                    Example:
                    Original: "MSCIインド指数" and "MSCIインド・インデックス"
                    Comment: These are different ways of referring to the same index, so no change is needed.
                Handling Uncertainty in Index Terminology:
                    
                    If there is uncertainty about whether multiple terms refer to the same index (e.g., it is unclear whether "ISM非製造業景況" and "ISM非製造業景況指数" refer to the same index), mark them without modification. Additionally, note that these terms might refer to the same index, but the exact nature of the index should be verified.
                    Example:
                    Original: "ISM非製造業景況" and "ISM非製造業景況指数"
                    Comment: These terms are potentially referring to the same index but require further clarification. Therefore, no changes are made in this case.
                Key Notes:

                Always ensure that consistency is maintained across the report. Even if different names are used for the same index, it is essential to mark them properly and explain that they are different terms for the same entity.
                Consistency applies not only to the formatting of the terms but also to how the terms are presented across the entire document. All references to a given index must follow the same format from the first mention to the last.


            - **上昇 or 下落に関する要因を明記:**
                文章内に 「上昇」 または 「下落」 という単語が含まれている場合、その要因を特定し、明記してください。
                output:
                「●●は上昇（または下落）しました。(理由: ○○)」
            - **指定用語の表記ルールを提示:**
                独Ifo企業景況感指数 または 独Ifo景況感指数 が含まれている場合、以下のメッセージを表示する。
                独Ifo企業景況感指数 または独Ifo景況感指数 の表記について、月報内での統一ルールを確認してください。(Ensure that the original text is not directly modified but follows this guideline.)」
                
                「独ZEW景気期待指数」または「独ZEW景況感指数」 が含まれている場合、以下のメッセージを表示する。
                「『独ZEW景気期待指数』または『独ZEW景況感指数』の表記を使用してください。ZEW単独使用の場合は括弧書き付き、または『欧州経済研究センター』のみとし、『ZEW』単独使用を避けてください。(Ensure that the original text is not directly modified but follows this guideline.)」
            - **特定の金融用語に対する表記ルールを確認・適用:**

                「リフレーション」 が含まれている場合、以下のメッセージを表示。
                「リフレーションとは、デフレーションから抜けて、まだインフレーションにはなっていない状況を指します。」

                「リスクプレミアム」 が含まれている場合、以下のメッセージを表示。
                「リスクプレミアムとは、あるリスク資産の期待収益率が、同期間の無リスク資産（国債など）の収益率を上回る幅を指します。」

                「モメンタム」 が含まれている場合、以下のメッセージを表示。
                「『モメンタム』の使用は避け、相場の『勢い』や『方向性』などの言葉に置き換えてください。」

                「ベージュブック」 が含まれている場合、以下のメッセージを表示。
                「『ベージュブック』は『FRB（米連邦準備制度理事会）が発表したベージュブック（地区連銀経済報告）』と明記してください。スペースが限られる場合は、『ベージュブック（米地区連銀経済報告）』としてください。」

                「フリーキャッシュフロー」 が含まれている場合、以下のメッセージを表示。
                「フリーキャッシュフローとは、税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたものです。」
                
                「システミック・リスク」 が含まれている場合、以下のメッセージを表示。
                「システミック・リスクとは、個別の金融機関の支払不能や特定の市場・決済システム等の機能不全が、他の金融機関、市場、または金融システム全体に波及するリスクを指します。」

                「クレジット（信用）市場」 が含まれている場合、以下のメッセージを表示。
                「クレジット（信用）市場とは、信用リスク（資金の借り手の信用度が変化するリスク）を内包する商品（クレジット商品）を取引する市場の総称であり、企業の信用リスクを取引する市場です。」
            - **特定の金融用語に対し、欄外に注記を加える指示を表示:**
                「格付別」 が含まれている場合:
                    格付別 -> 格付別
                「格付機関」 が含まれている場合:
                    格付機関 -> 格付機関
                「組入比率」 が含まれている場合:
                    組入比率 -> 組入比率
                「引締策」 が含まれている場合:
                    引締策 -> 引締策
                「国債買入れオペ」 が含まれている場合:
                    国債買入れオペ -> 国債買入オペ

                「投資適格債」 が含まれている場合:
                「※欄外に注記: 投資適格債とは、格付機関によって格付けされた公社債のうち、債務を履行する能力が十分にあると評価された公社債を指します。」

                「デュレーション」 が含まれている場合:
                「※欄外に注記: デュレーションとは、金利が一定の割合で変動した場合、債券の価格がどの程度変化するかを示す指標です。この値が大きいほど、金利変動に対する債券価格の変動率が大きくなります。」

                「デフォルト債」 が含まれている場合:
                「※欄外に注記: デフォルトとは、一般的に債券の利払いおよび元本返済の不履行、または遅延などを指し、このような状態にある債券を『デフォルト債』といいます。」

                「ディストレス債券」 が含まれている場合:
                「※欄外に注記: ディストレス債券とは、信用事由などにより価格が著しく下落した債券を指します。」
                
                「イールドカーブ」 が含まれている場合:
                「※欄外に注記: イールドカーブ（利回り曲線）とは、横軸に残存年数、縦軸に利回りをとった座標に、債券利回りを点描して結んだ曲線のことを指します。」

            - **組入上位10銘柄」について記述がある場合、以下のルールを適用:**
                「組入上位10銘柄を超える保有銘柄（個別銘柄の特定が可能な子会社名等を含む）は原則として開示禁止である」ことを明示。
                ただし、社内規程に基づき開示が認められているファンドは例外とすることを伝える。

            - **年度表記:**
                - Use four-digit notation for years.(Ensure that the original text is not directly modified but follows this guideline.)
                - Example: 22年（×） → 2022年（○）

            - **前年比 or 前年同月（同期）比の統一:**
                - 「前年同月（同期）比」に統一。
                - 通年の比較には「前年比」の使用可。
                - Ensure that the original text is not directly modified but follows this guideline.
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                    - Example:
                        前年比+3.0%（×） → 前年同月比+3.0%（○）
                        2023年のGDPは前年比+3.0％（○）

            - **年をまたいだ経済指標の記載:**
                - コメント内の初出のみに記載する。(In the case where there is a description of the economic indicator over the year, it is described only in the first comment.)
                    - Example:
                        - 2023年12月のCPIは～（○）
                        - 一方2024年1月のユーロ圏PMIは～ （○）
                        - 10-12月期のGDPは～（○）
            - **経済指標について:**
                    -"加速" の対象を明確に記載すること。文脈を考慮し、"加速" の対象を適切に補う。
                        文脈に応じて、"何が加速したのか" を判断し、適切な単語に修正してください。
                        修正ルール:

                        - 前月から加速しました（×） → 何が加速したのかを明記（○）
                        Exsample: 「前月から上昇が加速しました」 → 「景気回復の加速（○）」
                        - 前月から上昇が加速しました（×） → 具体的な経済活動を明記（○）
                        Exsample: 「景気加速（○）」「消費の回復が加速（○）」「投資の拡大が加速（○）」
                        - 経済（×） → 景気（○）（"経済" ではなく "景気" を用いる）
                        Exsample: 
                        前月から加速しました。（×）-> 企業の設備投資が加速しました。（○）
                        経済加速が見られます。（×）-> 景気加速が見られます。（○）
                        消費が前月から加速しました。（×）-> 個人消費の拡大が加速しました。（○）
                        インフレが前月から加速しました。（×）-> 物価上昇のスピードが加速しました。（○）


                    - 日付・国名の明記:
                        いつのものか特定できる場合、○月のものかを明記する。例: 「10月の製造業PMI（購買担当者景気指数）は～」
                        必要に応じて国名も記載する。（文脈に応じて記載していれば、位置は問わない）
                    - 日付・国名の変換ルール:

                        **「下旬」「上旬」**と記載されている場合、文脈から適切な日付に変更する。
                        **「ユーロ圏」**と記載されている場合、文脈に応じて適切な国名に置き換える。
                        Exsample:

                        修正前: 「下旬は、ユーロ圏総合PMI（購買担当者景気指数）が…」

                        修正後: 「10月下旬のドイツ総合PMI（購買担当者景気指数）が…」

                        修正前: 「上旬は、ユーロ圏総合PMI（購買担当者景気指数）が…」

                        修正後: 「10月上旬のフランス総合PMI（購買担当者景気指数）が…」

            - **業界指数の表記:**
                - 必ず対象となる「月」を明記する。
                - 月がない場合、最近3ヶ月以内か、3ヶ月以上前のものかを確認する。
                - 必要に応じて国名も記載する。（文脈に応じて位置は自由）
                    - Example:
                        - 製造業PMI（購買担当者景気指数）（×） → 10月の製造業PMI（○）
                        - 直近3ヶ月以内の指数は明示的に「○月」と記載する。（例:12月のCPI）
                        - 3ヶ月以上前の指数は、比較の文脈を明確にする。（例:2023年10月のGDP成長率）
                        - ユーロ圏の10月PMIは～（○）
            - **カタカナ表記の統一:**
                Katakana representation of foreign words should be unified within the document.
                    Ensure that the Katakana form is consistent throughout the text, and choose one version for the entire document.
                    Modification reason: To maintain consistency in the usage of Katakana for foreign words.
                    Example of text modifications:

                    サステナブル (×) → サスティナブル (○)

                    エンターテイメント (×) → エンターテインメント (○)
            
            - **レンジの表記について表記:**
                - Always append "%" when indicating a range.(Ensure that the original text is not directly modified but follows this guideline.)
                - Make modifications directly in this article and explain the reasons for the modifications.
                - Example: -1～0.5%（×） → -1%～0.5%（○）
            - **償還に関する記載:**
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                - 最終リリースの1ヵ月程前より、償還に関する内容を入れること。
                - 例）当ファンドは、●●月●●日に信託の終了日（償還日or繰上償還日）を迎える予定です。
                - ※（）内は、定時償還の場合には償還日、繰上償還の場合には繰上償還日とする。
            - **個別企業名の表記:**
                - 投資環境等においては、個別企業の名称を使わない表現を心掛ける。
                - 例:スイス金融大手クレディ・スイス（×） → スイスの大手金融グループ（○）
            - **プラスに寄与/影響の表記:**
                - Do not Make modifications directly in this article and explain the reasons for the modifications.
                - 「プラスに寄与」または「プラスに影響」どちらも可。
                - または、「～プラス要因となる」と表記。
            - **マイナスに影響の表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 「マイナスに寄与」（×）→「マイナスに影響」（○）
                - マイナスの際は「寄与」は使用しない。
                - または、「～マイナス要因となる」と表記。
            - **利回りの表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 利回りは「上昇（価格は下落）」または「低下（価格は上昇）」と表記。
            - **低下と下落の表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 債券利回りは「低下（○）」と表記し、「下落（×）」は使用しない。
                - 価格は「下落（○）」と表記し、「低下（×）」は使用しない。
                - 金利の「低下（〇）」と表記し、「下落（×）」は使用しない。

            - **資金流出入の表記:**
                - Do notMake modifications directly in this article and explain the reasons for the modifications.
                - 「外国人投資家の資金流出」を「外国人投資家からの資金流入」と記載。

            - **（金利の）先高感/先高観 の表記統一:**
                - 文中に「先高観」という表記がある場合でも、原文は修正しないでください。
                - その代わり、「先高観」の直後に「修正提案」として、「先高感」への統一理由を提示してください。
                - 表記がすでに「先高感」である場合は、何も追記せずそのままにしてください。

                - 表示フォーマット（例）:
                    先高観<span style="color:red;">（修正理由: 用語の統一 <s style="background:yellow;color:red">先高観</s> → 先高感）</span>

                - 必ず原文の構成と文脈を保持し、構文を壊さず、修正理由は補足的に後ろに追記してください。
                
            - **ポートフォリオの表記:**
                - Make modifications directly in this article and explain the reasons for the modifications.
                - 「●●への組み入れ（×）」ではなく、「●●の組み入れ（○）」と表記。
                - 「への投資比率」は使用可能。
                
            - **構成比の0％の表記:
                - 「0％程度」or「ゼロ％程度」の表記を使用すること
                - 変更前表記: 構成比は0％である
                - 統一後表記: 構成比は0％程度
                - Append a correction reason in the following format:
                        `<span style="color:red;">変更前表記</span> (<span>修正理由: 構成比表記 <s style="background:yellow;color:red">変更前表記</s> → 統一後表記</span>)`
                    
                    Example:
                    Input: 構成比は0％である
                    Output: 
                    <span style="color:red;">構成比は0％である</span> 
                    (<span>修正理由: 構成比表記 <s style="background:yellow;color:red">構成比は0％である</s> → 構成比は0％程度である。</span>)で売られています。

                
            - **'投資環境の記述:** 
                - Make modifications directly in this article and explain the reasons for the modifications.
                **「先月の投資環境」**の部分で「先月末」の記述が含まれる場合、「前月末」に変更してください。
                - Ensure that the original text is directly modified and follows this guideline.
                Example:
                修正前: 先月末の市場動向を分析すると…
                修正後: 前月末の市場動向を分析すると…

            - **通貨表記の統一:**
                - Standardize currency notation across the document.
                    - The first appearance of any currency symbol (e.g., ドル, $, 円, JPY) will be the standard.
                    - All following occurrences of that currency must match this format.

                    - For example, if "100ドル" appears first, then all future "$100" will be rewritten to "100ドル" for consistency.
                    - If "$100" appears first, then "100ドル" should be rewritten as "$100".

                    - Always apply this rule in the direction of "first-appeared" format.
                    - Append a correction reason in the following format:
                        `<span style="color:red;">統一後表記</span> (<span>修正理由: 通貨表記の統一 <s style="background:yellow;color:red">変更前表記</s> → 統一後表記</span>)`
                    
                    Example:
                    Input: このバッグは100ドルですが、アメリカでは$100で売られています。
                    Output:
                    このバッグは100ドルですが、アメリカでは
                    <span style="color:red;">100ドル</span>
                    (<span>修正理由: 通貨表記の統一 <s style="background:yellow;color:red">$100</s> → 100ドル</span>)で売られています。


            **Preferred and Recommended Terminology (置き換えが必要な用語/表現):**
            - **第1四半期:**
                - Ensure the period is clearly stated.
                - Example: 18年第4四半期（×） → 2018年10-12月期（○）
            - **約○％程度:**
                - Do not use "約" (approximately) and "程度" (extent) together. Choose either one.
                - Example: 約○％程度（×） → 約○％ or ○％程度（○）
            - **大手企業表記の明確化**  
                **Correction Rule:**
                - If a sentence contains vague expressions like「大手○○」, analyze the context to determine what type of company is being referred to.
                - Rewrite it in the format:「大手○○会社」/「大手○○企業」/「大手○○メーカー」depending on the company’s nature.
                - Use context clues (e.g., product type, industry references) to guess the appropriate company category (e.g., 不動産, 自動車, 電機, 金融).
                - Append a correction reason in this format:
                `<span style="color:red;">Changed Expression</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">Original Expression</s> → Changed Expression</span>)`

                **Example Input:**
                - 大手は業界全体に影響力を持つ。
                - 大手が新しい半導体を発表した。

                **Example Output:**
                - <span style="color:red;">大手不動産会社</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">大手</s> → 大手不動産会社</span>) は業界全体に影響力を持つ。
                - <span style="color:red;">大手半導体メーカー</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">大手</s> → 大手半導体メーカー</span>) が新しい半導体を発表した。

                **Important Notes:**
                - Always preserve the original sentence structure and paragraph formatting.
                - Only make corrections when「○○大手」is ambiguous and can be clarified using contextual information.
                - Do not modify proper nouns or known company names (e.g., トヨタ, ソニー).

            - **入力例:**  
                - 「大手メーカー/会社/企業」  
                - **出力:** 「大手不動産会社、大手半導体メーカー」  
            - **The actual company name must be found and converted in the article
            - **先月/前月の表記:
                - 1ヵ月前について言及する場合は、「前月」を使用。
            前期比○％の表記:

            - **前期比年率○％:**
                - 基本的に、期間比較の伸率は「年率」を記載してください。
                - 主に経済統計等で一般的に前期比で年率換算されているものについては、「前期比年率○％」と表記。
            - **第○四半期の表記:**
                **ルール:
                - If the input contains a format like "18年第4四半期", infer it as:
                    - "18年" → "2018年"
                    - "第1四半期" → "1-3月期"
                    - "第2四半期" → "4-6月期"
                    - "第3四半期" → "7-9月期"
                    - "第4四半期" → "10-12月期"
                - Modify the expression accordingly, converting the year to a 4-digit format and specifying the exact month range.
                - Add a correction reason in this format:
                `<span style="color:red;">修正後</span> (<span>修正理由: 四半期表記の明確化 <s style="background:yellow;color:red">修正前</s> → 修正後</span>)`

                ---

                **Example:**
                - Input: 18年第4四半期の売上が好調だった。
                - Output: 
                <span style="color:red;">2018年10-12月期</span> (<span>修正理由: 四半期表記の明確化 <s style="background:yellow;color:red">18年第4四半期</s> → 2018年10-12月期</span>) の売上が好調だった。

                ---

                **Additional Notes:**
                - Do not modify any proper names, organizations, or if the date range is already correct.
                - Apply to all similar shorthand expressions like "20年第2四半期", "21年第1四半期" etc.
                - Keep the structure and formatting of the original document.


        **Special Rules:**
        1. **Do not modify proper nouns (e.g., names of people, places, or organizations) unless they are clearly misspelled.**
            -Exsample:
            ベッセント氏: Since this is correct and not a misspelling, it will not be modified.
        2. **Remove unnecessary or redundant text instead of replacing it with other characters.**
            -Exsample:
            ユーロ圏域内の景気委: Only the redundant character 委 will be removed, and no additional characters like の will be added. The corrected text will be: ユーロ圏域内の景気.
        3. **Preserve spaces between words in the original text unless they are at the beginning or end of the text.**
            -Example:
            input: 月の 前半は米国の 債券利回りの上昇 につれて
            Output: 月の 前半は米国の 債券利回りの上昇 につれて (spaces between words are preserved).

        **Output Requirements:**
        1. **Highlight the original incorrect text in red and include additional details:**
        - For corrected parts:
            - Highlight the original incorrect text in red using `<span style="color:red;">`.
            - Append the corrected text in parentheses, marked with a strikethrough using `<s>` tags.
            - Provide the reason for the correction and indicate the change using the format `123 → 456`.
            - Example:
            `<span style="color:red;">123</span> (<span>修正理由: 一致性不足 <s style="background:yellow;color:red">123</s> → 456</span>)`
        
        2. **Preserve the original structure and formatting of the document:**
        - Maintain paragraph breaks, headings, and any existing structure in the content.

        3. **Use the uploaded correction rules for reference:**
        - {corrected}

        4. **Do not provide any explanations or descriptions in the output. Only return the corrected HTML content.**

         **Corrected Terminology Map (修正された用語リスト):
            {corrected_map}
        - Replace only when the **original** term in `corrected_map` appears in the input text.
        - Do **not** replace anything if the input already contains the `corrected` term (it is already correct).
        - Do **not** perform any reverse replacements (`corrected → original` は禁止).
        - Modify the original text only when the `original` term is found.

        - If the `corrected` term appears in the input, **do not modify it** (it is already correct).
        - Do **not** reverse substitutions (i.e., never convert corrected → `original`).
        
        - After replacing, add the reason in this format:
        Original Term (修正理由: 用語の統一 Original Term → Corrected Term)
        Example:
            `<span style="color:red;">Corrected Term</span> (<span>修正理由: 用語の統一 <s style="background:yellow;color:red">Original Term</s> → Corrected Term</span>)`
        
        Example:
        Input: 中銀
        Output: 
        `<span style="color:red;">中央銀行</span> (<span>修正理由: 用語の統一 <s style="background:yellow;color:red">中銀</s> → 中央銀行</span>)`
        ※ Note: Do **not** convert 中央銀行 → 中銀. All replacements must follow the direction from `original` to `corrected` only.

        Input: 中央銀行  
        Output:  
        中央銀行 ← (No correction shown because it is already the correct term)

        If the input already contains the corrected term, it should remain unchanged.
        For English abbreviations or foreign terms, the rule is the same: replace the original term with the corrected term and format as follows:
        Example:
        Input: BOE
        Output: <span style="color:red;">BOE（英中央銀行、イングランド銀行）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">BOE</s> → BOE（英中央銀行、イングランド銀行）</span>)
        Input: AAA
        Output: <span style="color:red;">AAA（全米自動車協会）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">AAA</s> → AAA（全米自動車協会）</span>)

        Input: インバウンド
        Output: <span style="color:red;">インバウンド（観光客の受け入れ）</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">インバウンド</s> → インバウンド（観光客の受け入れ）</span>)

        
        **Except Original Term
        Input: 等
        Output: 
        `<span style="color:red;">等</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">等</s> → など</span>)`

        Input: ローン
        Output: 
        `<span style="color:red;">ローン</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">ローン</s> → 貸し付け</span>)`
                            
        Input: ％を上回る
        Output: 
        `<span style="color:red;">％を上回る</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">％を上回る</s> → ％を超える</span>)`
                            
        Input: ％を下回る
        Output: 
        `<span style="color:red;">％を下回る</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">％を下回る</s> → ％を下回るマイナス幅</span>)`
        
        Input: 伝播（でんぱ）
        Output: 
        `<span style="color:red;">伝播（でんぱ）</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">伝播（でんぱ）</s> → 広ります</span>)`
        
        Input: 伝播しています
        Output:
        `<span style="color:red;">伝播しています</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">伝播しています</s> → 広がるしています</span>)`
        
        Input: 連れ高
        Output:
        `<span style="color:red;">連れ高</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">連れ高</s> → 影響を受けて上昇</span>)`
        
        Input: 相場
        Output:
        `<span style="color:red;">相場</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">相場</s> → 市場/価格</span>)`
        
        Input: ハト派
        Output:
        `<span style="color:red;">ハト派</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">ハト派</s> → 金融緩和重視、金融緩和に前向き</span>)`
        
        Input: タカ派
        Output:
        `<span style="color:red;">タカ派</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">タカ派</s> → 金融引き締め重視、金融引き締めに積極的</span>)`
        
        Input: 織り込む
        Output: 
        `<span style="color:red;">織り込む</span> (<span>修正理由: 用語の置き換え <s style="background:yellow;color:red">織り込む</s> → 反映され</span>)`
        
        Input: 積極姿勢とした
        Output: 
        `<span style="color:red;">積極姿勢とした</span> (<span>修正理由: 用語/表現 <s style="background:yellow;color:red">積極姿勢とした</s> → 長めとした</span>)`
        
        Input: 限定的
        Output: 
        `<span style="color:red;">限定的</span> (<span>修正理由: 効果や影響がプラスかマイナスか不明瞭なため <s style="background:yellow;color:red">限定的</s> → 他の適切な表現に修正</span>)`
        
        Input: 利益確定の売り
        Output: 
        `<span style="color:red;">利益確定の売り</span> (<span>修正理由: 断定的な表現では根拠が説明できないため <s style="background:yellow;color:red">利益確定の売り</s> → が出たとの見方</span>)`
        
        Input: 利食い売り
        Output: 
        `<span style="color:red;">利食い売り</span> (<span>修正理由: 断定的な表現では根拠が説明できないため <s style="background:yellow;color:red">利食い売り</s> → が出たとの見方</span>)`
        
        Input: ABS
        Output: 
        `<span style="color:red;">ABS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">ABS</s> → ABS（資産担保証券、各種資産担保証券）</span>)`
        
        Input: AI
        Output: 
        `<span style="color:red;">AI</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">AI</s> → AI（人工知能</span>)`
        
        Input: BRICS（5ヵ国）
        Output: 
        `<span style="color:red;">BRICS（5ヵ国）</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">BRICS（5ヵ国）</s> → BRICS（ブラジル、ロシア、インド、中国、南アフリカ）</span>)`
        
        Input: CMBS
        Output: 
        `<span style="color:red;">CMBS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">CMBS</s> → CMBS（商業用不動産ローン担保証券）</span>)`
        
        Input: ISM
        Output: 
        `<span style="color:red;">ISM</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">ISM</s> → ISM（全米供給管理協会）</span>)`
        
        Input: IT
        Output: 
        `<span style="color:red;">IT</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">IT</s> → IT（情報技術）</span>)`
        
        Input: MBS
        Output: 
        `<span style="color:red;">MBS</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">MBS</s> → MBS（住宅ローン担保証券）</span>)`
        
        Input: PMI
        Output: 
        `<span style="color:red;">PMI</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">PMI</s> → PMI（購買担当者景気指数）</span>)`
        
        Input: S&P
        Output: 
        `<span style="color:red;">S&P</span> (<span>修正理由: 英略語 <s style="background:yellow;color:red">S&P</s> → S&P（スタンダード・アンド・プアーズ）社</span>)`
        
        Input: アセットアロケーション
        Output: 
        `<span style="color:red;">アセットアロケーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">アセットアロケーション</s> → アセットアロケーション（資産配分）</span>)`
        
        Input: E-コマース
        Output: 
        `<span style="color:red;">Eコマース</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">E-コマース</s> → Eコマース（電子商取引）</span>)`
             
        Input: e-コマース
        Output: 
        `<span style="color:red;">eコマース</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">eコマース</s> → eコマース（電子商取引）</span>)`
           
        Input: EC
        Output: 
        `<span style="color:red;">EC（電子商取引）</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">EC</s> → EC（電子商取引）</span>)`
        
        Input: イールドカーブ
        Output: 
        `<span style="color:red;">イールドカーブ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">イールドカーブ</s> → イールドカーブ（利回り曲線）</span>)`
        
        Input: エクスポージャー
        Output: 
        `<span style="color:red;">エクスポージャー</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">エクスポージャー</s> → ＊積極的に使用しない。　（価格変動リスク資産の配分比率、割合）</span>)`
        
        Input: クレジット（信用）市場
        Output: 
        `<span style="color:red;">クレジット（信用）市場</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">クレジット（信用）市場</s> → 信用リスク（資金の借り手の信用度が変化するリスク）を内包する商品（クレジット商品）を取引する市場の総称。　企業の信用リスクを取引する市場。</span>)`
        
        Input: システミック・リスク
        Output: 
        `<span style="color:red;">システミック・リスク</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">システミック・リスク</s> → 個別の金融機関の支払不能等や、特定の市場または決済システム等の機能不全が、他の金融機関、他の市場、または金融システム全体に波及するリスク</span>)`
        
        Input: ディストレス債券
        Output: 
        `<span style="color:red;">ディストレス債券</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ディストレス債券</s> → 信用事由などにより、価格が著しく下落した債券</span>)`
        
        Input: ディフェンシブ
        Output: 
        `<span style="color:red;">ディフェンシブ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ディフェンシブ</s> → ディフェンシブ（景気に左右されにくい）</span>)`
        
        Input: テクニカル
        Output: 
        `<span style="color:red;">テクニカル</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">テクニカル</s> → テクニカル（過去の株価の動きから判断すること</span>)`
        
        Input: デフォルト
        Output: 
        `<span style="color:red;">デフォルト</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デフォルト</s> → デフォルト（債務不履行）</span>)`
        
        Input: デフォルト債
        Output: 
        `<span style="color:red;">デフォルト債</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デフォルト債</s> → デフォルトとは一般的には債券の利払いおよび元本返済の不履行、もしくは遅延などをいい、このような状態にある債券をデフォルト債といいます。</span>)`
        
        Input: デュレーション
        Output: 
        `<span style="color:red;">デュレーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">デュレーション</s> → デュレーション（金利感応度）</span>)`
        
        Input: 投資適格債
        Output: 
        `<span style="color:red;">投資適格債</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">投資適格債</s> → 格付機関によって格付けされた公社債のうち、債務を履行する能力が十分にあると評価された公社債</span>)`
        
        Input: ファンダメンタルズ
        Output: 
        `<span style="color:red;">ファンダメンタルズ</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ファンダメンタルズ</s> → ファンダメンタルズ（賃料や空室率、需給関係などの基礎的条件）※REITファンドで使用する</span>)`
        
        Input: フリーキャッシュフロー
        Output: 
        `<span style="color:red;">フリーキャッシュフロー</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">フリーキャッシュフロー</s> → 税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたもの</span>)`
        
        Input: ベージュブック
        Output: 
        `<span style="color:red;">ベージュブック</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">ベージュブック</s> → スペースがない場合は、ベージュブック（米地区連銀経済報告）</span>)`
        
        Input: モメンタム
        Output: 
        `<span style="color:red;">モメンタム</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">モメンタム</s> →  相場の勢い)が強く、投資家たちは短期的な利益を狙っています。</span>)`
        
        Input: リオープン
        Output: 
        `<span style="color:red;">リオープン</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リオープン</s> → リオープン/リオープニング（経済活動再開）</span>)`
        
        Input: リスクプレミアム
        Output: 
        `<span style="color:red;">リスクプレミアム</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リスクプレミアム</s> → 同じ投資期間内において、あるリスク資産の期待収益率が、無リスク資産（国債など）の収益率を上回る幅のこと。</span>)`
        
        Input: リフレーション
        Output: 
        `<span style="color:red;">リフレーション</span> (<span>修正理由: 外来語・専門用語 <s style="background:yellow;color:red">リフレーション</s> → リフレーション**デフレーションから抜けて、まだ、インフレーションにはなっていない状況のこと。</span>)`
        
        **【例外用語 – 修正しないこと】**
        - コロナ禍
        - コロナショック
        - 新型コロナ禍
        - 住宅ローン
        - 引き締め策
        - 引き締め政策
        - 組入比率
        - 格付機関
        - 格付別
        - 国債買入オペ

        **特定表現の言い換えルール（文脈判断を伴う修正）:
        文脈に応じて、具体的な表現に言い換えてください。
        「まちまち」の使用
        「まちまち」という曖昧な表現が出現した場合は、その語をそのまま保持した上で、後に「修正理由: 曖昧表現の明確化」を補足してください。

        変換語「異なる動き」も表示しますが、原文は変更せず、装飾で示すのみです。

        Output Format (Original term preserved, only correction reason shown):
        <span style="color:red;">まちまち</span>
        (<span>修正理由: 曖昧表現の明確化 <s style="background:yellow;color:red">まちまち</s> → 異なる動き</span>)

        -「行って来い」の表現

        文脈に応じて、「上昇（下落）したのち下落（上昇）」のように明確にしてください。
        Exsample:

        Input: 相場は行って来いの展開となった
        Output: 相場は上昇したのち下落する展開となった
        
        変換語「行って来い」も表示しますが、原文は変更せず、装飾で示すのみです。

        Output Format (Original term preserved, only correction reason shown):
        
        修正理由: 表現の明確化
        <span style="color:red;">行って来い</span> (<span>修正理由: 表現の明確化 <s style="background:yellow;color:red">Original Term</s> → 行って来い</span>)

        -「横ばい」表現の適正使用

        小幅な変動であれば「横ばい」を使用可能。
        大きな変動の末に同水準で終了した場合は、「ほぼ変わらず」「同程度となる」などに修正。
        
        変換語「横ばい」も表示しますが、原文は変更せず、装飾で示すのみです。
        Output Format (Original term preserved, only correction reason shown):

        修正理由: 用語の適正使用
        <span style="color:red;">横ばい</span> (<span>修正理由: 用語の適正使用 <s style="background:yellow;color:red">横ばい</s> → ほぼ変わらず</span>)

        -「（割安に）放置」表現の修正

        「割安感のある」など、より適切な表現に修正してください。

        Exsample:
        Input: 株価は割安に放置された
        Output: 株価には割安感がある状態が続いた

        変換語「（割安に）放置」も表示しますが、原文は変更せず、装飾で示すのみです。
        Output Format (Original term preserved, only correction reason shown):

        修正理由: 表現の明確化と客観性の向上
        <span style="color:red;">（割安に）放置</span> (<span>修正理由: 表現の明確化と客観性の向上 <s style="background:yellow;color:red"（割安に）放置</s> → 割安感のある</span>)

        """  
        # ChatCompletion Call
        response = openai.ChatCompletion.create(
        # OpenAI API 调用 asyncio
        # loop = asyncio.get_event_loop()
        # response = await loop.run_in_executor(None, lambda: openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=[
                {"role": "system", "content": "You are a professional Japanese text proofreading assistant."
                "This includes not only Japanese text but also English abbreviations (英略語), "
                "foreign terms (外来語),and specialized terminology (専門用語)."},
                {"role": "user", "content": prompt_result}
            ],
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
            seed=SEED
        )
        answer = response['choices'][0]['message']['content'].strip()
        re_answer = remove_code_blocks(answer)
        
        return jsonify({"success": True, "corrected_text": re_answer})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def convert_logs(items):
    converted_data = {
        "code": 200,
        "message": "成功",
        "data": []
    }
    
    for idx, item in enumerate(items):
        log_entries = item.get("logEntries", [])
        
        for log_idx, log_entry in enumerate(log_entries):
            log_parts = log_entry.split(" - ")
            timestamp_str = log_parts[0] if len(log_parts) > 1 else ""
            message = log_parts[1] if len(log_parts) > 1 else ""
            
            log_data = {
                "id": idx * len(log_entries) + log_idx + 1,  # ID
                "name": message,  # message[:30] message split pre 30 'name'
                "status": "完了(修正あり)" if "✅ SUCCESS" in message else "エラー",
                "timeclock": timestamp_str,
                "progress": "成功" if "✅ SUCCESS" in message else "エラー",
                "timestamp": timestamp_str,
                "selected": False
            }
            
            converted_data["data"].append(log_data)
    
    return converted_data

# appLog
APPLOG_CONTAINER_NAME='appLog'
@app.route('/api/applog', methods=['GET'])
def get_applog():
    # Cosmos DB 连接
    container = get_db_connection(APPLOG_CONTAINER_NAME)

    query = "SELECT * FROM c"
    items = list(container.query_items(query=query, enable_cross_partition_query=True))

    for item in items:
        item['id'] = item['id']

    converted_logs = convert_logs(items)
    return jsonify(converted_logs)

# azure Cosmos DB
@app.route('/api/faqs', methods=['GET'])
def get_faq():
    # Cosmos DB 链接客户端,ENV
    container=get_db_connection()

    query = "SELECT * FROM c"
    items = list(container.query_items(query=query, enable_cross_partition_query=True))

    for item in items:
        item['id'] = item['id']

    return jsonify(items)

@app.route('/api/tenbrend', methods=['POST'])
def tenbrend():
    data = request.get_json() or {}

    raw_fcode = data.get('fcode', '').strip()
    months = data.get('month', '').strip()
    stocks = data.get('stock', '').strip()
    fund_type = data.get('fundType', 'public').strip()  # 默认为公募

    # 根据 fundType 选择容器（即 Cosmos DB 的表）

    if fund_type == 'private':
        TENBREND_CONTAINER_NAME = 'tenbrend_private'
    else:
        TENBREND_CONTAINER_NAME = 'tenbrend'

    # Cosmos DB 链接客户端
    container = get_db_connection(TENBREND_CONTAINER_NAME)
    parameters = []
    if not raw_fcode:
        query = "SELECT * FROM c"

    else:
        # 构建 SQL 查询
        if '-' in raw_fcode:
            # 带 `-` 的直接用字符串模糊匹配
            query = "SELECT * FROM c WHERE CONTAINS(c.fcode, @fcode)"
            parameters.append({"name": "@fcode", "value": raw_fcode})
        else:
            try:
                fcode_num = raw_fcode
                query = "SELECT * FROM c WHERE c.fcode = @fcode"
                parameters.append({"name": "@fcode", "value": fcode_num})
            except ValueError:
                # fallback 到字符串查询
                query = "SELECT * FROM c WHERE CONTAINS(c.fcode, @fcode)"
                parameters.append({"name": "@fcode", "value": raw_fcode})

        if months:
            query += " AND c.months = @months"
            parameters.append({"name": "@months", "value": months})

        if stocks:
            query += " AND CONTAINS(c.stocks, @stocks)"
            parameters.append({"name": "@stocks", "value": stocks})

    items = list(container.query_items(
        query=query,
        parameters=parameters,
        enable_cross_partition_query=True
    ))

    filtered_items = [item for item in items if item.get('id')]
    return jsonify({"code": 200, "data": filtered_items})


@app.route('/api/tenbrend/months', methods=['POST'])
def tenbrend_months():
    data = request.get_json() or {}

    fcode = data.get('fcode', '').strip()
    stocks = data.get('stock', '').strip() if data.get('stock') else ''
    fund_type = data.get('fundType', 'public').strip()

    if not fcode:
        return jsonify({"code": 400, "message": "fcode is required"}), 400

    # ✅ 根据 fundType 切换容器（表）
    if fund_type == 'private':
        TENBREND_CONTAINER_NAME ='tenbrend_private'
    else:
        TENBREND_CONTAINER_NAME='tenbrend'

    # Cosmos DB 链接客户端
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    query = "SELECT c.months FROM c WHERE CONTAINS(c.fcode, @fcode)"
    parameters = [{"name": "@fcode", "value": fcode}]

    if stocks:
        query += " AND CONTAINS(c.stocks, @stocks)"
        parameters.append({"name": "@stocks", "value": stocks})

    try:
        items = list(container.query_items(
            query=query,
            parameters=parameters,
            enable_cross_partition_query=True
        ))

        months = sorted({item.get('months') for item in items if item.get('months')})
        return jsonify({"code": 200, "data": months})
    except Exception as e:
        print("❌ Cosmos DB query failed:", e)
        return jsonify({"code": 500, "message": "internal error"}), 500


@app.route('/api/tenbrend/stocks', methods=['POST'])
def tenbrend_stocks():
    data = request.get_json() or {}

    fcode = data.get('fcode', '').strip()
    months = data.get('month', '').strip() if data.get('month') else ''
    fund_type = data.get('fundType', 'public').strip()

    if not fcode:
        return jsonify({"code": 400, "message": "fcode is required"}), 400

    if fund_type == 'private':
        TENBREND_CONTAINER_NAME ='tenbrend_private'
    else:
        TENBREND_CONTAINER_NAME='tenbrend'

    # Cosmos DB 链接客户端
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    query = "SELECT c.stocks FROM c WHERE CONTAINS(c.fcode, @fcode)"
    parameters = [{"name": "@fcode", "value": fcode}]

    if months:
        query += " AND c.months = @months"
        parameters.append({"name": "@months", "value": months})

    try:
        items = list(container.query_items(
            query=query,
            parameters=parameters,
            enable_cross_partition_query=True
        ))

        stocks = sorted({item.get('stocks') for item in items if item.get('stocks')})
        return jsonify({"code": 200, "data": stocks})
    except Exception as e:
        print("❌ Cosmos DB query failed:", e)
        return jsonify({"code": 500, "message": "internal error"}), 500




@app.route('/api/tenbrend/template', methods=['GET'])
def download_excel_template():
    data = request.get_json() or {}
    # 默认是“公募”
    fund_type = data.get('fundType', 'public').strip()

    # 根据类型拼接路径
    if fund_type == '私募':
        file_url = ACCOUNT_URL + STORAGE_CONTAINER_NAME +"/10銘柄マスタ管理_私募.xlsx"
    else:
        file_url = ACCOUNT_URL + STORAGE_CONTAINER_NAME +"/10銘柄マスタ管理_公募.xlsx"

    try:
        # 注意:send_file 不能直接下载远程链接，改为重定向
        return redirect(file_url)
    except Exception as e:
        return jsonify({"code": 500, "message": str(e)}), 500



# Data transfer
def transform_data(items,fund_type):
    menu_data = {
        "公募": [],
        "私募": []
    }

    for item in items:
        if fund_type == 'public':
            fund_category = menu_data["公募"]
        elif fund_type == 'private':
            fund_category = menu_data["私募"]
        else:
            continue  # 잘못된 fund_type은 무시

        # 데이터 구조에 맞게 변환
        reference = {
            "id": "reference",
            "name": "📁 参照ファイル",
            "children": [
                {
                    "id": "report_data",
                    "name": "📂 レポートデータ",
                    "children": []
                },
                {
                    "id": "mingbing_data",
                    "name": "📂 10銘柄解説一覧表",
                    "children": []
                }
            ]
        }

        # report_data里添加item
        reference["children"][0]["children"].append({
            "id": item.get('id'),
            "name": item.get('fileName'),
            "icon": "📄",
            "file": item.get('fileName'),
            "pdfPath": extract_pdf_path(item.get('link')),
        })

        # mingbing_data里添加item
        reference["children"][1]["children"].append({
            "id": item.get('id'),
            "name": item.get('fileName'),
            "icon": "📄",
            "file": item.get('fileName'),
            "pdfPath": extract_pdf_path(item.get('link'))
        })

        fund_category.append(reference)

        # checked_files 追加session
        checked_files = {
            "id": "checked_files",
            "name": "📁 チェック対象ファイル",
            "children": [
                {
                    "id": "individual_comments",
                    "name": "📂 共通コメントファイル",
                    "children": []
                },
                {
                    "id": "kobetsucomment",
                    "name": "📂 個別コメントファイル",
                    "children": []
                }
            ]
        }

        # individual_comments里添加item
        checked_files["children"][0]["children"].append({
            "id": item.get('id'),  
            "name": item.get('fileName'),  
            "icon": "⚠️",
            "file": item.get('fileName'),  
            "status": item.get('comment_status'),  
            "readStatus": item.get('comment_readStatus'),  
            "pdfPath": extract_pdf_path(item.get('link'))  
        })

        # kobetsucomment里添加item
        checked_files["children"][1]["children"].append({
            "id": item.get('id'),  
            "name": item.get('fileName'),
            "icon": "❌",
            "file": item.get('fileName'),
            "status": item.get('individual_status'),
            "readStatus": item.get('individual_readStatus'),
            "pdfPath": extract_pdf_path(item.get('link'))  
        })

        fund_category.append(checked_files)

    return menu_data

def extract_pdf_path(link):
    match = re.search(r'href="([^"]+)"', link)
    return match.group(1) if match else ""

def extract_base_name(file_path):
    file_name = os.path.basename(file_path)
    base_name, _ = os.path.splitext(file_name)
    return base_name

# public_Fund and private_Fund
@app.route('/api/fund', methods=['POST'])
def handle_fund():
    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 连接
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        query = "SELECT * FROM c"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        
        # filtered_items = [item for item in items if item and item.get('id')]
        
        # return jsonify(filtered_items)
        formatted_data = transform_data(items,fund_type)

        return jsonify(formatted_data)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# 625 tenbrend
def convert_to_tenbrend(items):
    corrections = []

    for item in items:
        old_text = item.get("元組入銘柄解説", "").strip()
        new_text = item.get("新組入銘柄解説", "").strip()

        if old_text != new_text:
            corrections.append({
                "check_point": "組入銘柄解説",
                "comment": f"{old_text} → {new_text}",
                "intgr": False,
                "locations": [],
                "original_text": new_text,
                "page": '',
                "reason_type": item.get("stocks", "")
            })

    return corrections


# 509 debug
def convert_format(filtered_items):
    checkResults = {}

    for correction in filtered_items.get("result", {}).get("corrections", []):
        page = correction["page"] + 1
        position = {}
        colorSet = "rgb(172 228 230)"

        change = {
            "before": correction["original_text"],
            "after": correction["comment"].split("→")[-1].strip(),
        }
        if correction["intgr"]:
            name = "不一致"
            colorSet = "rgba(172, 228, 230, 0.5)"
        else:
            name = ""
            colorSet= "rgba(255, 255, 0, 0.5)"

        if correction["locations"]:
            # for idx, loc in enumerate(correction["locations"]): 
            # checkResults
            if page not in checkResults:
                checkResults[page] = [{"title": filtered_items["fileName"], "items": []}]

            # loc = correction["locations"][0]
            for loc in correction["locations"]:
                pdf_height = loc.get("pdf_height", 792)  # PDF height (Default: A4 , 792pt)

                # x = loc["x0"] - 22 if idx == 0 else loc["x0"]
                position = {
                    "x": loc["x0"],
                    "y": pdf_height - loc["y1"] + 50,
                    "width": loc["x1"] - loc["x0"],
                    "height": loc["y1"] - loc["y0"],
                }

                if correction["intgr"]:
                    checkResults[page][0]["items"].append({
                        "name": name,
                        "color": colorSet, #"rgba(255, 255, 0, 0.5)", # green background rgba(0, 255, 0, 0.5)
                        "page": page,
                        "position": position,
                        "changes": [change],
                        "reason_type":correction["reason_type"],
                        "check_point":correction["check_point"],
                        "original_text":correction["original_text"],
                        })
                else:
                        existing_item = any(
                                item["name"] == name and
                                item["changes"] == [change] and
                                item["position"] == position
                                for item in checkResults[page][0]["items"]
                            )
                        if not existing_item:
                            checkResults[page][0]["items"].append({
                                "name": name,
                                "color": colorSet, #"rgba(255, 255, 0, 0.5)", # green background rgba(0, 255, 0, 0.5)
                                "page": page,
                                "position": position,
                                "changes": [change],
                                "reason_type":correction["reason_type"],
                                "check_point":correction["check_point"],
                                "original_text":correction["original_text"],
                                })


    return {'data': checkResults, 'code': 200}

# public_Fund and check-results
@app.route('/api/check_results', methods=['POST'])
def handle_check_results():
    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400
    
    selected_id = request.json.get('selectedId')
    if not selected_id:
        return jsonify({"error": "selectedId is required"}), 400
    
    pageNumber = request.json.get('pageNumber')
    if not pageNumber:
        return jsonify({"error": "pageNumber is required"}), 400

    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 连接
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        query = "SELECT * FROM c WHERE c.id = @id"
        parameters = [{"name": "@id", "value": selected_id}]
        items = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))
        
        if not items:
            return jsonify({"error": "Item not found"}), 404

        converted_data = convert_format(items[0])

        return jsonify(converted_data)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# get side bar
@app.route('/api/menu', methods=['POST'])
def handle_menu():
    fund_type = request.json.get('type')
    page = int(request.json.get('page', 1))
    page_size = int(request.json.get('page_size', 10))
    # user_name = request.json.get('user_name')

    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    # container name Setting
    container_name = f"{fund_type}_Fund"

    
    try:
        # Cosmos DB 连接
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        # Query exe
        query = "SELECT * FROM c WHERE CONTAINS(c.id, '.pdf') OR c.upload_type='参照ファイル'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        
        # filter result
        filtered_items = [item for item in items if item and item.get('id')]

        # pagenations
        total = len(filtered_items)
        start = (page - 1) * page_size
        end = start + page_size
        paged_items = filtered_items[start:end]

        response = {
            "code": 200,
            "data": paged_items,
            "total": total
        }

        return jsonify(response)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/menu_all', methods=['POST'])
def handle_menu_all():
    # param check
    fund_type = request.json.get('type')

    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    # container name Setting
    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 连接
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        # Query exe
        query = "SELECT * FROM c WHERE CONTAINS(c.id, '.pdf') OR c.upload_type='参照ファイル'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        
        # filter result
        filtered_items = [item for item in items if item and item.get('id')]
        response = {
        "code": 200,
        "data": filtered_items
        }

        return jsonify(response)
        
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
# Cosmos DB 状态确认 endpoint
MONITORING_CONTAINER_NAME = "monitoring-status"

# Cosmos DB 状态确认 endpoint
@app.route('/api/monitoring-status', methods=['GET'])
def get_monitoring_status():
    try:
        # Cosmos DB 连接
        container = get_db_connection(MONITORING_CONTAINER_NAME)
        
        # Cosmos DB里取数据
        query = "SELECT * FROM c"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        for item in items:
            item['id'] = item['id']

        return jsonify(items), 200
        
    except CosmosResourceNotFoundError:
        logging.error("Monitoring status document not found")
        return jsonify({"error": "Status document not found"}), 404
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# 状态更新
@app.route('/api/monitoring-status', methods=['PUT'])
def update_monitoring_status():
    try:
        # Cosmos DB 连接
        container = get_db_connection(MONITORING_CONTAINER_NAME)
        
        new_status = request.json.get('status', 'off')
        
        status_item = {
            'id': 'monitoring_status',
            'type': 'control',
            'status': new_status,
            "timestamp": datetime.utcnow().isoformat()
        }
        
        container.upsert_item(body=status_item)
        logging.info(f"Monitoring status updated to {new_status}")
        return jsonify({'message': 'Status updated', 'new_status': new_status, 'code': 200}), 200
        
    except CosmosResourceNotFoundError as e:
        logging.error(f"Cosmos DB error: {str(e)}")
        return jsonify({"error": "Database operation failed"}), 500
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# read/unread status change
# Cosmos DB 状态确认 endpoint
@app.route('/api/update_read_status', methods=['POST'])
def get_read_status():
    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400
    
    selected_id = request.json.get('selectedId')
    if not selected_id:
        return jsonify({"error": "selectedId is required"}), 400

    # container name Setting
    container_name = f"{fund_type}_Fund"
    try:
        # Cosmos DB 连接
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        query = "SELECT * FROM c WHERE c.id = @id"
        parameters = [{"name": "@id", "value": selected_id}]
        items = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))
        
        if not items:
            return jsonify({"error": "Item not found"}), 404

        return jsonify(items[0]), 200
        
    except CosmosResourceNotFoundError:
        logging.error("read status document not found")
        return jsonify({"error": "read Status document not found"}), 404
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/api/update_read_status', methods=['PUT'])
def update_read_status():
    selected_id = request.json.get('selectedId')
    if not selected_id:
        return jsonify({"error": "selectedId is required"}), 400

    mark = request.json.get('mark')
    if mark not in ['read', 'unread']:
        return jsonify({"error": "Invalid mark value"}), 400

    fund_type = request.json.get('type')
    if fund_type not in ['public', 'private']:
        return jsonify({"error": "Invalid fund type"}), 400

    # container name Setting
    container_name = f"{fund_type}_Fund"
    
    try:
        # Cosmos DB 连接
        container = get_db_connection(container_name)
        logging.info(f"Connected to {container_name} container")
        
        query = "SELECT * FROM c WHERE c.id = @id"
        parameters = [{"name": "@id", "value": selected_id}]
        items = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))

        if not items:
            return jsonify({"error": "Item not found"}), 404

        status_item = items[0]

        # readStatus 和 timestamp
        status_item['readStatus'] = mark
        status_item['timestamp'] = datetime.utcnow().isoformat()
        
        container.upsert_item(body=status_item)
        logging.info(f"readStatus updated to {mark} for item {selected_id}")
        return jsonify({'message': 'Status updated', 'new_status': mark, 'code': 200}), 200
        
    except CosmosResourceNotFoundError as e:
        logging.error(f"Cosmos DB error: {str(e)}")
        return jsonify({"error": "Item not found"}), 404
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500


@app.route("/api/health")
def health_check():
    return "OK", 200

logging.basicConfig(level=logging.INFO)

def get_storage_container():
    """
    Azure AD RBAC 方式 Azure Blob Storage에 连接, 返回ContainerClient .
    :return: ContainerClient
    """
    try:
        # BlobServiceClient 
        blob_service_client = BlobServiceClient(account_url=ACCOUNT_URL, credential=credential)
        
        container_client = blob_service_client.get_container_client(STORAGE_CONTAINER_NAME)
        
        print("Connected to Azure Blob Storage via Azure AD RBAC")
        logging.info("Connected to Azure Blob Storage via Azure AD RBAC")
        
        return container_client
    except Exception as e:
        logging.error(f"Azure Blob Storage Connection Error: {e}")
        print(f"Azure Blob Storage Connection Error: {e}")
        raise e
    
def allowed_file(filename):
    """    
    :param filename:
    :return: bool
    """
    ALLOWED_EXTENSIONS = {'pdf', 'xlsx','txt','xls','XLSX','xlm','xlsm','xltx','xltm','xlsb','doc','docx'}   # PDF 和 Excel  
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/test_token', methods=['GET'])
def test_token():
    try:
        token_cache._refresh_token()
        token = token_cache.get_token()
        return jsonify({"access_token": token}), 200
    except Exception as e:
        logging.exception("Token Get Error")
        return jsonify({"message": f"Token Get Error: {str(e)}"}), 500

# uploadpdf,api/brand
def parse_escaped_json(raw_text: str):
    text = raw_text.strip()
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
    
    text = text.replace('```json', '')
    text = text.replace('```', '')

    text = text.replace('""', '"')

    parsed = json.loads(text)
    return parsed

def parse_gpt_response(answer):
    try:
        json_str = re.search(r'\{[\s\S]*?\}', answer).group()
        return json.loads(json_str)
    except (AttributeError, json.JSONDecodeError):
        dict_str = re.search(r'corrected_map\s*=\s*\{[\s\S]*?\}', answer, re.DOTALL)
        if dict_str:
            dict_str = dict_str.group().split('=', 1)[1].strip()
            return ast.literal_eval(dict_str)
        return {}

def detect_corrections(original, corrected):
    matcher = SequenceMatcher(None, original, corrected)
    corrections = {}
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'replace':
            orig_part = original[i1:i2].strip()
            corr_part = corrected[j1:j2].strip()
            if orig_part and corr_part:
                corrections[orig_part] = corr_part
    return corrections

def filter_corrected_map(corrected_map):
    keys_to_remove = [" ", "  "]
    for key in keys_to_remove:
        if key in corrected_map:
            del corrected_map[key]
    return corrected_map

# 512 debug
def apply_corrections(input_text, corrected_map):
    result = input_text


    for original, corrected in corrected_map.items():

        if result == corrected:
            continue

        if re.search(re.escape(corrected), result):
            continue

        pattern_already_corrected = re.compile(
            rf"<span style=\"color:red;\">{re.escape(corrected)}</span>\s*"
            rf"\(<span>修正理由: 用語の統一\s*<s style=\"background:yellow;color:red\">{re.escape(original)}</s>\s*→\s*{re.escape(corrected)}</span>\)"
        )
        if pattern_already_corrected.search(result):
            continue

        # original
        if re.search(original, result):
            replacement = (
                f'<span style="color:red;">{corrected}</span> '
                f'(<span>修正理由: 用語の統一 '
                f'<s style="background:yellow;color:red">{original}</s> → {corrected}</span>)'
            )
            result = result.replace(original, replacement)

    return result


DICTIONARY_CONTAINER_NAME = "dictionary"
def fetch_and_convert_to_dict():
    try:
        container = get_db_connection(DICTIONARY_CONTAINER_NAME)
        query = "SELECT c.original, c.corrected FROM c"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        corrected_dict = {item["original"]: item["corrected"] for item in items if "original" in item and "corrected" in item}

        return corrected_dict

    except CosmosHttpResponseError as e:
        print(f"❌ DB error: {e}")
        return {}
    
@app.route('/api/check_upload', methods=['POST'])
def check_upload():
    if 'files' not in request.files:
        return jsonify({"success": False, "message": "No files part"}), 400

    files = request.files.getlist('files')
    file_type = request.form.get("fileType")
    fund_type = request.form.get("fundType")

    for file in files:
        if file.filename == '':
            return jsonify({"success": False, "error": "No selected file"}), 400

        file_bytes = file.read()

        if file and allowed_file(file.filename):
            try:
                if file.filename.endswith('.pdf'):  
                    tenbrend_data = check_tenbrend(file.filename,fund_type)
                    reader = PdfReader(io.BytesIO(file_bytes))
                    text = ""
                    for page in reader.pages:
                        text += page.extract_text()

                    # Encode the PDF bytes to Base64
                    file_base64 = base64.b64encode(file_bytes).decode('utf-8')

                    return jsonify({
                        "success": True,
                        "original_text": extract_text_from_base64_pdf(file_bytes),  # file_bytesPDF text  , input = extract_text_from_base64_pdf(pdf_base64)
                        "pdf_bytes": file_base64,  # PDF Base64 
                        "file_name": file.filename,
                        "tenbrend_data":tenbrend_data,
                        # "fund_type": fund_type
                    })
                
                elif file.filename.endswith('.txt'):
                    text = file_bytes.decode('utf-8')  # UTF-8 

                    return jsonify({
                        "success": True,
                        "prompt_text": text
                    })

                elif file.filename.endswith(('.doc', '.docx')):
                    # Just Only DOCX format
                    # docx = Document(io.BytesIO(file_bytes))
                    # text = "\n".join([para.text for para in docx.paragraphs])

                    file_base64 = base64.b64encode(file_bytes).decode('utf-8')

                    return jsonify({
                        "success": True,
                        "original_text": "",
                        "docx_bytes": file_base64,
                        "file_name": file.filename
                    })

                elif regcheck.search(r'\.(xls|xlsx|XLSX|xlsm|xlm|xltx|xltm|xlsb)$',file.filename):
                    """
                    :param file_bytes: 上传的base64文件
                    :return: 修改完的base64 encoding
                    """
                    #--------------excel start------------------------------------------
                    # 🔹 1️⃣ corrected_map init
                    # corrected_map = fetch_and_convert_to_dict()
                    # all_text=[]

                    # # 🔹 2️⃣ 临时保存内存里 in-memory zip)
                    # in_memory_zip = zipfile.ZipFile(io.BytesIO(file_bytes), 'r')

                    # # new ZIP 的 BytesIO
                    # output_buffer = io.BytesIO()
                    # new_zip = zipfile.ZipFile(output_buffer, 'w', zipfile.ZIP_DEFLATED, allowZip64=True)

                    # # 🔹 3️⃣ 循环文件             
                    # for item in in_memory_zip.infolist():
                    #     file_data = in_memory_zip.read(item.filename)
                    #     # 🔹 4️⃣ 是否drawingN.xml 检查 (处理文本框)
                    #     if item.filename.startswith("xl/drawings/drawing") and item.filename.endswith(".xml"):
                    #         try:
                    #             tree = ET.fromstring(file_data)
                    #             ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
                    #             # 所有的 <a:t> 
                    #             text_elements = []
                    #             for t_element in tree.findall(".//a:t", ns):
                    #                 original_text = t_element.text
                    #                 if original_text:
                    #                     parent = t_element.getparent()
                    #                     if parent is not None:
                    #                         x = parent.attrib.get('x', 0)
                    #                         y = parent.attrib.get('y', 0)
                    #                         text_elements.append((float(y), float(x), original_text.strip()))
                    #             text_elements.sort(key=lambda item: (item[0], item[1]))
                    #             for _, _, text in text_elements:
                    #                 all_text.append(text)
                    #             file_data = ET.tostring(tree, encoding='utf-8', standalone=False)
                    #         except Exception as e:
                    #             print(f"Warning: Parsing {item.filename} failed - {e}")

                    #         try:
                    #             tree = ET.fromstring(file_data)
                    #             ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                    #             for row in tree.findall(".//ss:Row", ns):
                    #                 for cell in row.findall("ss:Cell", ns):
                    #                     value_element = cell.find("ss:Data", ns)
                    #                     if value_element is not None and value_element.text:
                    #                         all_text.append(value_element.text.strip())

                    #                     if cell.attrib.get('ss:MergeAcross') is not None:
                    #                         merged_value = value_element.text.strip() if value_element is not None else ""
                    #                         for _ in range(int(cell.attrib['ss:MergeAcross'])):
                    #                             all_text.append(merged_value)

                    #         except Exception as e:
                    #             print(f"Warning: Parsing {item.filename} failed - {e}")
                                
                    #     new_zip.writestr(item, file_data)

                    # # merge all text one string
                    # combined_text = ''.join(all_text)
                    
                    # # 612 debug
                    # # if file_type != "参照ファイル":
                    # #     result_map = gpt_correct_text(combined_text)
                    # #     corrected_map.update(result_map)  # 결과 맵 병합
                    # # else:
                    # #     corrected_map = ""


                    # in_memory_zip.close()
                    # new_zip.close()

                    # output_buffer.seek(0)
                    #--------------excel end------------------------------------------
                    # excel_base64 = base64.b64encode(output_buffer.getvalue()).decode('utf-8')
                    excel_base64 = base64.b64encode(file_bytes).decode('utf-8')

                    return jsonify({
                        "success": True,
                        "original_text": "",# combined_text,
                        "excel_bytes": excel_base64,
                        "combined_text": "",# combined_text,
                        "file_name": file.filename
                    })

            except Exception as e:
                logging.error(f"Error processing file {file.filename}: {str(e)}")
                return jsonify({"success": False, "error": str(e)}), 500

    return jsonify({"success": False, "error": "Invalid file type"}), 400


# 5007 debug
def remove_correction_blocks(html_text):
    pattern = re.compile(
        r'<span[^>]*?>.*?<\/span>\s*\(<span>提示:<s[^>]*?>.*?<\/s><\/span>\)',
        re.DOTALL
    )
    return pattern.sub('', html_text)

half_to_full_dict = {
    "ｦ": "ヲ", "ｧ": "ァ", "ｨ": "ィ", "ｩ": "ゥ", "ｪ": "ェ", "ｫ": "ォ",
    "ｬ": "ャ", "ｭ": "ュ", "ｮ": "ョ", "ｯ": "ッ", "ｰ": "ー",
    "ｱ": "ア", "ｲ": "イ", "ｳ": "ウ", "ｴ": "エ", "ｵ": "オ",
    "ｶ": "カ", "ｷ": "キ", "ｸ": "ク", "ｹ": "ケ", "ｺ": "コ",
    "ｻ": "サ", "ｼ": "シ", "ｽ": "ス", "ｾ": "セ", "ｿ": "ソ",
    "ﾀ": "タ", "ﾁ": "チ", "ﾂ": "ツ", "ﾃ": "テ", "ﾄ": "ト",
    "ﾅ": "ナ", "ﾆ": "ニ", "ﾇ": "ヌ", "ﾈ": "ネ", "ﾉ": "ノ",
    "ﾊ": "ハ", "ﾋ": "ヒ", "ﾌ": "フ", "ﾍ": "ヘ", "ﾎ": "ホ",
    "ﾏ": "マ", "ﾐ": "ミ", "ﾑ": "ム", "ﾒ": "メ", "ﾓ": "モ",
    "ﾔ": "ヤ", "ﾕ": "ユ", "ﾖ": "ヨ",
    "ﾗ": "ラ", "ﾘ": "リ", "ﾙ": "ル", "ﾚ": "レ", "ﾛ": "ロ",
    "ﾜ": "ワ", "ﾝ": "ン",
    "%": "％", "@": "＠"
}

full_to_half_dict = {
    '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
    '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
    'Ａ': 'A', 'Ｂ': 'B', 'Ｃ': 'C', 'Ｄ': 'D', 'Ｅ': 'E',
    'Ｆ': 'F', 'Ｇ': 'G', 'Ｈ': 'H', 'Ｉ': 'I', 'Ｊ': 'J',
    'Ｋ': 'K', 'Ｌ': 'L', 'Ｍ': 'M', 'Ｎ': 'N', 'Ｏ': 'O',
    'Ｐ': 'P', 'Ｑ': 'Q', 'Ｒ': 'R', 'Ｓ': 'S', 'Ｔ': 'T',
    'Ｕ': 'U', 'Ｖ': 'V', 'Ｗ': 'W', 'Ｘ': 'X', 'Ｙ': 'Y', 
    'Ｚ': 'Z','＋':'+','－':'-'
}

# 半角→,-全角
def half_and_full_process(text, mapping):
    return ''.join(mapping.get(c, c) for c in text)

replace_rules = {
    # 'AAA': 'AAA（全米自動車協会）', # 729 fix bug
    'ABS': 'ABS（資産担保証券、各種資産担保証券）',
    'ADB': 'ADB（アジア開発銀行）',
    'ADR': 'ADR（米国預託証券）',
    'AI': 'AI（人工知能）',
    'AIIB': 'AIIB（アジアインフラ投資銀行）',
    'APEC': 'APEC（アジア太平洋経済協力会議）',
    'API': 'API（全米石油協会）',
    'BIS': 'BIS（国際決済銀行）',
    'BOE': 'BOE（英中央銀行、イングランド銀行）',
    'BRICS（5ヵ国）': 'BRICS（ブラジル、ロシア、インド、中国、南アフリカ）',
    'CDS市場': 'CDS（クレジット・デフォルト・スワップ）市場',
    'CFROIC': 'CFROIC（投下資本キャッシュフロー利益率）',
    'Chat GPT': 'Chat GPT（AIを使った対話型サービス）',
    'CMBS': 'CMBS（商業用不動産ローン担保証券）',
    'COP26': 'COP26（国連気候変動枠組み条約第26回締約国会議）',
    'CPI': 'CPI（消費者物価指数）',
    'CSR': 'CSR（企業の社会的責任）',
    'DR': 'DR（預託証書）',
    'DRAM': 'DRAM（半導体素子を利用した記憶装置のひとつ）',
    'DX': 'DX（デジタルトランスフォーメーション）',
    'EC': 'EC（電子商取引）',
    'ECB': 'ECB（欧州中央銀行）',
    'EIA': 'EIA（米エネルギー省エネルギー情報局）',
    'EMEA': 'EMEA（欧州・中東・アフリカ）',
    'EPA': 'EPA（米環境保護局）',
    'EPS': 'EPS（一株当たり利益）',
    'ESM': 'ESM（欧州安定メカニズム）',
    'ESG': 'ESG（環境・社会・企業統治）',
    'EU': 'EU（欧州連合）',
    'EV': 'EV（電気自動車）',
    'EVA': 'EVA（経済的付加価値）',
    'FASB': 'FASB（米財務会計基準審議会）',
    'FDA': 'FDA（米国食品医薬品局）',
    'FFレート（米国の場合）': '政策金利（FFレート）',
    'FOMC': 'FOMC（米連邦公開市場委員会）',
    'FRB': 'FRB（米連邦準備制度理事会）',
    'FTA': 'FTA（自由貿易協定）',
    'G7': 'G7（主要7ヵ国会議）',
    'G8': 'G8（主要8ヵ国首脳会議）',
    'G20': 'G20（20ヵ国・地域）財務相・中央銀行総裁会議、首脳会議',
    'GDP': 'GDP（国内総生産）',
    'GPIF': '年金積立金管理運用独立行政法人（GPIF）',
    'GNP': 'GNP（国民総生産）',
    'GST　※インドの場合': 'GST（物品・サービス税）',
    'IEA': 'IEA（国際エネルギー機関）',
    'IMF': 'IMF（国際通貨基金）',
    'IoT': 'IoT（モノのインターネット）',
    'IPEF': 'IPEF（インド太平洋経済枠組み）',
    'IPO': 'IPO（新規株式公開）',
    'ISM非製造業景況': 'ISM非製造業景況指数',
    'IT': 'IT（情報技術）',
    'LBO': 'LBO（レバレッジド・バイアウト：対象企業の資産を担保に資金調達する買収）',
    'LED': 'LED（発光ダイオード）',
    'LME': 'LME（ロンドン金属取引所）',
    'LNG': 'LNG（液化天然ガス）',
    'M&A': 'M&A（企業の合併・買収）',
    'MAS': 'MAS（シンガポール金融通貨庁）',
    'MBA': 'MBA（全米抵当貸付銀行協会）',
    'MBO': 'MBO（経営陣による買収）',
    'MBS': 'MBS（住宅ローン担保証券）',
    'NAFTA': 'NAFTA（北米自由貿易協定）',
    'NAHB': 'NAHB（全米住宅建設業者協会）',
    'NAIC': 'NAIC（全米保険監督官協会）',
    'NAR': 'NAR（全米不動産業者協会）',
    'NDF': 'NDF（為替先渡取引のひとつ）',
    'NISA': 'NISA（少額投資非課税制度）',
    'OECD': 'OECD（経済協力開発機構）',
    'OEM': 'OEM（相手先ブランドによる生産）',
    'OPEC': 'OPEC（石油輸出国機構）',
    'OPECプラス': 'OPECプラス（OPEC（石油輸出国機構）と非加盟産油国で構成するOPECプラス）',
    'PBR': 'PBR（株価純資産倍率）',
    'PCE': 'PCE（個人消費支出）',
    'PCFR': 'PCFR（株価キャッシュフロー倍率）',
    'PER': 'PER（株価収益率）',
    'PMI': 'PMI（購買担当者景気指数）',
    'PPI': 'PPI（生産者物価指数）',
    'QE': 'QE（量的金融緩和）',
    'QT': 'QT（量的引き締め）',
    'Quad': 'Quad（日米豪印戦略対話）',
    'RBA': 'RBA（豪州準備銀行）',
    'RCEP': 'RCEP（地域的な包括的経済連携協定）',
    'RBI': 'RBI（インド準備銀行）',
    'ROA': 'ROA（総資産利益率）',
    'ROE': 'ROE（自己資本利益率）',
    'S&L': 'S&L（貯蓄貸付組合）',
    'SDGs': 'SDGs（持続可能な開発目標）',
    'SEC': 'SEC（米証券取引委員会）',
    'SQ': 'SQ（特別清算指数）',
    'SRI': 'SRI（社会的責任投資）',
    'SUV': 'SUV（スポーツ用多目的車）',
    'TALF': 'TALF（ターム物資産担保証券貸出制度）',
    'TOB': 'TOB（株式公開買付け）',
    'TPP': 'TPP（環太平洋経済連携協定）',
    'UAE': 'UAE（アラブ首長国連邦）',
    'UAW': 'UAW（全米自動車労働組合）',
    'USDA': 'USDA（米国農務省）',
    'USMCA': 'USMCA（米国・メキシコ・カナダ協定）',
    'USTR': 'USTR（米通商代表部）',
    'VAT': 'VAT（付加価値税）',
    'WTI': 'WTI（ウエスト・テキサス・インターミディエート）',
    'WTO': 'WTO（世界貿易機関）',
    'アセットアロケーション': 'アセットアロケーション（資産配分）',
    'アンダーウェイト': 'アンダーウェイト（ベンチマークに比べ低めの投資比率）',
    'オーバーウエイト': 'オーバーウエイト（ベンチマークに比べ高めの投資比率）',
    'E-コマース': 'Eコマース（電子商取引）',
    'e-コマース': 'eコマース（電子商取引）',
    'イールドカーブ': 'イールドカーブ（利回り曲線）',
    'イールドカーブ・コントロール': 'イールドカーブ・コントロール（長短金利操作）',
    'イールドカーブのスティープ化': 'イールドカーブのスティープ化（長・短金利格差の拡大）',
    'イールドカーブのフラット化': 'イールドカーブのフラット化（長・短金利格差の縮小）',
    'インカムゲイン': 'インカムゲイン（利子収入）',
    'インタラクティブ': 'インタラクティブ（双方向性）',
    'エクイティ・ファイナンス': 'エクイティ・ファイナンス（新株発行等による資金調達）',
    'オバマケア': 'オバマケア（医療保険制度改革法）',
    'オンデマンド': 'オンデマンド（注文生産）',
    'カントリー･アロケーション': 'カントリー･アロケーション（国別資産配分）',
    '逆イールド': '逆イールド（短期債券の利回りが長期債券の利回りを上回っている状態）',
    # 'キャッシュフロー': 'キャッシュフロー（現金収支）',
    'キャピタルゲイン': 'キャピタルゲイン（値上がり益）',
    'キャリートレード': 'キャリートレード（低金利の資金を調達して、高金利の資産で運用する取引）',
    'クレジットスプレッド': 'クレジットスプレッド（企業の信用力の差による利回りの差）',
    'グローバリゼーション': 'グローバリゼーション（地球規模化）',
    'コジェネレーション': 'コジェネレーション（熱電供給システム）',
    'コーポレート・ガバナンス': 'コーポレート・ガバナンス（企業統治）',
    'コングロマリット': 'コングロマリット（複合企業）',
    'コンソーシアム': 'コンソーシアム（共同事業）',
    'サーベイランス': 'サーベイランス（調査監視）',
    'サステナビリティ': 'サステナビリティ（持続可能性）',
    'サブプライムローン': 'サブプライムローン（信用度の低い個人向け住宅融資）',
    'サプライチェーン': 'サプライチェーン（供給網）',
    'ジェネリック医薬品': 'ジェネリック医薬品（後発薬）',
    'シクリカル': 'シクリカル（景気敏感）',
    'シャドーバンキング': 'シャドーバンキング（影の銀行）',
    'ショートポジション': 'ショートポジション（売り持ち）',
    '信用市場': '企業の信用リスクを取引する市場',
    'スティープ化': 'スティープ化（長短金利格差の拡大）',
    'ストレステスト': 'ストレステスト（健全性審査）',
    'スプレッド': 'スプレッド（利回り格差）',
    'スマートシティー': 'スマートシティー（ITを活用した次世代型の都市）',
    'スマートモビリティ': 'スマートモビリティ（従来の交通・移動を変える新たなテクノロジー）',
    'セーフティネット': 'セーフティネット（安全網）',
    '全人代': '全人代（全国人民代表大会）',
    'ソフトランディング': 'ソフトランディング（軟着陸）',
    'ダイバーシティ': 'ダイバーシティ（多様性）',
    'ターミナルレート': 'ターミナルレート（政策金利の最終到達水準）',
    'テーパリング': 'テーパリング（量的金融緩和の縮小）',
    'ディフェンシブ': 'ディフェンシブ（景気に左右されにくい）',
    'デフォルト': 'デフォルト（債務不履行）',
    'デュレーション': 'デュレーション（金利感応度）', 
    'デリバティブ': 'デリバティブ（金融派生商品）',
    'ドルペッグ制': 'ドルペッグ（連動）制',
    'バイオシミラー': 'バイオシミラー（後続薬）',
    'バイオマス': 'バイオマス（生物を利用して物質やエネルギーを得ること）',
    'バーチャル': 'バーチャル（仮想）',
    'バリュエーション': 'バリュエーション（投資価値評価）',
    'バリュー': 'バリュー（割安）',
    '5G': '5G（第5世代移動通信システム）',
    'フィンテック': 'フィンテック（金融と技術の融合）', 
    'フェアバリュー': 'フェアバリュー（適正価格）',
    'フェーズ2': 'フェーズ2（臨床試験の中間段階）',
    'フェーズ3': 'フェーズ3（臨床試験の最終段階）',
    'フードデリバリー': 'フードデリバリー（料理等の宅配サービス）',
    'フルインベストメント': 'フルインベストメント（高位組入）',
    'ブロードバンド': 'ブロードバンド（大容量･高速通信）',
    'ポテンシャル': 'ポテンシャル（潜在力）',
    'ポピュリズム': 'ポピュリズム（大衆迎合主義）',
    'モーゲージ': 'モーゲージ（不動産担保ローン）',
    'モーゲージ債': 'モーゲージ債（不動産ローン担保債券）',
    'モラルハザード': 'モラルハザード（倫理崩壊）',
    'リエンジニアリング': 'リエンジニアリング（業務の抜本的革新）',
    'リオープン': 'リオープニング（経済活動再開）',
    'リセッション': 'リセッション（景気後退）',
    'リターンリバーサル': 'リターンリバーサル（過剰反応効果）',
    'リバウンド': 'リバウンド（反発）',
    'リバランス': 'リバランス（投資比率の再調整）',
    'レパトリ減税': 'レパトリ（海外収益の本国還流）減税',
    'レバレッジドローン': '低格付け等の借り手向け融資',
    'レラティブ・バリュー': 'レラティブ・バリュー（相対価値）',
    'ロックダウン': 'ロックダウン（都市封鎖）',
    'ロングポジション': 'ロングポジション（買い持ち）',
    'EDAツール': 'EDAツール（電子設計自動化ツール）', #623
    'TOPIX':'TOPIX（東証株価指数）', #63207
    '利回りは上昇': '利回りは上昇（価格は下落）', #730
    '利回りは低下': '利回りは低下（価格は上昇）', #730

    '利回りの上昇': '利回りの上昇(価格は低下)', #730
    '利回りの低下': '利回りの低下(価格は上昇)', #730

    '国債利回りは、月間で上昇': '国債利回りは、月間で上昇(価格は低下)', #730
    '国債利回りは、月間で下落': '国債利回りは、月間で下落(価格は上昇)', #730

    '債券利回りは上昇': '債券利回りは上昇(価格は低下)', #730
    '債券利回りは下落': '債券利回りは上昇(価格は上昇)', #730
    
    'シャリア': 'シャリーア',
    'TTM': '仲値',
    '殆ど': 'ほとんど',
    '真似': 'まね',
    '亘る': 'わたる',
    '但し': 'ただし',
    '牽制': 'けん制',
    '牽引': 'けん引',
    '終焉': '終えん',
    '収斂': '収れん',
    '逼迫': 'ひっ迫',
    'ヶ月': 'ヵ月',
    '入替え': '入れ替え',
    '入替': '入れ替',
    '売付':'売り付け',
    '売付け':'売り付け',
    '格付': '格付け', 
    '買建て': '買い建て',
    '売建て': '売り建て',
    '切上げ': '切り上げ',
    '切捨て': '切り捨て',
    '組入れ': '組み入れ', 
    '繰上げ償還': '繰上償還',
    '先き行き': '先行き',
    '下支える': '下支えする',
    '取り引き': '取引',
    '引上げ': '引き上げ',
    '引下げ': '引き下げ',
    '引続き': '引き続き',
    '引締め': '引き締め',
    '薄商い': '取引量が少なく',
    'コア銘柄': '中核銘柄、コア（中核）銘柄',
    'トリガー': 'きっかけ',
    'ブルーチップ企業': '優良企業',
    'ハト派': '金融緩和に前向き',
    'タカ派': '金融引き締め重視',
    '相場': '市場',
    '連れ高': '影響を受けて上昇',
    '伝播': '広がる',
    'でんぱ': '広がる',
    'レンジ': '範囲',
    '回金': '円転',
    'ローン': '貸し付け',
    '所謂': 'いわゆる',
    '暫く': 'しばらく',
    '留まる': 'とどまる',
    '止まる': 'とどまる',
    '尚': 'なお',
    '筈': 'はず',
    '蓋然性': '可能性',
    '商い': '出来高',
    '後倒し': '延期',
    '経済正常化': '経済活動正常化',
    '金融正常化': '金融政策正常化',
    '日本銀行': '日銀',
    '政治的リスク': '政治リスク',
    '地政学リスク': '地政学的リスク',
    'への組み入れ': 'の組み入れ',
    'マイナスに寄与': 'マイナスに影響',
    'マイナス寄与': 'マイナス影響', #829
    '米国国債': '米国債',
    '新型コロナ': '新型コロナウイルス',
    'コロナウイルス': '新型コロナウイルス',
    '立ち後れ': '立ち遅れ',
    '伸張': '伸長',
    'ダウ平均': 'ダウ平均株価',
    'NYダウ': 'ダウ平均株価',
    '中銀': '中央銀行', #623
    '行われ': '行なわれ', #623
    '行い': '行ない', #623
    '行わない': '行なわない', #821
    '行った':'行なった',
    '行う': '行なう', #623
    '行って': '行なって', #623
    '行われる': '行なわれる',
    'なりしました': 'なしました', #180015,628
    '買い付けました': '買い付けしました',
    # '買い付け': '買い付けし', #64977 , 829 fix
    '買付':'買い付け',
    '買付け':'買い付けし', #63207
    '売り付けました':'売り付けしました', #628
    '売り立てました':'売り立てしました', #628
    '割安に': '割安感のある',
    'MSCIインド指数': 'MSCIインド・インデックス',
    'サステナブル': 'サスティナブル',
    'エンターテイメント': 'エンターテインメント',
    '亘': 'わた',
    'REIT': 'リート', #629
    '燻': 'くすぶ',
    'トランプ政権': 'トランプ米政権',
    'トランプ大統領': 'トランプ米大統領',
    '米トランプ大統領': '米トランプ大統領',
    '好感され、下落し': '好感され下落し',
    '嫌気され、下落し': '嫌気され下落し',
    '好感され、上昇し': '好感され上昇し',
    '嫌気され、上昇し': '嫌気され上昇し',
    '留': 'とど', #629
    '当社': '同社', #629
    '牽': 'けん', #629
    'こと目指している': 'ことを目指している', #630
    'グローバルで事業': 'グローバルに事業', #630
    '積み増す': '積み増しする', #630 
    '取組み': '取り組み',
    '魅力度': '<sup>※</sup>魅力度',
    'フリーキャッシュフロー': 'フリーキャッシュフロー(税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたもの )', #726
    'フリー・キャッシュフロー': 'フリーキャッシュフロー(税引後営業利益に減価償却費を加え、設備投資額と運転資本の増加を差し引いたもの )', #726

    'ボラティリティ': 'ボラティリティ（価格変動性）', #829
    'ファンダメンタルズ': 'ファンダメンタルズ（経済の基礎的条件）', #829

    
}

replace_rules1 ={
    'シャリア': 'シャリーア',
    'TTM': '仲値',
    '殆ど': 'ほとんど',
    '真似': 'まね',
    '亘る': 'わたる',
    '但し': 'ただし',
    '牽制': 'けん制',
    '牽引': 'けん引',
    '終焉': '終えん',
    '収斂': '収れん',
    '逼迫': 'ひっ迫',
    'ヶ月': 'ヵ月',
    '入替え': '入れ替え',
    '入替': '入れ替',
    '売付':'売り付け',
    '売付け':'売り付け',
    '格付': '格付け', 
    '買建て': '買い建て',
    '売建て': '売り建て',
    '切上げ': '切り上げ',
    '切捨て': '切り捨て',
    '組入れ': '組み入れ', 
    '繰上げ償還': '繰上償還',
    '先き行き': '先行き',
    '下支える': '下支えする',
    '取り引き': '取引',
    '引上げ': '引き上げ',
    '引下げ': '引き下げ',
    '引続き': '引き続き',
    '引締め': '引き締め',
    '薄商い': '取引量が少なく',
    'コア銘柄': '中核銘柄、コア（中核）銘柄',
    'トリガー': 'きっかけ',
    'ブルーチップ企業': '優良企業',
    'ハト派': '金融緩和に前向き',
    'タカ派': '金融引き締め重視',
    '相場': '市場',
    '連れ高': '影響を受けて上昇',
    '伝播': '広がる',
    'でんぱ': '広がる',
    'レンジ': '範囲',
    '回金': '円転',
    'ローン': '貸し付け',
    '所謂': 'いわゆる',
    '暫く': 'しばらく',
    '留まる': 'とどまる',
    '止まる': 'とどまる',
    '尚': 'なお',
    '筈': 'はず',
    '蓋然性': '可能性',
    '商い': '出来高',
    '後倒し': '延期',
    '経済正常化': '経済活動正常化',
    '金融正常化': '金融政策正常化',
    '日本銀行': '日銀',
    '政治的リスク': '政治リスク',
    '地政学リスク': '地政学的リスク',
    'への組み入れ': 'の組み入れ',
    'マイナスに寄与': 'マイナスに影響',
    'マイナス寄与': 'マイナス影響', #829
    '米国国債': '米国債',
    '新型コロナ': '新型コロナウイルス',
    'コロナウイルス': '新型コロナウイルス',
    '立ち後れ': '立ち遅れ',
    '伸張': '伸長',
    'ダウ平均': 'ダウ平均株価',
    'NYダウ': 'ダウ平均株価',
    '中銀': '中央銀行', #623
    '行われ': '行なわれ', #623
    '行い': '行ない', #623
    '行わない': '行なわない', #821
    '行った':'行なった',
    '行う': '行なう', #623
    '行って': '行なって', #623
    '行われる': '行なわれる',
    'なりしました': 'なしました', #180015,628
    '買い付けました': '買い付けしました',
    # '買い付け': '買い付けし', #64977 , 824fix
    '買付':'買い付け',
    '買付け':'買い付けし', #63207
    '売り付けました':'売り付けしました', #628
    '売り立てました':'売り立てしました', #628
    '割安に': '割安感のある',
    'MSCIインド指数': 'MSCIインド・インデックス',
    'サステナブル': 'サスティナブル',
    'エンターテイメント': 'エンターテインメント',
    '亘': 'わた',
    'REIT': 'リート', #629
    '燻': 'くすぶ',
    'トランプ政権': 'トランプ米政権',
    'トランプ大統領': 'トランプ米大統領',
    '米トランプ大統領': '米トランプ大統領',
    '好感され、下落し': '好感され下落し',
    '嫌気され、下落し': '嫌気され下落し',
    '好感され、上昇し': '好感され上昇し',
    '嫌気され、上昇し': '嫌気され上昇し',
    '留': 'とど', #629
    '当社': '同社', #629
    '牽': 'けん', #629
    'こと目指している': 'ことを目指している', #630
    'グローバルで事業': 'グローバルに事業', #630
    '積み増す': '積み増しする', #630 
    '取組み': '取り組み',
    '魅力度': '<sup>※</sup>魅力度'
}


replace_rules2 ={
    '政治的リスク': '政治リスク',
    '地政学リスク': '地政学的リスク',
}

def merge_brackets(content: str) -> str:
    """
    括号内换行符: 'CPI（消費者物\n価指数）' -> 'CPI（消費者物価指数）'
    """
    # return regcheck.sub(r'（[^）\n\r]*[\n\r]+[^）]*）', lambda m: m.group(0).replace("\n", "").replace("\r", ""), content)
    content = regcheck.sub(r'([^\s\n\r])[\s\n\r]+（', r'\1（', content)

    def replacer(match):
        inside = match.group(1)
        cleaned = regcheck.sub(r'[\s\u3000]+', '', inside)
        return f'（{cleaned}）'

    return regcheck.sub(r'（(.*?)）', replacer, content, flags=regcheck.DOTALL)


# (4月30日 → 2025年4月30日)
def insert_year_by_regex(date_str: str, full_text: str, date_pos: int) -> str:
    year_matches = list(regcheck.finditer(r'(\d{4})年', full_text[:date_pos]))
    if year_matches:
        last_year = year_matches[-1].group(1)
        return f'{last_year}年{date_str}'
    return date_str

# (4月30日 → 2025年4月30日)
def year_half_dict(text: str) -> str:
    full_half = {
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9'
    }
    return ''.join(full_half.get(c, c) for c in text)


def opt_check_eng(content, rules):
    if not isinstance(rules, dict):
        raise TypeError(f"`rules` must be a dict, got {type(rules)}")
    
    content = merge_brackets(content)
    content = content.replace("(", "（").replace(")", "）")
    lines = content.strip().splitlines()

    seen_raw = set()
    seen_full = set()
    results = []

    for line in lines:
        result = []
        normalized_line = line.replace("\n", "").replace(" ", "")

        for k, v in rules.items():
            raw_key = k.replace("(", "（").replace(")", "）")
            full_key = v.replace("(", "（").replace(")", "）")

            if '(' not in full_key and '（' not in full_key:
                continue
            
            escaped_k = regcheck.escape(raw_key)
            escaped_v = regcheck.escape(full_key)

            # ------------------------------
            # keyword 没有对应的pattern
            # ------------------------------
            new_k = escaped_k
            paren_pattern = f"{escaped_k}（[^）]+）"

            if raw_key.isalpha() or raw_key in ["S&L", "M&A"]:
                if raw_key == "OPEC":
                    new_k = f"(?<![a-zA-Z]){escaped_k}(?!プラス|[a-zA-Z])"
                elif raw_key == "スティープ化":
                    new_k = f"(?<!イールドカーブの){escaped_k}"
                elif raw_key == "イールドカーブ":
                    new_k = f"{escaped_k}(?!・コントロール|のスティープ化|のフラット化)"
                elif raw_key == "キャッシュフロー":
                    new_k = f"(?<!フリー){escaped_k}"
                elif raw_key == "キャリートレード":
                    new_k = f"(?<!円){escaped_k}"
                elif raw_key == "スプレッド":
                    new_k = f"(?<!クレジット){escaped_k}"
                elif raw_key == "バリュー":
                    new_k = f"(?<!レラティブ・|フェア){escaped_k}"
                elif raw_key == "モーゲージ":
                    new_k = f"{escaped_k}(?!債)"
                elif raw_key == "商い":
                    new_k = f"(?<!薄){escaped_k}"
                else:
                    new_k = f"(?<![a-zA-Z]){escaped_k}(?![a-zA-Z])"

            matched_full = regcheck.search(escaped_v, normalized_line)
            matched_raw_with_paren = regcheck.search(paren_pattern, normalized_line)
            matched_raw = regcheck.search(new_k, normalized_line)

            # ✅ 校验full_key,第一次出现
            if matched_full and full_key not in seen_full:
                seen_raw.add(raw_key)
                seen_full.add(full_key)
                continue

            # ✅ full_key ,第二次出现
            elif matched_full and full_key in seen_full:
                result.append({full_key: "删除"})
            
            elif matched_raw_with_paren:
                result.append({matched_raw_with_paren.group(): full_key})
                seen_raw.add(raw_key)
                seen_full.add(full_key)

            elif matched_raw and raw_key not in seen_raw:
                result.append({raw_key: full_key})
                seen_raw.add(raw_key)
                seen_full.add(full_key)

        results.append(result)

    return results

def opt_check_ruru1(content, rules):
    content = merge_brackets(content)

    result = []
    for k, v in rules.items():
        raw_key = k.replace("(", "（").replace(")", "）")
        full_key = v.replace("(", "（").replace(")", "）")

        escaped_k = regcheck.escape(raw_key)
        escaped_v = regcheck.escape(full_key)

        new_k = escaped_k
        if raw_key.isalpha() or raw_key in ["S&L", "M&A"]:
            if raw_key == "OPEC":
                new_k = f"(?<![a-zA-Z]){escaped_k}(?!プラス|[a-zA-Z])"
            elif raw_key == "スティープ化":
                new_k = f"(?<!イールドカーブの){escaped_k}"
            elif raw_key == "イールドカーブ":
                new_k = f"{escaped_k}(?!・コントロール|のスティープ化|のフラット化)"
            elif raw_key == "キャッシュフロー":
                new_k = f"(?<!フリー){escaped_k}"
            elif raw_key == "キャリートレード":
                new_k = f"(?<!円){escaped_k}"
            elif raw_key == "スプレッド":
                new_k = f"(?<!クレジット){escaped_k}"
            elif raw_key == "バリュー":
                new_k = f"(?<!レラティブ・|フェア){escaped_k}"
            elif raw_key == "モーゲージ":
                new_k = f"{escaped_k}(?!債)"
            elif raw_key == "商い":
                new_k = f"(?<!薄){escaped_k}"
            else:
                new_k = f"(?<![a-zA-Z]){escaped_k}(?![a-zA-Z])"
        #  中銀
        elif raw_key == "中銀":
            matches = regcheck.finditer(escaped_v, content)
            exclude = False
            for m in matches:
                prefix = content[max(0, m.start() - 2): m.start()]
                if prefix and not regcheck.match(r"[ \t\n\r]", prefix):
                    exclude = True
                    break
            if exclude:
                new_k = escaped_k
                full_match = None
            else:
                full_match = regcheck.search(escaped_v, content)

            
        raw_match = regcheck.search(new_k, content)
        full_match = regcheck.search(escaped_v, content)

        if raw_key != "中銀":
            if full_match and raw_match:
                if full_match.start() <= raw_match.start():
                    continue
            elif full_match and not raw_match:
                continue

        if raw_match:
            result.append({raw_key: full_key})
        
    return result

def keyword_pair_exists(content, keyword_a, keyword_b):
    return keyword_a in content and keyword_b in content

# 地政学リスク/政治的リスク
def opt_check_ruru2(content, replace_rules2):
    content = merge_brackets(content)

    result = []

    keyword_pairs = [
        ("地政学リスク", "地政学的リスク"),
        ("政治的リスク", "政治リスク")
    ]

    for a, b in keyword_pairs:
        if keyword_pair_exists(content, a, b):
            result.append({a: b})

    return result

# 0501 debug
def find_corrections(corrected_text,input_text,pageNumber):
    corrections = []
    pattern = r'<span\s+style="color:red;">([\s\S]*?)<\/span>\s*\(<span>\s*修正理由[::]\s*([\s\S]*?)\s*<s[^>]*>([\s\S]*?)<\/s>\s*→\s*([\s\S]*?)<\/span>\)'
    matches = re.findall(pattern, corrected_text)

    print("Matches found:", matches)
    # <span style="color:red;">上午12时00分</span> (<span>修正理由: 不要な中国語表記 <s style="background:yellow;color:red">上午12时00分</s> → （削除）</span>)

    for match in matches:
        if len(match) == 4:
            corrected_text_re = match[0]  #610 debug
            reason_type = match[1].strip()
            original_text = match[2].strip()
            target_text = match[3].strip()

            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": corrected_text_re,
                "comment": comment,
                "reason_type":reason_type,
                "check_point": input_text.strip(),
                "locations": [],
                "intgr": False, 
            })
    
    return corrections

# 814 ,add dotfind 句読点
#------------------------------------------------------------
def check_fullwidth_period(sentence):
    return sentence.endswith("。")

#---------------------------------------------------------------------------

# 0623 debug
def find_corrections_wording(input_text,pageNumber,tenbrend,fund_type,input_list):
    corrections = []

#-------------------
    #常用外汉字
    common_par = r"((啞|蛙|鴉|埃|挨|曖|靄|軋|斡|按|庵|鞍|闇|已|夷|畏|韋|帷|萎|椅|葦|彙|飴|謂|閾|溢|鰯|尹|咽|殷|淫|隕|蔭|于|迂|盂|烏|鬱|云|暈|穢|曳|洩|裔|穎|嬰|翳|腋|曰|奄|宛|怨|俺|冤|袁|婉|焉|堰|淵|焰|筵|厭|鳶|燕|閻|嚥|嗚|凰|嘔|鴨|甕|襖|謳|鶯|鷗|鸚|臆|俤|瓜|呵|苛|珂|迦|訛|訶|跏|嘩|瑕|榎|窩|蝦|蝸|鍋|顆|牙|瓦|臥|俄|峨|訝|蛾|衙|駕|芥|乖|廻|徊|恢|晦|堺|潰|鞋|諧|檜|蟹|咳|崖|蓋|漑|骸|鎧|喀|廓|摑|攪|愕|萼|諤|顎|鰐|樫|絣|筈|葛|闊|鰹|萱|奸|串|旱|函|咸|姦|宦|柑|竿|悍|桓|涵|菅|嵌|鉗|澗|翰|諫|瞰|檻|灌|玩|雁|翫|頷|癌|贋|几|卉|其|祁|耆|埼|悸|揆|毀|箕|畿|窺|諱|徽|櫃|妓|祇|魏|蟻|掬|麴|吃|屹|拮|謔|仇|臼|汲|灸|咎|邱|柩|笈|躬|厩|嗅|舅|炬|渠|裾|噓|墟|鋸|遽|欅|匈|怯|俠|脇|莢|竟|卿|僑|嬌|蕎|鋏|頰|橿|疆|饗|棘|髷|巾|僅|禽|饉|狗|惧|軀|懼|俱|喰|寓|窟|粂|偈|荊|珪|畦|脛|頃|痙|詣|禊|閨|稽|頸|髻|蹊|鮭|繫|睨|戟|隙|抉|頁|訣|蕨|姸|倦|虔|捲|牽|喧|硯|腱|鍵|瞼|鹼|呟|眩|舷|諺|乎|姑|狐|股|涸|菰|袴|壺|跨|糊|醐|齬|亢|勾|叩|尻|吼|肛|岡|庚|杭|肴|咬|垢|巷|恍|恰|狡|桁|胱|崗|梗|喉|腔|蛤|幌|煌|鉤|敲|睾|膏|閤|膠|篝|縞|薨|糠|藁|鮫|壙|曠|劫|毫|傲|壕|濠|嚙|轟|剋|哭|鵠|乞|忽|惚|昏|痕|渾|褌|叉|些|嗟|蓑|磋|坐|挫|晒|柴|砦|犀|賽|鰓|榊|柵|炸|窄|簀|刹|拶|紮|撒|薩|珊|餐|纂|霰|攢|讃|斬|懺|仔|弛|此|址|祀|屍|屎|柿|茨|恣|砥|祠|翅|舐|疵|趾|斯|覗|嗜|滓|獅|幟|摯|嘴|熾|髭|贄|而|峙|痔|餌|竺|雫|𠮟|悉|蛭|嫉|膝|櫛|柘|洒|娑|這|奢|闍|杓|灼|綽|錫|雀|惹|娶|腫|諏|鬚|呪|竪|綬|聚|濡|襦|帚|酋|袖|羞|葺|蒐|箒|皺|輯|鍬|繡|蹴|讐|鷲|廿|揉|絨|粥|戌|閏|楯|馴|杵|薯|藷|汝|抒|鋤|妾|哨|秤|娼|逍|廂|椒|湘|竦|鈔|睫|蛸|鉦|摺|蔣|裳|誦|漿|蕭|踵|鞘|篠|聳|鍾|醬|囁|杖|茸|嘗|擾|攘|饒|拭|埴|蜀|蝕|燭|褥|沁|芯|呻|宸|疹|蜃|滲|賑|鍼|壬|訊|腎|靱|塵|儘|笥|祟|膵|誰|錐|雖|隋|隧|芻|趨|鮨|丼|凄|栖|棲|甥|貰|蜻|醒|錆|臍|瀞|鯖|脆弱?|贅|脊|戚|晰|蹟|泄|屑|浙|啜|楔|截|尖|苫|穿|閃|陝|釧|揃|煎|羨|腺|詮|煽|箋|撰|箭|賤|蟬|癬|喘|膳|狙|疽|疏|甦|楚|鼠|遡|蘇|齟|爪|宋|炒|叟|蚤|曾|湊|葱|搔|槍|漕|箏|噌|瘡|瘦|踪|艘|薔|甑|叢|藪|躁|囃|竈|鰺|仄|捉|塞|粟|杣|遜|噂|樽|鱒|侘|咤|詫|陀|拿|荼|唾|舵|楕|驒|苔|殆|堆|碓|腿|頽|戴|醍|托|鐸|凧|襷|燵|坦|疸|耽|啖|蛋|毯|湛|痰|綻|憚|歎|簞|譚|灘|雉|馳|蜘|緻|筑|膣|肘|冑|紐|酎|厨|蛛|註|誅|疇|躊|佇|楮|箸|儲|瀦|躇|吊|帖|喋|貼|牒|趙|銚|嘲|諜|寵|捗|枕|槌|鎚|辻|剃|挺|釘|掟|梯|逞|啼|碇|鼎|綴|鄭|薙|諦|蹄|鵜|荻|擢|溺|姪|轍|辿|唸|塡|篆|顚|囀|纏|佃|淀|澱|臀|兎|妬|兜|堵|屠|賭|宕|沓|套|疼|桶|淘|萄|逗|棹|樋|蕩|鄧|橙|濤|檮|櫂|禱|撞|禿|瀆|栃|咄|沌|遁|頓|吞|貪|邇|匂|韮|涅|禰|捏|捻|撚|膿|囊|杷|爬|琶|頗|播|芭|罵|蟇|胚|徘|牌|稗|狽|煤|帛|柏|剝|粕|箔|莫|駁|瀑|曝|畠|捌|撥|潑|醱|筏|跋|噺|氾|汎|叛|袢|絆|斑|槃|幡|攀|挽|磐|蕃|屁|庇|砒|脾|痺|鄙|誹|臂|枇|毘|梶|媚|琵|薇|靡|疋|畢|逼|謬|豹|憑|瓢|屛|廟|牝|瀕|憫|鬢|斧|阜|訃|俯|釜|腑|孵|鮒|巫|葡|撫|蕪|諷|祓|吻|扮|焚|糞|幷|聘|蔽|餅|斃|袂|僻|璧|襞|蔑|瞥|扁|篇|騙|娩|鞭|哺|圃|蒲|戊|牡|姥|菩|呆|彷|庖|苞|疱|捧|逢|蜂|蓬|鞄|鋒|牟|芒|茫|虻|榜|膀|貌|鉾|謗|吠|卜|勃|梵|昧|邁|枡|俣|沫|迄|曼|蔓|瞞|饅|鬘|鰻|蜜|鵡|冥|瞑|謎|麵|蒙|朦|勿|籾|悶|揶|爺|鑓|喩|揄|愈|楡|尤|釉|楢|猷|飫|輿|孕|妖|拗|涌|痒|傭|熔|瘍|蠅|沃|螺|萊|蕾|洛|埒|拉|辣|瀾|爛|鸞|狸|裡|罹|籬|戮|慄|掠|笠|溜|榴|劉|瘤|侶|梁|聊|菱|寥|蓼|淋|燐|鱗|屢|蛉|蠣|櫟|礫|轢|煉|漣|憐|簾|鰊|攣|賂|魯|濾|廬|櫓|蘆|鷺|弄|牢|狼|榔|瘻|﨟|臘|朧|蠟|籠|聾|肋|勒|漉|麓|窪|歪|猥|隈|或|罠|椀|碗|彎|一旦).{,5})"
    common_list = regcheck.findall(common_par, input_text)
    
    for word in common_list:
        reason_type = "常用外漢字の使用"
        corrections.append({
            "page": pageNumber,
            "original_text": word[0],  # original_text,
            "comment": word[0],
            "reason_type": reason_type,
            "check_point": word[1],
            "locations": [],
            "intgr": False,  
        })
#-------------------
    if fund_type == 'public':
        # （半角→全角） -0.09% → -0.09％
        pattern_half_width_katakana = r"[ｦ-ﾝ%＠]+"
        half_width_katakana_matches = regcheck.findall(pattern_half_width_katakana, input_text)

        for match in half_width_katakana_matches:
            corrected_text_re = half_and_full_process(match,half_to_full_dict)  # 半角→全角
            reason_type = "半角を全角統一"
            original_text = match
            target_text = corrected_text_re
            # 「％」表記の統一（半角→全角） -0.09% → -0.09％
            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": original_text,#corrected_text_re
                "comment": comment,
                "reason_type": reason_type,
                "check_point": reason_type,
                "locations": [],
                "intgr": False, 
            })

        # # （半角括弧 → 全角括弧） -() → () ,with date format: \((?!\d{4}年\d{1,2}月\d{1,2}日)([^)]+)\)
        # pattern_half_width_kuohao = r"\(([^)]+)\)"
        # half_width_kuohao_matches = regcheck.findall(pattern_half_width_kuohao, input_text)

        # for match in half_width_kuohao_matches:
        #     corrected_text_re = half_and_full_process(match,half_to_full_dict)  # 半角→全角
        #     reason_type = "半角括弧を全角括弧に統一"
        #     original_text = match
        #     converted = corrected_text_re
        #     target_text = re.sub(r'\(([^)]+)\)', r'（\1）', converted)
        #     # ()表記の統一(分配金再投資)） -(分配金再投資) → （分配金再投資）
        #     comment = f"{reason_type} {original_text} → {target_text}"

        #     corrections.append({
        #         "page": pageNumber,
        #         "original_text": original_text,#corrected_text_re
        #         "comment": comment,
        #         "reason_type": reason_type,
        #         "check_point": input_text.strip(),
        #         "locations": [],
        #         "intgr": False, 
        #     })

        # 半角→全角
        pattern_full_width_numbers_and_letters = r"[０-９Ａ-Ｚ＋－]+"
        full_width_matches = regcheck.findall(pattern_full_width_numbers_and_letters, input_text)

        for match in full_width_matches:
            corrected_text_re = half_and_full_process(match,full_to_half_dict)  # 全角→半角
            reason_type = "全角を半角統一"
            original_text = match
            target_text = corrected_text_re

            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": original_text,
                "comment": comment,
                "reason_type": reason_type,
                "check_point": reason_type,
                "locations": [],
                "intgr": False, 
            })
            
        # （注0-9）--删除
        pattern_full_delete = r"（注[0-9]+）"
        full_width_matches_delete = regcheck.findall(pattern_full_delete, input_text)

        for match in full_width_matches_delete:
            corrected_text_re = match
            reason_type = "删除"
            original_text = match
            target_text = corrected_text_re

            comment = f"{reason_type} {original_text} → {target_text}"

            corrections.append({
                "page": pageNumber,
                "original_text": original_text,
                "comment": comment,
                "reason_type": reason_type,
                "check_point": reason_type,
                "locations": [],
                "intgr": False,
            })
#-------------------
    # 年度
    # if fund_type == 'public':
    #     cleaned_text = regcheck.sub(r'\n\s*', '', input_text)
    #     date_pattern = r'(?<!\d{4}年)(\d{1,2})月(\d{1,2})日'

    #     for match in regcheck.finditer(date_pattern, cleaned_text):
    #         date_str = match.group(0)               # '4月30日'
    #         date_pos = match.start()            
    #         full_date = insert_year_by_regex(date_str, cleaned_text, date_pos)
    #         half_date = year_half_dict(full_date)

    #         context_pattern = r'.{0,8}' + regcheck.escape(date_str)
    #         context_match = regcheck.search(context_pattern, cleaned_text)
    #         original_text = context_match.group() if context_match else date_str

    #         comment = f"{original_text} → {half_date}"
    #         corrections.append({
    #             "page": pageNumber,
    #             "original_text": original_text,
    #             "comment": comment,
    #             "reason_type": '年度用語の統一',
    #             "check_point": '年度用語の統一',
    #             "locations": [],
    #             "intgr": False,  # for debug
    #         })
#-------------------
    # 英略词
    if fund_type == 'public':
        results = opt_check_eng(input_text, replace_rules)

        for line_result in results:
            if line_result:
                for item in line_result:
                    if isinstance(item, dict):
                        for original_text, corrected_text_re in item.items():
                            reason_type = "用語の統一"
                        
                            if corrected_text_re == "删除":
                                comment = f"{original_text} → トルは不要"
                            else:
                                comment = f"{original_text} → {corrected_text_re}"

                            corrections.append({
                                "page": pageNumber,
                                "original_text": original_text,
                                "comment": comment,
                                "reason_type": reason_type,
                                "check_point": reason_type,
                                "locations": [],
                                "intgr": False,
                            })


        results_ruru1 = opt_check_ruru1(input_text, replace_rules1)
    
        for item in results_ruru1:
            for k, v in item.items():
                original_text = k
                corrected_text_re = v
                reason_type = "用語の統一"

                comment = f"{reason_type} {original_text} → {corrected_text_re}"

            corrections.append({
                "page": pageNumber,
                "original_text": extract_text(input_text, original_text),# original_text,
                "comment": comment,
                "reason_type": reason_type,
                "check_point": reason_type,
                "locations": [],
                "intgr": False,  
            })

# 英略词，only 地政学
    if fund_type == 'private':
        results_ruru2 = opt_check_ruru2(input_text, replace_rules2)
    
        for item in results_ruru2:
            for k, v in item.items():
                original_text = k  # original_text save to AI
                corrected_text_re = v  # value(v)을 corrected_text_re save to AI（人工知能）
                reason_type = "用語の統一"

                comment = f"{reason_type} {original_text} → {corrected_text_re}"

            corrections.append({
                "page": pageNumber,
                "original_text": extract_text(input_text, original_text),# original_text,
                "comment": comment,
                "reason_type": reason_type,
                "check_point": reason_type,
                "locations": [],
                "intgr": False,
            })

# -----------------
    if fund_type == 'public':
        word_re = regcheck.findall(r"外国人投資家からの資金流入|外国人投資家の資金流出|加速", input_text)
        for word_result in word_re:
            corrections.append({
                "page": pageNumber,
                "original_text": word_result,
                "comment": f"{word_result} → ", #word_result,
                "reason_type": "メッセージの表示",
                "check_point": word_result,
                "locations": [],  
                "intgr": False,  
            })

        day_re = regcheck.findall(r"\d{1,2}[~～]\d{1,2}月期|\d{1,2}月\d{1,2}日[~～]\d{1,2}月\d{1,2}日", input_text)
        for day_result in day_re:
            cor_day = day_result.replace("~", "-").replace("～", "-")
            corrections.append({
                "page": pageNumber,
                "original_text": day_result,
                "comment": f"{day_result} → {cor_day}",
                "reason_type": "波ダッシュの修正",
                "check_point": day_result,
                "locations": [],  
                "intgr": False,  
            })

        score_re = regcheck.findall(r"[\d.]+?[～~][\d.]+?[%％]", input_text)
        for score_result in score_re:
            cor_score = score_result.replace("～", "％～").replace("~", "％~")
            corrections.append({
                "page": pageNumber,
                "original_text": score_result,
                "comment": f"{score_result} → {cor_score}",
                "reason_type": "波ダッシュの修正",
                "check_point": score_result,
                "locations": [],  
                "intgr": False,  
            })

    half_re = regcheck.findall(r"\d{2,4}年第[1-4一二三四]四半期", input_text)
    for half_result in half_re:
        half_num = half_result[-3]
        if half_num in ["1", "一"]:
            time_range = "1-3"
        elif half_num in ["2", "二"]:
            time_range = "4-6"
        elif half_num in ["3", "三"]:
            time_range = "7-9"
        else:
            time_range = "10-12"
        cor_half = half_result[: -4] + time_range + half_result[-3:]
        cor_half_len = len(cor_half.split("年", 1)[0])
        if cor_half_len < 4:
            cor_half = "20" + cor_half[cor_half_len - 2:]
        corrections.append({
            "page": pageNumber,
            "original_text": half_result,
            "comment": f"{half_result} → {cor_half}",
            "reason_type": "日付の修正",
            "check_point": half_result,
            "locations": [],  
            "intgr": False,  
        })


#-------------------
    # tenbrend
    if isinstance(tenbrend, list):
        for item in tenbrend:
            if not isinstance(item, dict):
                continue

            old_text = item.get("元組入銘柄解説", "").strip()
            new_text = item.get("新組入銘柄解説", "").strip()

            corrections.append({
                "check_point": "組入銘柄解説",
                "comment": f"{old_text} → {new_text}",
                "intgr": False,
                "locations": [],
                "original_text": new_text[:20],
                "page": pageNumber,
                "reason_type": item.get("分類", "")
            })

#---------------------
# dotfind 句点の追加  句読点  
    for sentence in input_list:
        if "（出所）" in sentence:
            continue

        if not check_fullwidth_period(sentence):
            sentence_split = re.sub(r"\s+$", "", sentence)[-30:]
            corrections.append({
                "check_point": "句点の追加",
                "comment": f"{sentence_split} → {sentence_split}。",
                "intgr": False,
                "locations": [],
                "original_text": sentence_split,
                "page": pageNumber,
                "reason_type": "句点の追加",
            })
#--------------------add--0905--
            # 主語欠落を検知する正規表現 pattern
            # 「〜と示唆した」 の直前に「が|は」などの主語 
            pattern = r"([^、。]+?と示唆した)"

            matches = re.finditer(pattern, input_text)

            for match in matches:
                original_text = match.group(0)
                # 修正案: 「同社が〜ことを示唆した」
                corrected_text_re = f"同社が{match.group(1)}ことを示唆した"
                reason_type = "主語の欠落"

                comment = f"{reason_type}: {original_text} → {corrected_text_re}"

                corrections.append({
                    "page": pageNumber,
                    "original_text": original_text,
                    "comment": comment,
                    "reason_type": reason_type,
                    "check_point": reason_type,
                    "locations": [],   # 必要なら位置情報を追加
                    "intgr": False,
                })
#--------------------
    return corrections

def extract_text(input_text, original_text):
    pattern = rf"{original_text}（[^）]*）|{original_text}"

    match = regcheck.search(pattern, input_text)
    
    if match:
        return match.group(0)
    else:
        return None

def clean_percent_prefix(value: str):
    if not isinstance(value, str):
        return None
    for symbol in ['％', '%', 'ポイント']:
        if symbol in value:
            value = value.split(symbol)[0].strip()
            return f"{value}{symbol}"
    return value.strip()
                
def extract_parts_with_direction(text: str):
    parts = re.split(r'[、。\n]', text)
    
    segments = []

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # pattern = r'[^％%ポイント上下、。\n]*[+-−]?\d+(?:\.\d+)?(?:％|%|ポイント)'
        pattern = r'[^％%、。\n]*[+-−]{0,2}\d+(?:\.\d+)?(?:％|%|ポイント)'
        segments.extend(re.findall(pattern, part))

        # 上下方向
        # direction_match = re.findall(r'(上回りました|下回りました)', part)
        # segments.extend(direction_match)

    return segments

def extract_corrections(corrected_text, input_text,pageNumber):
    corrections = []
    
    # correction span
    pattern_alt = re.compile(
        r'<span.*?>(.*?)<\/span>\s*'
        r'\(<span>提示:\s*(.*?)\s*<s.*?>(.*?)<\/s>\s*→\s*(.*?)<\/span>\)',
        re.DOTALL
    )

    matches = pattern_alt.findall(corrected_text)

    for match in matches:
        original = match[0].strip()
        reason = match[2].strip()
        reason_type = match[1].strip()
        corrected = match[3].strip()

        comment = f"{reason} → {corrected}" if corrected else reason
        # "%": "％"
        corrections.append({
            "page": pageNumber,
            "original_text": clean_percent_prefix(reason),
            "comment": comment, # +0.2% → 0.85% , 上升 -> 下落
            "reason_type": reason_type, # ファンドの騰落率，B-xxx

            "check_point": input_text.strip(), # 当月のファンドの騰落率は+0.2%となりました。 A B -xxx
            "locations": [],
            "intgr": True,
        })

    return corrections

    
def add_comments_to_pdf(pdf_bytes, corrections):
    if not isinstance(pdf_bytes, bytes):
        raise ValueError("pdf_bytes must be a bytes object.")
    if not isinstance(corrections, list):
        raise ValueError("corrections must be a list of dictionaries.")
    for correction in corrections:
        if not all(key in correction for key in ["page", "original_text", "comment"]):
            raise ValueError("Each correction must contain 'page', 'original_text', and 'comment' keys.")

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        raise ValueError(f"Invalid PDF file: {str(e)}")

    for idx, correction in enumerate(corrections):
        page_num = correction["page"]
        comment = correction["comment"]
        reason_type = correction["reason_type"]
        locations = correction["locations"][0]
        text_instances = [fitz.Rect(locations["x0"], locations["y0"], locations["x1"], locations["y1"])]
        if int(text_instances[0][0]) == 0:
            continue
        colorSetFill= (1, 1, 0)

        if page_num < 0 or page_num >= len(doc):
            raise ValueError(f"Invalid page number: {page_num}")

        page = doc.load_page(page_num)

        if correction["intgr"]:
            colorSetFill = (172/255, 228/255, 230/255)
        else:
            colorSetFill= (1, 1, 0)

        for rect in text_instances:
            highlight = page.add_rect_annot(rect)
            highlight.set_colors(stroke=None, fill=colorSetFill)
            highlight.set_opacity(0.5)
            highlight.set_info({
                "title": reason_type,  # 可选：显示在注释框标题栏
                "content": comment
            })
            highlight.update()

    output = io.BytesIO()
    doc.save(output)
    output.seek(0)
    doc.close()

    return output


def add_comments_to_excel(excel_bytes, corrections):
    excel_file = io.BytesIO(excel_bytes)
    workbook = load_workbook(excel_file)  # openpyxl

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for correction in corrections:
            if correction["sheet"] == sheet_name:
                cell = correction["cell"]  # : "A1", "B2"
                original_text = correction["original_text"]
                comment = correction["comment"]

                if sheet[cell].value and original_text in str(sheet[cell].value):
                    fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    sheet[cell].fill = fill

                    sheet[cell].comment = Comment(comment, "Author")

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

# pre processing
def normalize_text_for_search(text: str) -> str:
    import re
    replacements = {
        "（": "(", "）": ")", "【": "[", "】": "]",
        "「": "\"", "」": "\"", "『": "\"", "』": "\"",
        "　": " ", "○": "〇", "・": "･", 
        "–": "-", "―": "-", "−": "-", "ー": "-"
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[\u200b\u200c\u200d\u00a0]", "", text)
    return re.sub(r"\s+", " ", text).strip()


#0617 debug
def find_locations_in_pdf(pdf_bytes, corrections):
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        raise ValueError(f"Invalid PDF file: {str(e)}")

    for idx, correction in enumerate(corrections):
        page_num = correction.get("page", 0)
        original_text = correction["original_text"]

        if page_num < 0 or page_num >= len(doc):
            print(f"Warning: Invalid page number: {page_num}")
            continue

        page = doc[page_num]
        found_locations = []
        # pre processing
        text_instances = page.search_for(original_text)

        if not text_instances:
            print(f"Warning: Text '{original_text}' not found on page {page_num}.")
            found_locations.append({
                "x0": 0,
                "y0": 0,
                "x1": 0,
                "y1": 0
            })
            
        else:
            for inst in text_instances:
                rect = fitz.Rect(inst)
                found_locations.append({
                    "x0": rect.x0,
                    "y0": rect.y0,
                    "x1": rect.x1,
                    "y1": rect.y1
                })

        if "locations" not in corrections[idx]:
            corrections[idx]["locations"] = []
        corrections[idx]["locations"].extend(found_locations)

    doc.close()
    return corrections


# db and save blob
PUBLIC_FUND_CONTAINER_NAME = "public_Fund"
PRIVATE_FUND_CONTAINER_NAME = "private_Fund"
CHECKED_PDF_CONTAINER = "checked_pdf"

public_container = get_db_connection(PUBLIC_FUND_CONTAINER_NAME)
private_container = get_db_connection(PRIVATE_FUND_CONTAINER_NAME)
checked_pdf_container = get_db_connection(CHECKED_PDF_CONTAINER)

def upload_to_azure_storage(pdf_bytes, file_name, fund_type):
        """Azure Blob Storage PDF"""
        container_name = PUBLIC_FUND_CONTAINER_NAME if fund_type == 'public' else PRIVATE_FUND_CONTAINER_NAME
        
        container_client = get_storage_container()

        try:
            blob_client = container_client.get_blob_client(file_name)
            blob_client.upload_blob(pdf_bytes, overwrite=True)
            logging.info(f"✅ Blob uploaded: {file_name} to {container_name}")
            return blob_client.url
        except Exception as e:
            logging.error(f"❌ Storage Upload error: {e}")
            return None
def upload_checked_pdf_to_azure_storage(pdf_bytes, file_name, fund_type):
        """Azure Blob Storage PDF"""
        container_name = CHECKED_PDF_CONTAINER

        container_client = get_storage_container()

        try:
            blob_client = container_client.get_blob_client(file_name)
            blob_client.upload_blob(pdf_bytes, overwrite=True)
            logging.info(f"✅ Blob uploaded: {file_name} to {container_name}")
            return blob_client.url
        except Exception as e:
            logging.error(f"❌ Storage Upload error: {e}")
            return None
def download_checked_pdf_from_azure_storage(file_name: str, fund_type: str = None) -> bytes:
    """
    从 Azure Blob Storage 下载 PDF
    :param file_name: 文件名，例如 "a_checked.pdf"
    :param fund_type: 公募或者私募
    :return: PDF 文件的字节流（bytes），失败时返回 None
    """
    container_name = CHECKED_PDF_CONTAINER
    container_client = get_storage_container()

    try:
        blob_client = container_client.get_blob_client(file_name)
        # 下载 blob 到内存
        download_stream = blob_client.download_blob()
        pdf_bytes = download_stream.readall()
        logging.info(f"📥 Blob downloaded: {file_name} from {container_name}")
        return pdf_bytes
    except Exception as e:
        logging.error(f"❌ Storage Download error: {e}")
        return None

def save_to_cosmos(file_name, response_data, link_url, fund_type, upload_type='', comment_type='',icon=''):
    """Cosmos DB Save"""
    # Cosmos DB 连接
    container = public_container if fund_type == 'public' else private_container

    # match = re.search(r'(\d{0,}(?:-\d+)?_M\d{4})', file_name)
    # if match:
    #     file_id = match.group(1)
    # else:
    #     file_id = file_name

    item = {
        'id': file_name,
        'fileName': file_name,
        'result': response_data,
        'link': link_url,
        'updateTime': datetime.utcnow().isoformat(),
        'status': "issue", 
        'readStatus': "unread",
        'icon': icon,
    }
    if upload_type:
        item.update(upload_type=upload_type)
    if comment_type:
        item.update(comment_type=comment_type)


    try:
        existing_item = list(container.query_items(
                query="SELECT * FROM c WHERE c.id = @id",
                parameters=[{"name": "@id", "value": file_name}],
                enable_cross_partition_query=True
            ))

        if not existing_item:
                container.create_item(body=item)
                logging.info(f"✅ Cosmos DB は保存されています: {file_name}")
        else:
            existing_item[0].update(item)
            container.upsert_item(existing_item[0])

            logging.info(f"🔄 Cosmos DB 更新完了: {file_name}")
                
    except CosmosHttpResponseError as e:
        logging.error(f"❌Cosmos DB save error: {e}")

def save_checked_pdf_cosmos(file_name, response_data, link_url, fund_type, upload_type='', comment_type='',icon=''):
    """Cosmos DB Save"""
    # Cosmos DB 连接
    container = 'checked_pdf'

    item = {
        'id': file_name,
        'fileName': file_name,
        'result': response_data,
        'link': link_url,
        'updateTime': datetime.utcnow().isoformat(),
        'status': "checked", 
        'readStatus': "unread",
        'icon': icon,
    }
    if upload_type:
        item.update(upload_type=upload_type)
    if comment_type:
        item.update(comment_type=comment_type)


    try:
        existing_item = list(container.query_items(
                query="SELECT * FROM c WHERE c.id = @id",
                parameters=[{"name": "@id", "value": file_name}],
                enable_cross_partition_query=True
            ))

        if not existing_item:
                container.create_item(body=item)
                logging.info(f"✅ Cosmos DB は保存されています: {file_name}")
        else:
            existing_item[0].update(item)
            container.upsert_item(existing_item[0])

            logging.info(f"🔄 Cosmos DB 更新完了: {file_name}")
                
    except CosmosHttpResponseError as e:
        logging.error(f"❌Cosmos DB save error: {e}")

@app.route('/api/file_status', methods=['POST'])
def get_file_status():
    data = request.json
    fund_type = data.get("fund_type", "public_Fund")
    file_name = data.get("file_name")
    container = public_container if fund_type == 'public_Fund' else private_container
    if file_name:
        query = f"SELECT * FROM c WHERE c.fileName = '{file_name}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        if items:
            return jsonify({"success": True, "status": True}), 200
    return jsonify({"success": True, "status": False}), 200


@app.route('/api/download_checked_pdf', methods=['POST'])
def download_checked_pdf():
    try:
        data = request.json
        fund_type = data.get("fund_type", "public_Fund")
        file_name = data.get("file_name")
        root, ext = os.path.splitext(file_name)
        if ext.lower() == ".pdf":
            file_name = root + "_checked" + ext
        container = get_db_connection(CHECKED_PDF_CONTAINER)

        query = f"SELECT link_url FROM c WHERE c.file_name = '{file_name}' AND c.fund_type = '{fund_type}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))
        if items:
            link_url = items[0].get("link_url")
            return jsonify({"success": True, "status": True,"link_url":link_url}), 200
    


    except Exception as e:
        logging.error(f"❌ Error in write_checked_pdf: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/write_upload_save', methods=['POST'])
def write_upload_save():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ token upload")

        data = request.json
        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        docx_base64 = data.get("docx_bytes", "")
        resutlmap = data.get("original_text", "")
        fund_type = data.get("fund_type", "public")  # 'public'
        file_name_decoding = data.get("file_name", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")
        icon = data.get("icon", "")
        change_flag = data.get("change_flag", "")

        # URL Decoding
        file_name = urllib.parse.unquote(file_name_decoding)

        #---------EXCEL-----------
        if excel_base64:
            try:
                excel_bytes = base64.b64decode(excel_base64)
                response_data = {
                    "success": True,
                    "corrections": []
                }

                # Blob Upload
                link_url = upload_to_azure_storage(excel_bytes, file_name, fund_type)
                if not link_url:
                    return jsonify({"success": False, "error": "Blob upload failed"}), 500

                # Cosmos DB Save
                save_to_cosmos(file_name, response_data, link_url, fund_type, upload_type, comment_type,icon)
                if upload_type != "参照ファイル" and change_flag == "change":
                    container = get_db_connection(FILE_MONITOR_ITEM)
                    container.upsert_item({"id": str(uuid.uuid4()), "file_name": file_name, "flag": "wait",
                                            "link": link_url, "fund_type": fund_type})

            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

            # 3) return xlsx
            return jsonify({
                "success": True,
                "corrections": [],
                "code": 200,
            })
        # ---------PDF -----------
        if pdf_base64:
            try:
                pdf_bytes = base64.b64decode(pdf_base64)

                response_data = {
                    "corrections": []
                }

                # Blob Upload
                link_url = upload_to_azure_storage(pdf_bytes, file_name, fund_type)
                if not link_url:
                    return jsonify({"success": False, "error": "Blob upload failed"}), 500

                # Cosmos DB Save
                save_to_cosmos(file_name, response_data, link_url, fund_type, upload_type, comment_type,icon)

            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500
            
        # ---------DOCX -----------
        if docx_base64:
            try:
                docx_bytes = base64.b64decode(docx_base64)

                response_data = {
                    "corrections": []
                }

                # Blob Upload
                link_url = upload_to_azure_storage(docx_bytes, file_name, fund_type)
                if not link_url:
                    return jsonify({"success": False, "error": "Blob upload failed"}), 500

                # Cosmos DB Save
                save_to_cosmos(file_name, response_data, link_url, fund_type, upload_type, comment_type,icon)
                if upload_type != "参照ファイル" and change_flag == "change":
                    container = get_db_connection(FILE_MONITOR_ITEM)
                    container.upsert_item({"id": str(uuid.uuid4()), "file_name": file_name, "flag": "wait",
                                            "link": link_url, "fund_type": fund_type})

            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500
            
            # return JSON
            return jsonify({
                "success": True,
                "corrections": [],
                "code": 200,
            })

        # return JSON
        return jsonify({
            "success": True,
            "corrections": [],
            "code": 200,
        })

    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def apply_manual_corrections(text, correction_map):
    if text in correction_map:
        result = correction_map[text]
    # for old_text, new_text in correction_map.items():
    #     if old_text in text:
    #         text = text.replace(old_text, new_text)
    return result

def gpt_correct_text(prompt):
    token = token_cache.get_token()
    openai.api_key = token
    print("✅ Token Update SUCCESS")
    
    if not prompt.strip():
        return prompt
    hyogaiKanjiList = []
    
    corrected_text = detect_hyogai_kanji(prompt, hyogaiKanjiList)
    prompt_result = f"""
    You are a professional Japanese text proofreading assistant. Please carefully proofread the following Japanese text and provide corrections in a structured `corrected_map` format.

    **Text to Proofread:**
    - {prompt}

    **Proofreading Requirements:**
    1. **Check for typos and missing characters (誤字脱字がないこと):**
    - Ensure there are no spelling errors or missing characters in the content of the report.
    - If errors are found, add them to the `corrected_map` in the format: "incorrect": "correct".

    2. **Follow the Fund Manager Comment Terminology Guide (ファンドマネージャコメント用語集に沿った記載となっていること):**
    - **Consistent Terminology (表記の統一):**
        - Ensure the writing format of terms is consistent throughout the report.
    - **Prohibited Words and Phrases (禁止（NG）ワード及び文章の注意事項):**
        - Check if any prohibited words or phrases are used in the report and correct them as per the guidelines.
    - **Replaceable and Recommended Terms/Expressions (置き換えが必要な用語/表現、置き換えを推奨する用語/表現):**
        - If you find terms or expressions that need to be replaced, revise them according to the provided rules.
    - **Use of Hiragana (ひらがなを表記するもの):**
        - Ensure the report follows the rules for hiragana notation, replacing content that does not conform to commonly used kanji.
    - **Kana Notation for Non-Standard Kanji (一部かな書き等で表記するもの):**
        - Ensure non-standard kanji are replaced with kana as the standard writing format.
    - **Correct Usage of Okurigana (一般的な送り仮名など):**
        - Ensure the correct usage of okurigana is applied.
    - **English Abbreviations, Loanwords, and Technical Terms (英略語、外来語、専門用語など):**
        - Check if English abbreviations, loanwords, and technical terms are expressed correctly.
    - **Identify and mark any 常用外漢字 (Hyōgai kanji):**
        - Identify and mark any 常用外漢字 (Hyōgai kanji) in the following text.
        - 常用外漢字 refers to Chinese characters that are not included in the [常用漢字表 (Jōyō kanji list)](https://ja.wikipedia.org/wiki/常用漢字), the official list of commonly used kanji in Japan.
        - For any 常用外漢字 identified, mark the character with (常用外漢字) next to it.

        1. 入力された全文（Report Content to Proofread）を **一文字ずつ** 走査してください（単語単位ではなく文字単位の照合です）。
        2. 各文字を、指定された hyogaiKanjiList の文字と **完全一致** で比較してください。
        3. 一致する文字がある場合、その文字を「常用外漢字」として検出してください。
        4. 一致しない文字は、常用漢字として無視してください（誤検出を避けるため）。
        5. 検出された常用外漢字は、以下のフォーマットで注釈をつけて表示してください。

        ---

        【注釈フォーマット】

        次のように、元の漢字に <s> タグと背景色を付け、読みまたは代替語を赤字で示し、その後に理由を添えてください。

        例:
        <span style="color:red;">ぜい</span> (<span>修正理由: 常用外漢字の使用 <s style="background:yellow;color:red">脆</s> → ぜい</span>)

        ---

        【Report Content to Proofread】:
        {corrected_text}

        **1: Typographical Errors (脱字・誤字) Detection**
            -Detect any missing characters (脱字) or misused characters (誤字) that cause unnatural expressions or misinterpretation.

            **Proofreading Requirements**:
            - Detect and correct all genuine missing characters (脱字) or misused characters (誤字) that cause grammatical errors or change the intended meaning.
            - Always detect and correct any incorrect conjugations, misused readings, or wrong kanji/verb usage, even if they superficially look natural.
            - Do not point out stylistic variations, natural auxiliary expressions, or acceptable conjugations unless they are grammatically incorrect.
            - Confirm that each kanji matches the intended meaning precisely.
            - Detect cases where non-verb terms are incorrectly used as if they were verbs.

            - ”と”を脱字しました:
                -Example:
                input: 月間ではほぼ変わらずなりました。
                output:
                <span style="color:red;">月間ではほぼ変わらずなりました。</span> (<span>修正理由: 誤字 <s style="background:yellow;color:red">月間ではほぼ変わらずなりました。</s> → 月間ではほぼ変わらずとなりました。</span>)
            - The kanji '剤' was incorrectly used instead of '済', resulting in a wrong word formation.
                -Example:
                input: 経剤成長
                output:
                <span style="color:red;">経済成長</span> (<span>修正理由: 誤字 <s style="background:yellow;color:red">経剤成長</s> → 経済成長</span>)
            - The verb "遊ぶ" was incorrectly conjugated into a non-existent form "あそぼれる" instead of the correct passive form "あそばれる".
                -Example:
                input: あそぼれますか。
                output:
                <span style="color:red;">あそばれますか。</span> (<span>修正理由: 動詞活用の誤り <s style="background:yellow;color:red">あそぼれますか。</s> → あそばれますか。</span>)
            - "と"を省略したら、「アンダーウェイト」は名詞であり、動詞のように「〜した」と活用するのは文法的に誤りです。
                -Example:
                input: アンダーウェイト（参考指数と比べ低めの投資比率）した
                output:
                <span style="color:red;">アンダーウェイト（参考指数と比べ低めの投資比率）とした</span> (<span>修正理由: 動詞活用の誤り <s style="background:yellow;color:red">アンダーウェイト（参考指数と比べ低めの投資比率）した</s> → アンダーウェイト（参考指数と比べ低めの投資比率）とした</span>)


            **correct Example*:
            - "取り組みし"は自然な連用形表現のため、修正不要'
                -Example:
                input: グローバルで事業を展開する
                output:
                <span style="color:red;">グローバルに事業を展開する</span> (<span>修正理由: 格助詞の誤用 <s style="background:yellow;color:red">グローバルで事業を展開する</s> → グローバルに事業を展開する</span>)

        **2: Punctuation (句読点) Usage Check**
            -Detect missing, excessive, or incorrect use of punctuation marks (、。).

            **Proofreading Requirements**:
            -Ensure sentences correctly end with「。」where appropriate.
            -Avoid redundant commas「、」in unnatural positions.
            -Maintain standard business writing style.

            -Example:
            input: 収益見通しが期待できる企業を中心に投資を行なう方針です
            output:
            <span style="color:red;">収益見通しが期待できる企業を中心に投資を行なう方針です</span> (<span>修正理由: 文末句点の欠如 <s style="background:yellow;color:red"収益見通しが期待できる企業を中心に投資を行なう方針です</s> → 収益見通しが期待できる企業を中心に投資を行なう方針です。</span>)

        **3: Unnatural Spaces (不自然な空白) Detection**
            -Detect unnecessary half-width or full-width spaces within sentences.

            **Proofreading Requirements**:
            -Remove any redundant spaces between words or inside terms.
            -Confirm that spacing follows standard Japanese document conventions.

            -Example:
            input: 送 配電設備
            output:
            <span style="color:red;">送配電設備</span> (<span>修正理由: 不要スペース削除 <s style="background:yellow;color:red">送 配電設備</s> → 送配電設備</span>)

            -Example:
            input: マイクロコントローラーや 関連の複合信号製品
            output:
            <span style="color:red;">マイクロコントローラーや関連の複合信号製品</span> (<span>修正理由: 不要スペース削除 <s style="background:yellow;color:red">マイクロコントローラーや 関連の複合信号製品</s> → マイクロコントローラーや関連の複合信号製品</span>)


        **4: Omission or Misuse of Particles (助詞の省略・誤用) Detection**
            - Detect omissions and misuses of grammatical particles (助詞), especially「の」「を」「に」, that lead to structurally incorrect or unnatural expressions.

            **Proofreading Requirements**:

            - Carefully examine whether all necessary particles—particularly「の」「を」「に」—are correctly used in every sentence.
            - Do not tolerate the omission of any structurally required particle, even if the sentence appears understandable or natural overall.
            - Focus on grammatical correctness, not perceived readability.
            - In long texts, perform sentence-by-sentence proofreading to ensure no required particle is missing at any position.
            - If a particle should be present according to standard Japanese grammar but is omitted, it must be explicitly identified and corrected.

            -Example:
            input: 欧州など市場調査開始して
            output:
            <span style="color:red;">欧州などの市場調査を開始して</span> (<span>修正理由: 連体修飾の助詞省略 <s style="background:yellow;color:red">欧州など市場調査開始して</s> → 欧州などの市場調査を開始して</span>)

            -Example:
            input: ECB（欧州中央銀行）など海外主要中銀による
            output:
            <span style="color:red;">ECB（欧州中央銀行）などの海外主要中銀による</span> (<span>修正理由: 所有格助詞「の」の省略 <s style="background:yellow;color:red">ECB（欧州中央銀行）など海外主要中銀による</s> → ECB（欧州中央銀行）などの海外主要中銀による</span>)

            -Example:
            input: 5000億円
            output:
            <span style="color:red;">5,000億円</span> (<span>修正理由: 金額カンマ区切り <s style="background:yellow;color:red">5000億円</s> → 5,000億円</span>)

        **5: Monetary Unit & Number Format (金額表記・数値フォーマット) Check**

            -Detect mistakes in number formatting, especially monetary values.
            -Proofreading Requirements:
            -Apply comma separator every three digits for numbers over 1,000.
            -Ensure currency units (円、兆円、億円) are correctly used.
            -Standardize half-width characters where needed.

            -Example:
            input: 対応には新たな技術開発や制度改革の必要性が指摘されています。
            output:
            <span style="color:red;">対応は新たな技術開発や制度改革の必要性が指摘されています。</span> (<span>修正理由: 格助詞「には」の誤用 <s style="background:yellow;color:red">対応には新たな技術開発や制度改革の必要性が指摘されています。</s> → 対応は新たな技術開発や制度改革の必要性が指摘されています。</span>)


            **Special Instructions**:
            - Always annotate all detected Hyōgai Kanji.
            - Never replace or modify the character unless explicitly instructed.

        **6: Detection of Misused Enumerative Particle「や」**
            **Proofreading Targets**:
            - Detect inappropriate use of the enumerative particle「や」when it connects elements with different grammatical structures.
            - The particle「や」must only be used to list **nouns or noun phrases** that are grammatically equivalent.
            - If the item following「や」is a **verb phrase**, **adverbial clause**, or a structurally different element, then「や」is incorrect.
            - In such cases, replace「や」with a comma「、」to properly separate clauses or adjust the sentence structure.


        **Special Rules:**
        1. **Do not modify proper nouns (e.g., names of people, places, or organizations) unless they are clearly misspelled.**
            -Exsample:
            ベッセント氏: Since this is correct and not a misspelling, it will not be modified.
        2. **Remove unnecessary or redundant text instead of replacing it with other characters.**
            -Exsample:
            ユーロ圏域内の景気委: Only the redundant character 委 will be removed, and no additional characters like の will be added. The corrected text will be: ユーロ圏域内の景気.
        3. **Preserve spaces between words in the original text unless they are at the beginning or end of the text.**
            -Example:
            input: 月の 前半は米国の 債券利回りの上昇 につれて
            Output: 月の 前半は米国の 債券利回りの上昇 につれて (spaces between words are preserved).
        - **Task**: Header Date Format Validation & Correction  
        - **Target Area**: Date notation in parentheses following "今後運用方針 (Future Policy Decision Basis)"  
        ---
        ### Validation Requirements  
        1. **Full Format Compliance Check**:  
        - Must follow "YYYY年MM月DD日現在" (Year-Month-Day as of)  
        - **Year**: 4-digit number (e.g., 2024)  
        - **Month**: 2-digit (01-12, e.g., April → 04)  
        - **Day**: 2-digit (01-31, e.g., 5th → 05)  
        - **Suffix**: Must end with "現在" (as of)  

        2. **Common Error Pattern Detection**:  
        ❌ "1月0日" → Missing month leading zero + invalid day 0  
        ❌ "2024年4月1日" → Missing month leading zero (should be 04)  
        ❌ "2024年12月" → Missing day value  
        ❌ "2024-04-05現在" → Incorrect separator usage (hyphen/slash)  
        ---
        ### Correction Protocol  
        1. **Leading Zero Enforcement**  
        - Add leading zeros to single-digit months/days (4月 → 04月, 5日 → 05日)  

        2. **Day 0 Handling**  
        - Replace day 0 with YYYYMMDD Date Format  
        - Example: 2024年4月0日 → 2024年04月00日

        3. **Separator Standardization**  
        - Convert hyphens/slashes to CJK characters:  
            `2024/04/05` → `2024年04月05日`  


        4. **Consistency with Report Data Section (レポートのデータ部との整合性確認):**
        - Ensure the textual description in the report is completely consistent with the data section, without any logical or content-related discrepancies.

        5. **Eliminate language fluency(単語間の不要なスペース削除):**
        - Ensure that there are no extra spaces.
            -Example:
            input:景気浮揚が意 識されたことで
            output:景気浮揚が意識されたことで
        
        6.  **Layout and Formatting Rules (レイアウトに関する統一):**
            - **文頭の「○」印と一文字目の間隔を統一:**
                - When a sentence begins with the "○" symbol, ensure the spacing between the symbol and the first character is consistent across the document.
            - **文章の間隔の統一:**
                - If a sentence begins with "○", ensure that the spacing within the frame remains consistent.
            - **上位10銘柄 コメント欄について、枠内に適切に収まっているかチェック:**
                - If the stock commentary contains a large amount of text, confirm whether it fits within the designated frame. 
                - If the ranking changes in the following month, adjust the frame accordingly.
                - **Check point**
                    1. **文字数制限内に収まっているか？**
                    - 1枠あたりの最大文字数を超えていないか？
                    - 適切な行数で収まっているか？

                    2. **次月の順位変動に伴う枠調整の必要性**
                    - 順位が変更されると枠調整が必要なため、調整が必要な箇所を特定

                    3. **枠内に収まらない場合の修正提案**
                    - 必要に応じて、短縮表現や不要な情報の削除を提案
                    - 重要な情報を損なわずに適切にリライト

                    output Format:
                    - **コメントの枠超過チェック**
                    - (枠超過しているか: はい / いいえ)
                    - (超過している場合、オーバーした文字数)

                    - **順位変動による枠調整の必要性**
                    - (調整が必要なコメントリスト)

                    - **修正提案**
                    - (枠内に収めるための修正後のコメント)

            **Standardized Notation (表記の統一):**
            - **基準価額の騰落率:**
                - When there are three decimal places, round off using the round-half-up method to the second decimal place. If there are only two decimal places, keep the value unchanged.

            - **％（パーセント）、カタカナ:**
                - **半角カタカナ → 全角カタカナ**（例:「ｶﾀｶﾅ」→「カタカナ」）
                - **半角記号 → 全角記号**（例:「%」→「％」、「@」→「＠」）
                    Example:
                        input: % ｶﾀｶﾅ 
                        output: ％ カタカナ

            - **数字、アルファベット、「＋」・「－」:**
                - **全角数字・アルファベット → 半角数字・アルファベット**（例:「１２３」→「123」、「ＡＢＣ」→「ABC」）
                - **全角「＋」「－」 → 半角「+」「-」**（例:「＋－」→「+-」
                    Example:
                        input: １２３ ＡＢＣ ｱｲｳ ＋－
                        output: 123 ABC アイウ +-

            - **スペースは変更なし**  

            - **「※」の使用:**
                - 「※」は可能であれば **上付き文字（superscript）※** に変換してください。
                - 出力形式の例:
                - 「重要事項※」 → 「重要事項<sup>※</sup>」

            - **（カッコ書き）:**
                - Parenthetical notes should only be included in their first occurrence in a comment.
                以下の日本語テキストにおいて、カッコ書き（bracket "（ ）"）が適切に使用されているかをチェックしてください。

                **Check point**
                    1. **カッコ書きは、コメントの初出のみに記載されているか？**
                    - 同じカッコ書きが2回以上登場していないか？
                    - 初出ページ以降のコメントにカッコ書きが重複して記載されていないか？

                    2. **ディスクロのページ番号順に従ってルールを適用**
                    - シートの順番ではなく、実際のページ番号を基準にする。

                    3. **例外処理**
                    - 「一部例外ファンドあり」とあるため、例外的にカッコ書きが複数回登場するケースを考慮する。
                    - 例外として認められるケースを判断し、適切に指摘。

                    output Format:
                    - **カッコ書きの初出リスト**（どのページに最初に登場したか）
                    - **重複チェック結果**（どのページで二重記載されているか）
                    - **修正提案**（どのページのカッコ書きを削除すべきか）
                    - **例外ファンドが適用される場合、補足情報**

            - **会計期間の表記:**
                - The use of "～" is prohibited; always use "-".
                - Example: 6～8月期（×） → 6-8月期（○）
            - **年度表記:**
                - Use four-digit notation for years.
                - Make modifications directly in this article and explain the reasons for the modifications.
                - Example: 22年（×） → 2022年（○）
            - **レンジの表記:**
                - Always append "%" when indicating a range.
                - Make modifications directly in this article and explain the reasons for the modifications.
                - Example: -1～0.5%（×） → -1%～0.5%（○）
            - **投資環境の記述:**
                **「先月の投資環境」**の部分で「先月末」の記述が含まれる場合、「前月末」に変更してください。
                Example:
                修正前: 先月末の市場動向を分析すると…
                修正後: 前月末の市場動向を分析すると…

            - **通貨表記の統一:**
                - Standardize currency notation across the document.
                - 日本円は「JPY」または「¥」で統一 円.
                - Exsample: 
                input: ¥100 or JPY 100
                output: 100 円

            **Preferred and Recommended Terminology (置き換えが必要な用語/表現):**
            - **第1四半期:**
                - Ensure the period is clearly stated.
                - Example: 18年第4四半期（×） → 2018年10-12月期（○）
            - **約○％程度:**
                - Do not use "約" (approximately) and "程度" (extent) together. Choose either one.
                - Example: 約○％程度（×） → 約○％ or ○％程度（○）
            - **大手企業表記の明確化**  
            - **「○○大手」** が含まれる場合、文中の **会社名を抽出** し、  
                **「大手○○会社/企業」** の形式に修正する。  
            - **入力例:**  
                - 「大手メーカー/会社/企業」  
                - **出力:** 「大手不動産会社、大手半導体メーカー」  
            - **The actual company name must be found and converted in the article


        **Special Rules:**
        1. **Do not modify proper nouns (e.g., names of people, places, or organizations) unless they are clearly misspelled.**
            -Exsample:
            ベッセント氏: Since this is correct and not a misspelling, it will not be modified.
        2. **Remove unnecessary or redundant text instead of replacing it with other characters.**
            -Exsample:
            ユーロ圏域内の景気委: Only the redundant character 委 will be removed, and no additional characters like の will be added. The corrected text will be: ユーロ圏域内の景気.
        3. **Preserve spaces between words in the original text unless they are at the beginning or end of the text.**
            -Example:
            input: 月の 前半は米国の 債券利回りの上昇 につれて
            Output: 月の 前半は米国の 債券利回りの上昇 につれて (spaces between words are preserved).
        ---

        ### **Correction Rules:**
        1. **Only output the corrected_map dictionary. No explanations or extra text.**
        2. **Only incorrect words and their corrected versions should be included.**
        3. **Do not include full sentence corrections.**
        4. **Ensure the corrected_map output is in valid Python dictionary format.**
        5. **Return only the following structure:**
        
        **Output Format:**
        {{
            "incorrect1": "corrected1",
            "incorrect2": "corrected2"
        }}

        """

    
    # ChatCompletion Call
    response = openai.ChatCompletion.create(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system", "content": "You are a professional Japanese text proofreading assistant."},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=MAX_TOKENS,
        temperature=TEMPERATURE,
        seed=SEED  #seed
    )

    try:
        answer = response['choices'][0]['message']['content']
        corrected_map = parse_gpt_response(answer)
        
        full_corrected = prompt
        for k, v in corrected_map.items():
            full_corrected = full_corrected.replace(k, v)

        dynamic_corrections = detect_corrections(prompt, full_corrected)
        corrected_map.update(dynamic_corrections)

        return {k: v for k, v in corrected_map.items() if k and v and k != v}

    except Exception as e:
        print(f"req error: {e}")
        return {}

def correct_text_box_in_excel(input_bytes,corrected_map):
    # 1)  in-memory zip
    in_memory_zip = zipfile.ZipFile(io.BytesIO(input_bytes), 'r')
    
    # BytesIO
    output_buffer = io.BytesIO()
    new_zip = zipfile.ZipFile(output_buffer, 'w', zipfile.ZIP_DEFLATED)

    for item in in_memory_zip.infolist():
        file_data = in_memory_zip.read(item.filename)

        # 3) drawingN.xml
        if item.filename.startswith("xl/drawings/drawing") and item.filename.endswith(".xml"):
            try:
                tree = ET.fromstring(file_data)
                ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

                # 4) find <a:t> tag 
                for t_element in tree.findall(".//a:t", ns):
                    original_text = t_element.text
                #----------------------------------------------------------------
                    # if original_text:  # None 
                    #     original_text_gpt = gpt_correct_text(original_text)

                    #     if original_text_gpt and original_text_gpt.strip() in corrected_map:
                    #         t_element.text = corrected_map[original_text_gpt.strip()]
                #----------------------------------------------------------------
                    # resultMap = gpt_correct_text(original_text)

                    if original_text in corrected_map:
                        t_element.text = corrected_map[original_text]
                #----------------------------------------------------------------

                file_data = ET.tostring(tree, encoding='utf-8', standalone=False)
                
            except Exception as e:
                print(f"Warning: Parsing {item.filename} failed - {e}")

        new_zip.writestr(item, file_data)

    in_memory_zip.close()
    new_zip.close()
    output_buffer.seek(0)
    return output_buffer.getvalue()

# Excel read -for debug
@app.route("/api/excel_upload", methods=["POST"])
def excel_upload():
    file = request.files["file"]  # XLSX
    original_bytes = file.read()

    corrected_map = {
        "地政学リスク": "地政学的リスク"
    }

    # 2) 수정
    modified_bytes = correct_text_box_in_excel(original_bytes, corrected_map)

    # 3) return xlsx
    return send_file(
        io.BytesIO(modified_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="annotated.xlsx"
    )

# T-STARヘルプ API --for debug
@app.route('/api/prompt_upload', methods=['POST'])
def prompt_upload():
    try:
        data = request.json
        prompt = data.get("input", "")
        original_text = data.get("original_text", "")

        if not prompt:
            return jsonify({"success": False, "error": "No input provided"}), 400
        
        prompt_result = f"""
        Please analyze the provided {original_text} and generate results based on the specified {prompt}.

        **Requirements**:
        1. **Extract relevant information**:
        - Extract only the information that directly answers the {prompt}.
        2. **Process the content**:
        - Process the extracted information to provide a clear and concise response.
        3. **Output in Japanese**:
        - Provide the results in Japanese, strictly based on the {prompt}.
        - Do not include any unrelated information or additional explanations.

        **Output**:
        - The output must be accurate, concise, and fully aligned with the {prompt}.
        - Only provide the response in Japanese.

        **Example**:
        - If the {prompt} is "売上成長率を教えてください", the output should be:
        "2023年の売上成長率は15％です。"

        Ensure the output is accurate, concise, and aligned with the given {prompt} requirements.
        """

        # ChatCompletion Call
        response = openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=[
                {"role": "system", "content": "You are a professional Japanese text proofreading assistant."},
                {"role": "user", "content": prompt_result}
            ],
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
            seed=SEED  # seed
        )
        answer = response['choices'][0]['message']['content'].strip()
        re_answer = remove_code_blocks(answer)

        # return JSON
        return jsonify({
            "success": True,
            "original_text": prompt,
            "corrected_text": re_answer,
            # "corrections": corrections
        })

    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

#------Auto app update API

@app.route('/api/auto_save_cosmos', methods=['POST'])
def auto_save_cosmos():
    try:
        data = request.json
        response_data = data['result']
        link_url = data['link']
        container_name = data['containerName']
        file_name_decoding = data['fileName']

        # URL Decoding
        file_name = urllib.parse.unquote(file_name_decoding)

        container = get_db_connection(container_name)

        item = {
            'id': file_name,
            'fileName': file_name,
            'result': response_data,
            'link': link_url,
            'updateTime': datetime.utcnow().isoformat(),
        }

        existing_item = list(container.query_items(
            query="SELECT * FROM c WHERE c.id = @id",
            parameters=[{"name": "@id", "value": file_name}],
            enable_cross_partition_query=True
        ))

        if not existing_item:
            container.create_item(body=item)
            logging.info(f"✅ Cosmos DB Update Success: {file_name}")
        else:
            existing_id = existing_item[0]['id']
            item['id'] = existing_id
            container.replace_item(item=existing_item[0], body=item)
            logging.info(f"🔄 Cosmos DB update success: {file_name}")

        return jsonify({"success": True, "message": "Data Update Success"}), 200

    except CosmosHttpResponseError as e:
        logging.error(f"❌ Cosmos DB Save error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    except Exception as e:
        logging.error(f"❌ API Save error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    
#----auto app save to blob
@app.route('/api/auto_save_blob', methods=['POST'])
def auto_save_blob():
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "no find file."}), 400

        file = request.files['file']
        blob_name = file.filename
        
        container_client = get_storage_container()

        blob_client = container_client.get_blob_client(blob_name)

        blob_client.upload_blob(file, overwrite=True)

        file_url = blob_client.url
        logging.info(f"✅ Azure Blob Storage Update Success: {blob_name}")

        return jsonify({"success": True, "url": file_url}), 200

    except Exception as e:
        logging.error(f"❌ Azure Blob Storage update error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

#----auto app save log 
@app.route('/api/auto_save_log_cosmos', methods=['POST','PUT'])
def auto_save_log_cosmos():
    """log Cosmos DB Save to API"""
    try:
        container = get_db_connection(APPLOG_CONTAINER_NAME)

        log_data = request.json
        log_by_date = log_data.get("logs", {})

        # ✅ Cosmos DB Save
        for log_id, logs in log_by_date.items():
            existing_logs = list(container.query_items(
                query="SELECT * FROM c WHERE c.id = @log_id",
                parameters=[{"name": "@log_id", "value": log_id}],
                enable_cross_partition_query=True
            ))

            if existing_logs:
                existing_log = existing_logs[0]
                existing_log["logEntries"].extend(logs)
                existing_log["timestamp"] = datetime.utcnow().isoformat(), 
                #update
                container.replace_item(item=existing_log["id"], body=existing_log)
                logging.info(f"🔄 SUCCESS: Update Log Success: {log_id}")
            else:
                log_data = {
                    "id": log_id,  # YYYYMMDD format ID
                    "logEntries": logs,
                    "timestamp": datetime.utcnow().isoformat(),
                }
                #create
                container.create_item(body=log_data)
                logging.info(f"✅ SUCCESS: Save to Log Success: {log_id}")

        return jsonify({"code": 200, "message": "Logs saved successfully."}), 200

    except Exception as e:
        logging.error(f"❌ ERROR: Save Log Error: {e}")
        return jsonify({"code": 500, "message": "Error saving logs."}), 500


# integeration ruru

@app.route('/api/integeration_ruru_cosmos', methods=['POST'])
def integeration_ruru_cosmos():
    try:
        data = request.json

        base_month = data['Base_Month']
        fundType = data['fundType']
        fcode = data['Fcode']
        org_sheet_name = data['Org_SheetName']
        org_title = data['Org_Title']
        org_text = data['Org_Text']
        org_type = data['Org_Type']
        target_sheet_name = data['Target_SheetName']
        target_text = data['Target_Text']
        target_type = data['Target_Type']
        target_condition = data['Target_Condition']
        result = data['result']
        Target_Consult = data['Target_Consult']
        flag = data['flag']
        id = data['id']
        No = data['No']

        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        query = f"SELECT * FROM c WHERE c.Fcode = '{data['Fcode']}' AND c.Base_Month = '{data['Base_Month']}' AND c.fundType = '{data['fundType']}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        common_item = {
            "id": id,
            "No": No,
            "fundType": fundType,
            "Base_Month": base_month,
            "Fcode": fcode,
            "Org_SheetName": org_sheet_name,
            "Org_Title": org_title,
            "Org_Text": org_text,
            "Org_Type": org_type,
            "Target_SheetName": target_sheet_name,
            "Target_Text": target_text,
            "flag": flag,
            "Target_Type": target_type,
            "Target_Condition": target_condition,
            "updateTime": datetime.utcnow().isoformat(),  
        }

        if flag == 'close':
            common_item["result"] = result

        elif flag == 'open':
            common_item["Target_Consult"] = Target_Consult

        item = common_item

        if items:
            # container.upsert_item(item)
            items[0].update(item)
            container.upsert_item(items[0])
            logging.info("✅ Data updated in Cosmos DB successfully.")
            return jsonify({"success": True, "message": "Data updated successfully."}), 200
        else:
            container.upsert_item(item)
            logging.info("✅ Data inserted into Cosmos DB successfully.")
            return jsonify({"success": True, "message": "Data inserted successfully."}), 200

        # if items:
        #     item["id"] = items[0]["id"]
        #     container.replace_item(item=items[0], body=item)
        #     logging.info("✅ Data updated in Cosmos DB successfully.")
        #     return jsonify({"success": True, "message": "Data updated successfully."}), 200
        # else:
        #     container.create_item(body=item)
        #     logging.info("✅ Data inserted into Cosmos DB successfully.")
        #     return jsonify({"success": True, "message": "Data inserted successfully."}), 200

    except Exception as e:
        logging.error(f"❌ Cosmos DB save error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/integeration_ruru_cosmos', methods=['GET'])
def get_integeration_ruru_cosmos():
    # Cosmos DB 连接
    container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

    flag = request.args.get("flag")
    base_month = request.args.get("Base_Month")

    query = "SELECT * FROM c"
    parameters = []

    if flag and base_month:
        query += " WHERE c.flag = @flag AND c.Base_Month = @base_month"
        parameters = [
            {"name": "@flag", "value": flag},
            {"name": "@base_month", "value": base_month}
        ]
    elif flag:
        query += " WHERE c.flag = @flag"
        parameters = [{"name": "@flag", "value": flag}]
    elif base_month:
        query += " WHERE c.Base_Month = @base_month"
        parameters = [{"name": "@base_month", "value": base_month}]

    users = list(container.query_items(
        query=query,
        parameters=parameters,
        enable_cross_partition_query=True
    ))

    response = {
        "code": 200,
        "data": users
    }
    return jsonify(response), 200

# common ruru add logic
def common_ruru_text(text):
    corrections = []
    seen = set()

    # ① ファンド＋ベンチマーク両方 → 超過収益 
    pattern_excess = (
        r"基準価額の騰落率は([+-]?\d+(\.\d+)?)％、"
        r"ベンチマークの騰落率は([+-]?\d+(\.\d+)?)％"
    )
    match = re.search(pattern_excess, text)

    if match:
        fund_return = float(match.group(1))
        benchmark_return = float(match.group(3))

        # round 2
        calculated_excess = round(fund_return - benchmark_return, 2)

        result = {
            "騰落率": fund_return,
            "ベンチマークの騰落率": benchmark_return,
            "超過収益（ポイント差）": calculated_excess,
            "reason": "基準価額とベンチマークの差を計算しました"
        }

        key = str(result)  # dict set duipli
        if key not in seen:
            seen.add(key)
            corrections.append(result)


    else:
        # pass
        # ② 個別パターンチェック
        patterns = {
            # "fund_only": r"月間の基準価額の騰落率は[+-]?\d+(\.\d+)?％",
            "benchmark_only": r"ベンチマークの騰落率は[+-]?\d+(\.\d+)?％",
            "course_multi": r"([Ａ-ＺA-Zぁ-んァ-ン一-龥]+コース)が[+-]?\d+(\.\d+)?％",
            "hedge": r"(為替ヘッジあり|為替ヘッジなし)は[+-]?\d+(\.\d+)?％",
            "currency_type": r"(円投資型|米ドル投資型)の月間騰落率は[+-]?\d+(\.\d+)?％",
            # "global_type": r"【[^】]+】[+-]?\d+(\.\d+)?％",
            # "point_value": r"[+-]?\d+(\.\d+)?ポイント",
            "select_course": r"通貨セレクトコース.*?(上昇|下落)",
            "fund_updown": r"基準価額（分配金再投資）は.*?(上昇|下落)"
        }

        # ファンド型: 当ファンド + ベンチマーク
        pattern_fund = r"当ファンドの月間騰落率.*?ベンチマーク[^。]*?ポイント[^。]"
        # 市況型: 株式市場 + TOPIX
        pattern_market = r"TOPIX（東証株価指数）[^。]*"

        # --- ファンド型 ---
        fund_sentences = re.findall(pattern_fund, text)
        for sentence in fund_sentences:
            for m in re.finditer(r"[^、。]+?(％|ポイント)", sentence):
                extracted = m.group(0).strip()
                if extracted not in seen:
                    seen.add(extracted)
                    corrections.append({"extract": extracted})

        # --- 市況型  ---
        market_sentences = re.findall(pattern_market, text)
        for sentence in market_sentences:
            for m in re.finditer(r"[^、。]+?(％|ポイント)", sentence):
                extracted = m.group(0).strip()
                if extracted not in seen:
                    seen.add(extracted)
                    corrections.append({"extract": extracted})


        # その他のパターン一括抽出
        for name, pat in patterns.items():
            for m in re.finditer(pat, text):
                extracted_other = m.group(0)
                if extracted_other not in seen:
                    seen.add(extracted_other)
                    corrections.append({"extract": extracted_other})

    return corrections

# --- common ruru api
@app.route('/api/common_ruru', methods=['POST'])
def common_ruru():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        input_list = data.get("input", "")
        pdf_base64 = data.get("pdf_bytes", "")
        pageNumber = data.get('pageNumber',0)

        if not input_list:
            return jsonify({"success": False, "error": "No input provided"}), 400
        
        corrections = []
        if isinstance(input_list, list):
            for idx, t in enumerate(input_list, start=1):
                part_result = common_ruru_text(t) 
                for pr in part_result:
                    corrections.append({
                        "page": pageNumber,
                        "original_text": pr.get("extract", t),
                        "comment": f"{pr.get("extract", t)} → ",
                        "reason_type": pr.get("reason", "整合性"),
                        "check_point": pr.get("extract", t),
                        "locations": [], 
                        "intgr": True
                    })
        else:
            part_result = common_ruru_text(input_list)
            for pr in part_result:
                corrections.append({
                    "page": pageNumber,
                    "original_text": input_list,
                    "comment": f"{input_list} → {pr.get('extract', pr.get('超過収益（ポイント差）', ''))}",
                    "reason_type": pr.get("reason", "整合性"),
                    "check_point": pr.get("extract", input_list),
                    "locations": [],
                    "intgr": True
                })
        
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
            find_locations_in_pdf(pdf_bytes, corrections)

            return jsonify({
                "success": True,
                "corrections": corrections,
            })

        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
        
        
    except Exception as e:
        # exception return JSON 
        return jsonify({"success": False, "error": str(e)}), 500

# --- ruru test api

@app.route('/api/ruru_search_db', methods=['POST'])
def ruru_search_db():
    try:
        data = request.json

        fcode = data.get('fcode')
        base_month = data.get('Base_Month')
        fund_type = data.get('fundType', 'private')

        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        query = f"SELECT * FROM c WHERE c.Fcode = '{fcode}' AND c.Base_Month = '{base_month}' AND c.fundType = '{fund_type}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        if items:
            # results = [{"id": item["id"], "result": item["result"],"Org_Text":item["Org_Text"],"Org_Type":item["Org_Type"],"Target_Condition":item["Target_Condition"]} for item in items]
            results = [item if item.get("flag") else {"id": item["id"], "result": item["result"],"Org_Text":item["Org_Text"],"Org_Type":item["Org_Type"],"Target_Condition":item["Target_Condition"]} for item in items]
            return jsonify({"success": True, "data": results}), 200
        else:
            return jsonify({"success": False, "message": "No matching data found in DB."}), 200

    except Exception as e:
        logging.error(f"❌ Error occurred while searching DB: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/refer_operate', methods=['GET'])
def get_rule():
    try:
        data = request.args
        flag = data.get('flag', "")
        fund_type = data.get('fundType', 'private')

        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        query = f"SELECT * FROM c WHERE c.flag = '{flag}' AND c.fundType = '{fund_type}'"
        items = list(container.query_items(query=query, enable_cross_partition_query=True))

        if items:
            items_map = list(map(lambda y: dict(filter(lambda x: not x[0].startswith("_"), y.items())), items))
            return jsonify({"success": True, "data": items_map}), 200
        else:
            return jsonify({"success": False, "message": "No matching data found in DB."}), 404

    except Exception as e:
        logging.error(f"❌ Error occurred while searching DB: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    

@app.route('/api/open_cosmos_data', methods=['POST'])
def get_open_data():
    data = request.json
    f_code = data.get("f_code", "")
    flag = data.get("flag", "open")
    base_month = data.get("base_month", "M2411")
    query = f"SELECT * FROM c WHERE c.flag = '{flag}' AND c.Base_Month = '{base_month}' and c.Fcode = '{f_code}'"
    items = list(integeration_container.query_items(query=query, enable_cross_partition_query=True))

    if items:
        return jsonify({"success": True, "data": items}), 200
    else:
        return jsonify({"success": False, "message": "No matching data found in DB."}), 200

@app.route('/api/save_cosmos_data', methods=['POST'])
def save_open_data():
    data = request.json
    item = data.get("item")

    if item:
        integeration_container.upsert_item(item)
        return jsonify({"success": True}), 200
    else:
        return jsonify({"success": False}), 200


@app.route('/api/refer_operate', methods=['POST'])
def insert_rule():
    try:
        data = request.json

        base_month = data.get('Base_Month', '')
        fund_type = data.get('fundType', '')
        fcode = data.get('Fcode', '')
        org_sheet_name = data.get('Org_SheetName', '')
        org_title = data.get('Org_Title', '')
        org_text = data.get('Org_Text', '')
        org_type = data.get('Org_Type', '')
        target_sheet_name = data.get('Target_SheetName', '')
        target_title = data.get('Target_Title', '')
        target_text = data.get('Target_Text', '')
        target_type = data.get('Target_Type', '')
        target_condition = data.get('Target_Condition', '')
        target_consult = data.get('Target_Consult', '')
        id = str(uuid.uuid4())

        container = get_db_connection(INTEGERATION_RURU_CONTAINER_NAME)

        item = {
            "id": id,
            "No": id,
            "fundType": fund_type,
            "Base_Month": base_month,
            "Fcode": fcode,
            "Org_SheetName": org_sheet_name,
            "Org_Title": org_title,
            "Org_Text": org_text,
            "Org_Type": org_type,
            "Target_SheetName": target_sheet_name,
            "Target_Text": target_text,
            "Target_Title": target_title,
            "Target_Type": target_type,
            "Target_Condition": target_condition,
            "Target_Consult": target_consult,
            "flag": "open",
            "updateTime": datetime.now().isoformat()
        }
        container.upsert_item(item)
        return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

async def get_original(input_data, org_text, file_name="", target_text=""):
    dt = [
        "文章から原文に類似したテキストを抽出してください",
        "出力は以下のJSON形式でお願いします:",
        "- [{'target': '[抽出されたテキスト:]'}]",
        "- 類似したものがない場合は、空の文字列を返してください。",
        "抽出ルール：",
        "- 原文と「語順・文構造・文法パターン」が高く一致している文は、たとえ語句（名詞や主語など）が一部違っていても、**必ず抽出してください**。",
        "- 類似度が50%以上の文をすべて抽出してください。",
        "- **原文と構造が似ている文も見落とさずに抽出してください。**",
        "- **抽出する文が原文の言い換え・文型パターンの共通性がある場合、キーワードの違いがあっても対象に含めてください。**",
        "- 最も類似した一文だけを返さず、条件を満たすすべての文を必ず抽出してください。",

        f"原文:{org_text}\n文章:{input_data}"
    ]
    input_data = "\n".join(dt)

    question = [
        {"role": "system", "content": "あなたはテキスト抽出アシスタントです"},
        {"role": "user", "content": input_data},
        {"role": "user", "content": input_data}
    ]
    response = await openai.ChatCompletion.acreate(
        deployment_id=deployment_id,  # Deploy Name
        messages=question,
        max_tokens=MAX_TOKENS,
        temperature=TEMPERATURE,
        seed=SEED  # seed
    )
    answer = response['choices'][0]['message']['content'].strip().replace("`", "").replace("json", "", 1)

    src_score = 0.5
    src_content = ""
    if answer:
        parsed_data = ast.literal_eval(answer)
        for once in parsed_data:
            similar_content = once.get("target")
            if file_name.startswith("180015"):
                if org_text[:4] in similar_content[:6]:
                    src_content = similar_content
                    break
            elif re.search("180332|180358|180359|180360|180344|180345", file_name):
                if org_text[:5] in similar_content[:10]:
                    if target_text in ["セクター別配分", "セクター別寄与度"]:
                        re_content = re.search("(セクター別.*)個別の寄与度", similar_content, re.DOTALL)
                        if re_content:
                            src_content = re_content.groups(1)[0]
                            break
                    elif target_text in ["寄与度", "寄与度＞【上位5銘柄】"]:
                        re_content = re.search("個別の寄与度.*", similar_content, re.DOTALL)
                        if re_content:
                            src_content = re_content.group()
                            break
            elif re.search("180001|180002|180003|180004|180015|180021|180022|180023", file_name):
                if org_text[1: 5] in similar_content[: 10]:
                    src_content = similar_content
                    break
            elif re.search("140672", file_name):
                if org_text[1: 6] in similar_content[: 10]:
                    src_content = similar_content
                    break
            if similar_content:
                score = SequenceMatcher(None, org_text, similar_content).ratio()
                if score > src_score:
                    src_score = score
                    src_content = similar_content
    return src_content, answer


LOCAL_LINK = "local_link"
@app.route('/api/getaths', methods=['GET'])
def get_local_link():
    try:
        container = get_db_connection(LOCAL_LINK)
        log_data = list(container.query_items(
            query=f"SELECT * FROM c",
            enable_cross_partition_query=True
        ))
        log_map = list(map(lambda y: dict(filter(lambda x: x[1] and not x[0].startswith("_"), y.items())), log_data))
        return jsonify({"success": True, "data": log_map}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 404


@app.route('/api/saveaths', methods=['POST'])
def save_local_link():
    try:
        data = request.json
        commonComment = data.get("commonComment")
        individualCheckPath = data.get("individualCheckPath")
        individualComment = data.get("individualComment")
        individualExcelPath = data.get("individualExcelPath")
        individualPdfPath = data.get("individualPdfPath")
        meigaramaster = data.get("meigaramaster")
        reportData = data.get("reportData")
        simu = data.get("simu")
        resultngPath = data.get("resultngPath")
        resultokPath = data.get("resultokPath")
        fund_type = data.get("fund_type")
        container = get_db_connection(LOCAL_LINK)
        link_data = list(container.query_items(
           query=f"SELECT * FROM c WHERE c.fund_type='{fund_type}'",
            enable_cross_partition_query=True
        ))
        update_data = dict(
                fund_type=fund_type,
                commonComment=commonComment,
                individualCheckPath=individualCheckPath,
                individualComment=individualComment,
                individualExcelPath=individualExcelPath,
                individualPdfPath=individualPdfPath,
                meigaramaster=meigaramaster,
                reportData=reportData,
                simu=simu,
                resultngPath=resultngPath,
                resultokPath=resultokPath
        )

        if not link_data:
            update_data.update(id=str(uuid.uuid4()))
            container.upsert_item(update_data)
        else:
            effective_data = dict(filter(lambda x: x[1] is not None, update_data.items()))
            link_data[0].update(effective_data)
            container.upsert_item(link_data[0])
        return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 404


@app.route('/api/log_operate')
def get_log():
    try:
        # 1)  page=1, size=15
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 15))
        file_name = request.args.get('fileName', "")
        log_controller = get_db_connection(LOG_RECORD_CONTAINER_NAME)
        offset = (page - 1) * size
        if file_name:
            file_query = f"SELECT * FROM c WHERE CONTAINS(c.fileName, '{file_name}') OFFSET {offset} LIMIT {size}"
            total_file = list(log_controller.query_items(
                query=file_query,
                enable_cross_partition_query=True
            ))
            name_count = f"SELECT VALUE COUNT(1) FROM c WHERE CONTAINS(c.fileName, '{file_name}')"
            count_result = list(log_controller.query_items(
                query=name_count,
                enable_cross_partition_query=True
            ))[0]
            return jsonify({
                "success": True,
                "data": total_file,
                "total": count_result

            }), 200
        count_query = "SELECT VALUE COUNT(1) FROM c"
        total_count = list(log_controller.query_items(
            query=count_query,
            enable_cross_partition_query=True
        ))[0]

        query = f"""
                SELECT * FROM c
                ORDER BY c.created_at DESC
                OFFSET {offset} LIMIT {size}
                """
        log_data = list(log_controller.query_items(
            query=query,
            enable_cross_partition_query=True
        ))

        log_map = list(map(lambda y: dict(filter(lambda x: x[1] and not x[0].startswith("_"), y.items())), log_data))

        return jsonify({
            "success": True,
            "data": log_map,
            "total": total_count
        }), 200
    
        # return jsonify({"success": True, "data": log_map}), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/check_file', methods=['POST'])
def check_file_statue():
    try:
        data = request.json
        file_name = data.get("file_name")
        fund_type = data.get("fund_type")
        comment_type = data.get("comment_type")
        upload_type = data.get("upload_type", "")
        container = get_db_connection(FILE_MONITOR_ITEM)
        file_data = list(container.query_items(
            query=f"SELECT * FROM u WHERE u.file_name = '{file_name}'",
            enable_cross_partition_query=True
        ))
        if file_data:
            file_info = file_data[0]
            if file_info.get("flag") == "success":
                pdf_name = re.sub(r"\.(xlsx|xlsm|xls|docx|doc)", ".pdf", file_name)
                result = {
                    "corrections": file_info.get("corrections", [])
                }
                is_url = file_info.get("link", "")
                link_url = re.sub(r"\.(xlsx|xlsm|xls|docx|doc)", ".pdf", is_url)
                save_to_cosmos(pdf_name, result, link_url, fund_type, comment_type=comment_type, upload_type=upload_type)
                return jsonify({"success": True}), 200
        return jsonify({"success": False}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 303


@app.route('/api/file_status_update', methods=['POST'])
def file_update():
    try:
        data = request.json
        id = data.get("id", "")
        flag = data.get("flag", "")
        file_name = data.get("file_name", "")
        link_url = data.get("link", "")
        error_space = data.get("error_space", "")
        if id and flag and file_name:
            container = get_db_connection(FILE_MONITOR_ITEM)
            corrections = list(map(lambda x: dict(
                check_point=x.get("original_text"),
                original_text=x.get("original_text"),
                comment=x.get("original_text"),
                intgr=False,
                page=0,
                reason_type=x.get("reason_type"),
                locations=[{"x0": 0, "x1": 0, "y0": 0, "y1": 0}]
                ), error_space))
            container.upsert_item({"id": id, "flag": flag, "file_name": file_name, "link": link_url, "corrections": corrections})
            return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 303


@app.route('/api/file_status_search', methods=['GET'])
def file_search():
    try:
        container = get_db_connection(FILE_MONITOR_ITEM)
        file_data = list(container.query_items(
            query=f"SELECT * FROM u WHERE u.flag = 'wait'",
            enable_cross_partition_query=True
        ))
        if file_data:
            results = []
            for file_info in file_data:
                results.append(file_info)
            return jsonify({"success": True, "data": results}), 200
        else:
            return jsonify({"success": False, "data": []}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 303


@app.route('/api/ruru_ask_gpt_enhance', methods=['POST'])
def integrate_enhance():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update Done")

        data = request.json
        _content = data.get("input", "")
        condition = data.get("Target_Condition", "")
        category = data.get("Org_Type", "")
        consult = data.get("Target_Consult", "")
        base_month = data.get("Base_month", "")
        pageNumber = data.get('pageNumber',0)
        file_name = data.get("file_name", "")
        target_text = data.get("Target_Text", "")

        org_text = data.get("Org_Text", "")
        __answer = ""

        if org_text == "リスク抑制戦略の状況":
            if "リスク抑制戦略の状況" in _content:
                return jsonify({
                    "success": True,
                    "corrections": [{
                        "page": pageNumber,
                        "original_text": "リスク抑制戦略の状況",
                        "check_point": "リスク抑制戦略の状況",
                        "comment": f"リスク抑制戦略の状況 → ",
                        "reason_type":"整合性", 
                        "locations": [{"x0": 0, "x1": 0, "y0": 0, "y1": 0}],
                        "intgr": True, 
                    }]
                })
            else:
                return jsonify({
                    "success": True,
                    "corrections": [{
                        "page": pageNumber,
                        "original_text": "リスク抑制戦略の状況",
                        "check_point": "リスク抑制戦略の状況",
                        "comment": f"リスク抑制戦略の状況 → ",
                        "reason_type": "リスク抑制戦略の状況が存在していません。",  
                        "locations": [{"x0": 0, "x1": 0, "y0": 0, "y1": 0}],
                        "intgr": True,  
                    }]
                })

        elif org_text == "銘柄名1～10":
            content = _content
        elif org_text == "【銘柄名】L’Occitane en Provence（欧州）":
            content_re = re.search("【銘柄名】.{,100}", _content)
            if content_re:
                content = content_re.group()
            else:
                content = ""


        else:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            content, __answer = loop.run_until_complete(get_original(_content, org_text, file_name, target_text))

            if not content:
                return jsonify({
                    "success": True,
                    "answer": __answer,
                    "corrections": []
                })

        pdf_base64 = data.get("pdf_bytes", "")

        file_name_decoding = data.get("file_name", "")

        # URL Decoding
        file_name = urllib.parse.unquote(file_name_decoding)

        if condition:
            result_temp = []
            table_list = condition.split("\n")
            for data in table_list:
                if data:
                    if category in ["比率", "配分"]:
                        re_num = re.search(r"([-\d. ]+)(%|％)", content)
                        if re_num:
                            num = re_num.groups()[0]
                            float_num = len(str(num).split(".")[1]) if "." in num else 0
                            old_data = pd.read_json(StringIO(data))
                            result_temp.append(old_data.applymap(
                                lambda x: (str(round(x * 100, float_num)) + "%" if float_num != 0 else str(
                                    int(round(x * 100, float_num))) + "%")
                                if not pd.isna(x) and isinstance(x, float) else x).to_json(force_ascii=False))
                        else:
                            result_temp.append(pd.read_json(StringIO(data)).to_json(force_ascii=False))
                    else:
                        result_temp.append(pd.read_json(StringIO(data)).to_json(force_ascii=False))
            if len(result_temp) > 1:
                result_data = "\n".join(result_temp)
            else:
                result_data = result_temp[0]
        else:
            result_data = ""

        input_list = [
            "以下の内容に基づいて、原文の記述が正しいかどうかを判断してください", "要件:",
            "- 『参考データ』に該当する情報がない場合、その記述については判断を行わず、「判定対象外」と明記してください。",
            "- 最後に原文の記述が正しいかどうかを明確に判断し、文末に『OK』または『NG』を記載してください",
            f"- **現在の参考データは20{base_month[1:3]}年{base_month[3:]}月の参考データです**",
            f"- 文中に『先月末』『前月末』『○月末』などの表現があっても、現在の参考データ（月）を基準として判断してください",
            f"原文の判断:'{content}'\n参考データ:\n'{result_data}'",
        ]

        if consult:
            input_list.insert(3, consult)
        input_data = "\n".join(input_list)
        question = [
            {"role": "system", "content": "あなたは日本語文書の校正アシスタントです"},
            {"role": "user", "content": input_data}
        ]

        response = openai.ChatCompletion.create(
            deployment_id=deployment_id,  # Deploy Name
            messages=question,
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
            seed=SEED  # seed
        )
        answer = response['choices'][0]['message']['content'].strip()
        if answer:
            dt = [
                "以下の分析結果に基づき、原文中の誤りを抽出してください",
                "出力は以下のJSON形式でお願いします:",
                "- [{'original': '[原文中の誤っている部分:]', 'reason': '[理由:]'}]",
                "- 原文の末尾に「OK」がある場合は、空文字列を返してください",
                f"原文:'{content}'\n分析結果:'{answer}'"
            ]
            summarize = "\n".join(dt)
            _question = [
                {"role": "system", "content": "あなたは日本語文書の校正アシスタントです"},
                {"role": "user", "content": summarize}
            ]
            _response = openai.ChatCompletion.create(
                deployment_id=deployment_id,  # Deploy Name
                messages=_question,
                max_tokens=MAX_TOKENS,
                temperature=TEMPERATURE,
                seed=SEED  # seed
            )
            _answer = _response['choices'][0]['message']['content'].strip().replace("`", "").replace("json", "", 1)
            parsed_data = ast.literal_eval(_answer)
            corrections = []
            if parsed_data:
                for once in parsed_data:
                    error_data = once.get("original", "")
                    reason = once.get("reason", "")
                    corrections.append({
                        "page": pageNumber,
                        "original_text": get_src(error_data, _content).replace("。○","").replace("。◯","").strip().rsplit('\n', 1)[0],
                        "check_point": content,
                        "comment": f"{error_data} → {reason}", #
                        "reason_type":reason, 
                        "locations": [],
                        "intgr": True, 
                    })
            else:
                corrections.append({
                    "page": pageNumber,
                    "original_text": get_src(content, _content).replace("。○","").replace("。◯","").strip().rsplit('\n', 1)[0],
                    "check_point": content,
                    "comment": f"{content} → ",
                    "reason_type": "整合性",  
                    "locations": [],
                    "intgr": True,
                })

            try:
                pdf_bytes = base64.b64decode(pdf_base64)
                find_locations_in_pdf(pdf_bytes, corrections)
                

            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

            return jsonify({
                "success": True,
                "answer": __answer,
                "first_answer": answer,
                "input_data": input_data,
                "corrections": corrections
            })
        else:
            return jsonify({
                "success": True,
                "corrections": [{
                    "page": pageNumber,
                    "original_text": content,
                    "check_point": content,
                    "comment": f"{content} → ",
                    "reason_type":"整合性", 
                    "locations": [{"x0": 0, "x1": 0, "y0": 0, "y1": 0}],
                    "intgr": True, 
                }]  
            })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 200

def extract_or_return(sentence):
    pattern = (
    r"(?P<fund_return>(?:ファンド|基準価額(?:（分配金再投資）)?|基準価額の変動率|基準価額騰落率|騰落率)[の]?(?:変動率|騰落率)?[-−]?\d+\.?\d*％?)?.*?"
    r"(?P<benchmark_return>(?:BM|ベンチマーク|参考指数)[の]?(?:騰落率|変動率)?[-−]?\d+\.?\d*％?)?.*?"
    r"(?P<diff_points>\d+\.?\d*ポイント)?.*?"
    r"(?P<direction>(上回[り]*|下回[り]*))?"
    )

    match = re.search(pattern, sentence)

    extracted = [v for v in match.groupdict().values() if v]

    return extracted if extracted else [sentence]

@app.route('/api/ruru_ask_gpt', methods=['POST'])
def ruru_ask_gpt():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        _input = data.get("input", "")
        result = data.get("result", "")
        orgtext = data.get("Org_Text", "")
        OrgType = data.get("Org_Type", "")
        TargetCondition = data.get("Target_Condition", "")
        pageNumber = data.get('pageNumber',0)
        
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        input, __answer = loop.run_until_complete(get_original(_input, orgtext))
        corrections = []
        pdf_base64 = data.get("pdf_bytes", "")
        if not input:
            dt = [
            "文章から原文に類似したテキストを抽出してください",
            "出力は以下のJSON形式でお願いします:",
            "- {'target': '[抽出されたテキスト:]'}",
            "- 類似したものがない場合は、空の文字列を返してください",
            "- 類似したものが存在する場合は、最も類似度の高いものを抽出してください",

            f"原文:{orgtext}\n文章:{_input}"
            ]
            input_data = "\n".join(dt)

            question = [
                {"role": "system", "content": "あなたはテキスト抽出アシスタントです"},
                {"role": "user", "content": input_data}
            ]
            response = openai.ChatCompletion.create(
                deployment_id=deployment_id,  # Deploy Name
                messages=question,
                max_tokens=MAX_TOKENS,
                temperature=TEMPERATURE,
                seed=SEED  # seed
            )
            _answer = response['choices'][0]['message']['content'].strip().strip().replace("`", "").replace("json", "", 1)
            _parsed_data = ast.literal_eval(_answer)
            _similar = _parsed_data.get("target")

            pattern = r'([ABCDEF]コース.?[+-]?\d+(?:\.\d+)?％|[ABCDEF]コース.?基準価額は(?:下落|上昇)(?:ました)?)'

            matches_list = re.findall(pattern, _similar)
            for re_result in matches_list:
                                
                corrections.append({
                        "page": pageNumber,
                        "original_text": re_result,
                        "check_point": re_result,
                        "comment": f"{re_result} → ", # +0.2% → 0.85% f"{reason} → {corrected}"
                        "reason_type": "整合性",  
                        "locations": [],
                        "intgr": True,
                    })
                
            try:
                pdf_bytes = base64.b64decode(pdf_base64)
                
                find_locations_in_pdf(pdf_bytes, corrections)
                
            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        else:
            if not input:
                return jsonify({"success": False, "error": "No input provided"}), 400
            
            # add the write logic
            dt = [
                "文章から原文に類似したテキストを抽出してください",
                "出力は以下のJSON形式でお願いします:",
                "- [{'original': '[原文中の誤っている部分:]', 'reason': '[理由:]'}]",
                "- 類似したものがない場合は、空の文字列を返してください",
                "- 類似したものが存在する場合は、最も類似度の高いものを抽出してください",

                "あなたは日本の金融レポートを専門とするプロの校正者です。",
                "以下の要約文(Input)を、結果(Result)と比較し、数値や意味に関して正しいかをチェックしてください。",
                "特に次のような誤りがないかを確認してください:",
                "- 騰落率（%）の不一致",
                "- 参考指数（ベンチマーク）の騰落率の不一致",
                "- ポイント",
                "- 上回った／下回ったの方向性の誤り",
                "- 月や期間の不一致",

                f"原文(Input): {input}",
                f"構結果(Result): {result}",
                f"原文種別(original): {OrgType}"
            ]

            input_data = "\n".join(dt)

            question = [
                {"role": "system", "content": "あなたはテキスト抽出アシスタントです"},
                {"role": "user", "content": input_data}
            ]
            response = openai.ChatCompletion.create(
                deployment_id=deployment_id,  # Deploy Name
                messages=question,
                max_tokens=MAX_TOKENS,
                temperature=TEMPERATURE,
                seed=SEED  # seed
            )
            _answer = response['choices'][0]['message']['content'].strip().strip().replace("`", "").replace("json", "", 1)
            _parsed_data = ast.literal_eval(_answer)
            corrections = []
            if _parsed_data:
                for once in _parsed_data:
                    error_data = once.get("original", "")
                    reason = once.get("reason", "")
                    corrections.append({
                        "page": pageNumber,
                        "original_text": clean_percent_prefix(error_data),
                        "check_point": input,
                        "comment": f"{error_data} → {reason}", 
                        "reason_type":reason, 
                        "locations": [],
                        "intgr": True, 
                    })
            else:
                segments = []
                segments= extract_parts_with_direction(input)
                corrections = []
                for part in segments:
                    if part:
                        corrections.append({
                            "page": pageNumber,
                            "original_text": part.strip(),
                            "check_point": input,
                            "comment": f"{part.strip()} → ",
                            "reason_type": "整合性",  
                            "locations": [],
                            "intgr": True,
                        })
                
            if pdf_base64:
                try:
                    pdf_bytes = base64.b64decode(pdf_base64)
                    
                    find_locations_in_pdf(pdf_bytes, corrections)
                    
                except ValueError as e:
                    return jsonify({"success": False, "error": str(e)}), 400
                except Exception as e:
                    return jsonify({"success": False, "error": str(e)}), 500
        
        if not corrections:
        #     match = re.search(r"超過収益[^-+0-9]*([+-]?\d+(?:\.\d+)?)", input)
        #     if match:
        #         value = match.group(1)
        #     else:
        #         value = input
            corrections.append({
                        "page": pageNumber,
                        "original_text": clean_percent_prefix(input),  # 倒数4个字符 [:15]
                        "check_point": input,
                        "comment": f"{input} → ", # +0.2% → 0.85% f"{reason} → {corrected}"
                        "reason_type": "整合性",
                        "locations": [],
                        "intgr": True,
                    })
                
            try:
                pdf_bytes = base64.b64decode(pdf_base64)
                
                find_locations_in_pdf(pdf_bytes, corrections)
                
            except ValueError as e:
                return jsonify({"success": False, "error": str(e)}), 400
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500
            
        # return JSON
        return jsonify({
            "success": True,
            "corrections": corrections,  
            "input": input, 
            "answer": _parsed_data, 
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    
# 611 opt - debug new prompt
def extract_text_from_base64_pdf(pdf_base64: bytes) -> list:
    # Base64 -> PDF bytes
    # pdf_bytes = base64.b64decode(pdf_base64)

    pdf_document = fitz.open(stream=pdf_base64, filetype="pdf")

    text_all = []
    keyword_pages = []
    page_list = []
    for page_num in range(pdf_document.page_count):
        page = pdf_document.load_page(page_num)

        full_text = page.get_text()

        keyword_pos = -1
        for keyword in ["組入銘柄解説", "組入銘柄","組入上位10銘柄の解説"]:
            keyword_pos = full_text.find(keyword)
            if keyword_pos != -1:
                keyword_pages.append(page_num)
                break

        if page_num in keyword_pages:
            # text_all.append(full_text)
            page_text = full_text
            

        else:
            blocks = page.get_text("blocks")  # (x0, y0, x1, y1, "text", block_no, block_type)
            blocks.sort(key=lambda b: b[1])
            page_text = "".join(block[4] for block in blocks)
            # text_all.append(page_text)

        # page_list.append(("".join(text_all), page_num))

        page_list.append((page_text, page_num))
            
    return page_list


# add pre-half logic
half_to_full_map = {
    '%': '％',
    '@': '＠',
    '&': '＆',
    '!': '！',
    '?': '？',
    '#': '＃',
    '$': '＄',
    '(': '（',
    ')': '）',
    '+': '＋'
}
def convert_halfwidth_to_fullwidth_safely(text):
    # (修正理由)
    protected_blocks = {}
    
    def protect_span(match):
        key = f"__PROTECT_{len(protected_blocks)}__"
        protected_blocks[key] = match.group(0)
        return key

    text = re.sub(r'<span[^>]*?>修正理由:.*?</span>\)', protect_span, text)

    def replace_half(match):
        char = match.group(0)
        full = half_to_full_map[char]
        return (
            f'<span style="color:red;">{full}</span> '
            f'(<span>修正理由: 半角記号を全角に統一 '
            f'<s style="background:yellow;color:red">{char}</s> → {full}</span>)'
        )

    pattern = re.compile('|'.join(map(re.escape, half_to_full_map.keys())))
    text = pattern.sub(replace_half, text)

    for key, val in protected_blocks.items():
        text = text.replace(key, val)

    return text

def get_num(num):
    if num:
        num_str = str(num)
        num_len = len(num_str)
        num_list = []
        for i in range(num_len, 0, -3):
            if i - 3 < 0:
                num_r = 0
            else:
                num_r = i - 3
            num_list.insert(0, num_str[num_r: i])
        return ",".join(num_list)
    return ""


def get_src(no_space, src_content):
    content_flag = "".join([i + "☆" for i in no_space])
    content_re = regcheck.escape(content_flag).replace("☆", ".?")
    res = regcheck.search(content_re, src_content, flags=regcheck.DOTALL)
    if res:
        return res.group()
    else:
        return no_space



# async call ,need FE promises
def opt_common(input, prompt_result, pdf_base64, pageNumber, re_list, rule_list, rule1_list, rule3_list,symbol_list):  
    # ChatCompletion Call
    response = openai.ChatCompletion.create(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system", "content": "You are a Japanese text extraction tool capable of accurately extracting the required text."},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=MAX_TOKENS,
        temperature=TEMPERATURE,
        seed=SEED  # seed
    )
    answer = response['choices'][0]['message']['content'].strip().replace("`", "").replace("json", "", 1).replace("\n", "")
    parsed_data = ast.literal_eval(answer)
    combine_corrections = []
    src_corrections = []
    if isinstance(parsed_data, list):
        for re_index, data in enumerate(parsed_data):
            _re_rule = ".{,2}"
            data["original"] = get_src(data["original"], input)
            _original_re = regcheck.search(f"{_re_rule}{regcheck.escape(data["original"])}{_re_rule}", input)
            if _original_re:
                _original_text = _original_re.group()
            else:
                _original_text = data["original"]
            combine_corrections.append({
                "page": pageNumber,
                "original_text": _original_text,
                "comment": f'{_original_text} → {data["correct"]}',
                "reason_type": data["reason"],
                "check_point": _original_text,
                "locations": [],
                "intgr": False,  
            })
            src_corrections.append(f'{data["original"]} → {data["correct"]}')

    if rule_list:
        for rule_result in rule_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": str(rule_result),  
                "comment": f"{str(rule_result)} → 当月の投資配分",
                "reason_type": "誤字脱字",
                "check_point": str(rule_result),
                "locations": [],
                "intgr": False,  
            })

    if re_list:
        for re_result in re_list:
            correct = get_num(re_result)
            combine_corrections.append({
                "page": pageNumber,
                "original_text": str(re_result),  
                "comment": correct,
                "reason_type": "数値千位逗号分隔修正",
                "check_point": str(re_result),
                "locations": [],
                "intgr": False,  
            })

    if rule1_list:
        for rule1_result in rule1_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": rule1_result,  
                "comment": f"{rule1_result} →  ",
                "reason_type": "削除",
                "check_point": rule1_result,
                "locations": [],
                "intgr": False,  
            })

    if rule3_list:
        for rule3_result in rule3_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": rule3_result,  
                "comment": f"{rule3_result} → {rule3_result[1:]}",
                "reason_type": "削除",
                "check_point": rule3_result,
                "locations": [],
                "intgr": False,  
            })

    # if word_list:
    #     for word_result in word_list:
    #         combine_corrections.append({
    #             "page": pageNumber,
    #             "original_text": word_result,  
    #             "comment": f"{word_result} → 値上がりし",
    #             "reason_type": "動詞固定用法",
    #             "check_point": word_result,
    #             "locations": [],
    #             "intgr": False,  
    #         })
    
    # され、下落し
    if symbol_list:
        for symbol_result in symbol_list:
            combine_corrections.append({
                "page": pageNumber,
                "original_text": symbol_result,  
                "comment": f"{symbol_result} → され下落し",
                "reason_type": "読点を削除する",
                "check_point": symbol_result,
                "locations": [],
                "intgr": False,  
            })

    if pdf_base64:
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
            
            find_locations_in_pdf(pdf_bytes, combine_corrections)
            for idx, _comment in enumerate(src_corrections):
                combine_corrections[idx]["comment"] = _comment

        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    # return JSON
    return jsonify({
        "success": True,
        "corrections": combine_corrections,  
        "parsed_data": parsed_data
    })

async def opt_common_wording(file_name,fund_type,input,prompt_result,excel_base64,pdf_base64,resutlmap,upload_type,comment_type,icon,pageNumber):
    # ChatCompletion Call
    response = await openai.ChatCompletion.acreate(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system", "content": "あなたは曖昧な表現を定型語に変換する、厳格な金融校正AIです。出力形式・修正ルールはすべて厳守してください。"},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=MAX_TOKENS,
        temperature=TEMPERATURE,
        seed=SEED  # seed
    )
    answer = response['choices'][0]['message']['content'].strip()
    re_answer = remove_code_blocks(answer)

    # add the write logic
    corrections = find_corrections(re_answer,input,pageNumber)

    corrections_wording = find_corrections_wording(input,pageNumber)

    combine_corrections = corrections + corrections_wording

    if excel_base64:
        try:
            excel_bytes_decoding = base64.b64decode(excel_base64)
            modified_bytes = correct_text_box_in_excel(excel_bytes_decoding,resutlmap)

            # 3) return xlsx
            return send_file(
                io.BytesIO(modified_bytes),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="annotated.xlsx"
            )
        except Exception as e:
            return jsonify({
                "success": False,
                "error": str(e)
            })


    if pdf_base64:
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
            
            find_locations_in_pdf(pdf_bytes, combine_corrections)
            
        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    # return JSON
    return jsonify({
        "success": True,
        "corrections": combine_corrections,  
        "debug_re_answer":re_answer, #610 debug
    })

@app.route('/api/prompt_test', methods=['GET'])
def get_prompt_data():
    prompt_result1 = get_prompt("\"" + "111111111111111111111111111" + "\"")
    prompt_result2 = loop_in_ruru("\"" + "1111111111111111111111111111" + "\"")
    return jsonify(dict(xu=list(prompt_result1), tang=list(prompt_result2)))


@app.route('/api/opt_typo', methods=['POST'])
def opt_typo():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        input = data.get("input", "")

        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        resutlmap = data.get("original_text", "")

        fund_type = data.get("fund_type", "public")  #  'public'
        file_name_decoding = data.get("file_name", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")
        icon = data.get("icon", "")
        pageNumber = data.get('pageNumber',0)

        # URL Decoding
        file_name = urllib.parse.unquote(file_name_decoding)

        if not input:
            return jsonify({"success": False, "error": "No input provided"}), 400
        
        if len(input) < 5:
            return jsonify({"success": True, "corrections": [],})

        prompt_result = get_prompt("\"" + input.replace('\n', '') + "\"")
        async def run_tasks():
            tasks = [handle_result(once) for once in prompt_result]
            return await asyncio.gather(*tasks)

        results = asyncio.run(run_tasks())
        sec_input = "\n".join(results)

        dt = [
            "以下の分析結果に基づき、原文中の誤りを抽出してください。",
            "- 出力結果は毎回同じにしてください（**同じ入力に対して結果が変動しないように**してください）。",
            "- originalには必ず全文や長い文ではなく、**reason_typeで指摘されている最小限の誤りポイント（単語や助詞など）**のみを記載してください。",
            "- 1単語またはごく短いフレーズ単位でoriginalを抽出してください。",
            "- originalはreason_typeの説明に該当する部分のみを抽出してください（例：『など』の後に助詞『の』が必要→originalは必ず『など』）。",
            "- 同じ入力には常に**同じJSON形式の出力**を返してください（推論の揺れを避けてください）。",
            
            "出力は以下のJSON形式でお願いします:",
            "- [{'original': '[原文中の誤っている最小単位の部分]', 'correct': '[正しいテキスト]', 'reason': '[理由:]'}]",
            "- 分析結果に修正部分がある場合は、必ず空のリストを返さないでください。",
            "【例】",
            "reason_type: '年表記は4桁（西暦）に統一'",
            "原文: \"22年の経済成長率は-1～0.5の範囲で推移しました。\"",
            "出力例:",
            "[",
            "  {",
            "    \"original\": \"22年\",",
            "    \"correct\": \"2022年\",",
            "    \"reason\": \"年表記は4桁（西暦）に統一\"",
            "  }",
            "]",
            "reason_type: '例示『など』の後には助詞『の』が必要。『など海外主要中銀』は文法的に不自然なため。'",
            "原文: \"など海外主要中銀による\"",
            "出力例:",
            "[",
            "  {",
            "    \"original\": \"など\",",
            "    \"correct\": \"などの\",",
            "    \"reason\": \"例示『など』の後には助詞『の』が必要\"",
            "  }",
            "]",
            f"原文:'{input}'\n分析結果:'{sec_input}'"

        ]
        sec_prompt = "\n".join(dt)
        re_list = regcheck.findall(r"(\d{4,})[人種万円兆億]", input)
        # word_list = regcheck.findall(r".{,2}値上がり(?!し).{,2}", input)
        rule_list = regcheck.findall(r"当月投資配分", input)
        rule1_list = regcheck.findall(r"【(先月の投資環境|先月の運用経過|今後の運用方針)】", input)
        rule3_list = regcheck.findall(r"-[\d.％]{4,6}下落", input)
        symbol_list = regcheck.findall(r"され、下落し", input)

        _content = opt_common(input, sec_prompt, pdf_base64,pageNumber,re_list,rule_list,rule1_list,rule3_list,symbol_list)
        return _content

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

async def handle_result(prompt_result):
    response = await openai.ChatCompletion.acreate(
        deployment_id=deployment_id,  # Deploy Name
        messages=[
            {"role": "system",
            "content": "You are a professional Japanese business document proofreader specialized in financial and public disclosure materials."},
            {"role": "user", "content": prompt_result}
        ],
        max_tokens=MAX_TOKENS,
        temperature=TEMPERATURE,
        seed=SEED  # seed
    )
    answer = response['choices'][0]['message']['content'].strip()
    return answer

def get_prompt(corrected):
    example_0 = "'original': '月間ではほぼ変わらずなりました。', 'correct': '月間ではほぼ変わらずとなりました。', 'reason': '誤字'"
    example_1 = "'original': '経剤成長', 'correct': '経済成長', 'reason': '誤字'"
    example_10 = "'original': '子供たちは公園で自由にあそぼれますか。', 'correct': '子供たちは公園で自由にあそばれますか。', 'reason': '動詞活用の誤り（「遊ばれる」→「遊ぼれる」）'"
    example_11 = "'original': '我々は新しいプロジェクトに取り組みし、成果を上げました。'"
    example_110 = "'original': 'セクター配分において特化型（物流施設）をアンダーウェイト（参考指数と比べ低めの投資比率）したことなどがプラスに寄与しました。', 'correct': 'セクター配分において特化型（物流施設）をアンダーウェイト（参考指数と比べ低めの投資比率）としたことなどがプラスに寄与しました。', 'reason': '動詞活用の誤り（「遊ばれる」→「遊ぼれる」）'"
    example_111 = "'original': '電子部品や通信機器などの製造・販売を行なうグローバルで事業を展開する電子モジュール・部品メーカー。', 'correct': '電子部品や通信機器などの製造・販売を行なうグローバルに事業を展開する電子モジュール・部品メーカー。', 'reason': 'グローバルは「に」を使用する'"
    # example_2 = "'original': '今後はトランプ次期米大統領が掲げる減税や規制緩和の政策が米景気を押し上げることが、市場の下支えになると考えています。引き続き、FRBによる金融政策や新政権の政策により影響を受けるセクターなどを注視しながら、銘柄を選定して運用を行ないます', 'correct': '今後はトランプ次期米大統領が掲げる減税や規制緩和の政策が米景気を押し上げることが、市場の下支えになると考えています。引き続き、FRBによる金融政策や新政権の政策により影響を受けるセクターなどを注視しながら、銘柄を選定して運用を行ないます。', 'reason': '文法誤用'"
    # example_4 = "'original': '半導体メーカー。マイクロコントローラーや 関連の複合信号製品', 'correct': '半導体メーカー。マイクロコントローラーや関連の複合信号製品', 'reason': '不要スペース削除'"
    example_6 = "'original': '残りについてT-bill（米国財務省短期証券）及び現金等となりました。', 'correct': '残りについてはT-bill（米国財務省短期証券）及び現金等となりました。', 'reason': '助詞「は」の脱落修正'"
    example_60 = "'original': '当月投資配分についてはノムラ・プライベート・クレジット・アクセス・カンパニーに46.4%、', '当月の投資配分についてはノムラ・プライベート・クレジット・アクセス・カンパニーに46.4%、', 'reason': '助詞「の」の脱落修正'"
    example_61 = "'original': '変えること目指している。', 'correct': '変えることを目指している。', 'reason': '助詞「を」の脱落修正'"

    example_70 = "'original': '○月間の基準価額（分配金再投資）の騰落率は、毎月分配型が0.37％、年2回決算型は0.36％の上昇となり、参考指数の騰落率（0.58％の上昇）を下回りました。', 'correct': '○月間の基準価額（分配金再投資）の騰落率は、毎月分配型が0.37％の上昇、年2回決算型は0.36％の上昇となり、参考指数の騰落率（0.58％の上昇）を下回りました。', 'reason': 'Aが◯%、Bは△%の上昇の場合、「の上昇」がBだけにかかっていて、Aにもつけた方がわかりやすいため。'"
    prompt_list = [
        f"""
        **Typographical Errors（脱字・誤字）Detection**
        - Detect only character-level errors that clearly break grammar or meaning.
        **Proofreading Requirements**：
        - Only correct missing or misused characters that clearly break grammar or meaning.
        - Correct obvious verb/kanji errors, even if they seem superficially natural.
        - Do not flag stylistic or acceptable variations unless clearly wrong.
        - Ensure each kanji accurately reflects the intended meaning.
        - Detect cases where non-verb terms are incorrectly used as if they were verbs.
        - Do **not** treat orthographic variants involving okurigana omission or abbreviation（e.g., 書き換え vs 書換え, 読み取る vs 読取る, 取り込む vs 取込）as typographical errors
        -Detect expressions where omitted repeated phrases (e.g., "の上昇", "の低下") may cause ambiguity between multiple items, and suggest repeating the term explicitly for each item to ensure clarity.
        - Do not modify expressions that are grammatically valid and commonly accepted in Japanese, even if alternative phrasing may seem more natural. For example, do not rewrite "中国、米国など" as "中国や米国など" unless required. However, grammatically incorrect forms like "中国、米国など国" must be corrected to "中国、米国などの国".
        
        **missing Example*：
        {example_0}  ”と”を脱字しました
        {example_1}  The kanji '剤' was incorrectly used instead of '済', resulting in a wrong word formation.
        {example_10} The verb "遊ぶ" was incorrectly conjugated into a non-existent form "あそぼれる" instead of the correct passive form "あそばれる".
        {example_110}  "と"を省略したら、「アンダーウェイト」は名詞であり、動詞のように「〜した」と活用するのは文法的に誤りです。
        {example_111}
        **correct Example*：
        {example_11}
        "取り組みし"は自然な連用形表現のため、修正不要'
        {example_70}
        """,
    #   f"""
    #    **Punctuation (句読点) Usage Check**
    #     -Check the sentence-ending punctuation and comma usage only within complete sentences.
    #     **Proofreading Requirements:**
    #     -Only detect missing「。」at the end of grammatically complete sentences.
    #     -If the sentence already ends with「。」, do not suggest any correction.
    #     -Do not flag missing or extra「。」in sentence fragments, headings, bullet points, or intentionally incomplete expressions.
    #     -Check for excessive or missing「、」only within grammatically complete sentences.
    #     -Do not flag cases where comma omission is stylistically natural and grammatically acceptable in Japanese (e.g.,「好感され月間では下落し」).

    #     **Example**：
    #     {example_2}
    #     """,
    #     f"""
    #    **Punctuation (「。」and 「、」) Usage Check**
    #     【Scope】
    #     - Sentences containing both a subject and predicate, ending in a terminal (sentence-final) form
    #     - Only check punctuation within a complete sentence

    #     【Exclusions】
    #     - Sentence fragments, headings, bullet points, or intentionally incomplete expressions
    #     - Conversational or poetic styles where punctuation is intentionally omitted

    #     【Complete Sentence Detection Logic Example】
    #     1. Check if the sentence ends with one of the following terminal forms:
    #     - Verb terminal form (e.g., 「行う」「行いました」「行ないます」)
    #     - Adjective terminal form (e.g., 「高い」「低かった」)
    #     - Noun + auxiliary verb “だ/です” (e.g., 「方針です」「必要だ」)
    #     - Noun + particle “である” (e.g., 「重要である」)
    #     2. If the sentence ends with a comma 「、」, treat it as incomplete
    #     3. If the sentence ends with closing brackets or quotation marks (「」, （）), check the part outside the brackets for terminal form
    #     4. If the sentence ends in a terminal form but lacks 「。」, flag as missing punctuation

    #     【Checks】
    #     1. Sentence-ending punctuation:
    #     - If a complete sentence does not end with 「。」, suggest adding it
    #     - If it already ends with 「。」, no correction is needed
    #     2. Comma usage:
    #     - Excessive: 「、」 appears repeatedly in an unnatural way within the same clause
    #     - Missing: The sentence is too long and hard to read without commas
    #     - Do not flag stylistically natural omissions (e.g., 「好感され月間では下落し」)

    #     **Example**：
    #     {example_2}
    #     """,
        f"""
        **Omission of Particles (助詞の省略・誤用) Detection**
        - Detect omissions of the particles「の」「を」「は」.All other cases are excluded from the check.

        **Example**：
        {example_61}
        {example_6}     
        {example_60}
        """,
        f"""
        **Monetary Unit(金額表記) Check**
        -Proofreading Requirements：
        -Ensure currency units (円、兆円、億円) are correctly used.
        """,
        f"""
        **Incorrect Verb Usage of Compound Noun Phrases（複合名詞の誤動詞化）**
        - Detect grammatically incorrect use of compound noun phrases such as「買い付け」「売り付け」「買い建て」when used in verb forms like「買い付けた」「売り付けた」.
        
        **Proofreading Requirements**:
        - Compound noun phrases such as「値上がり」「買い付け」「売り付け」「買い建て」must not be used as if they were conjugatable verbs.
        - Expressions like「買い付けた」「売り付けた」are grammatically incorrect and must be corrected to「買い付けした」「売り付けした」.
        - Similarly, when followed by a comma such as「〜買い付け、〜」, the correct form is「〜買い付けし、〜」.
        - These terms function as fixed nominal expressions, not inflectable verbs. All such cases must be explicitly identified and corrected.

        """
    ]

    for target_prompt in prompt_list:
        if "助詞の省略" in target_prompt:
            special_word = "- **動詞の連用形や文中の接続助詞前の活用形は正しい表現として認め、文末形などへの変更を求めないこと。**"
        else:
            special_word = ""
        common_result = f"""
        You are a professional Japanese proofreading assistant specializing in official financial documents.
        あなたは金融機関の公式文書に特化した日本語校正アシスタントです。
        校正の目的は「明らかな誤りのみに限定し、余計な修正を一切行わないこと」です。
    
        以下の校正基準を厳守すること：  
        - 文法的に明確な誤り以外は修正禁止。
        - 意味や機能に問題がない表現には、一切手を加えないこと。
        - 表現の改善提案は不要かつ禁止。
        - あくまで機械的・ルールベースの確認のみ行い、スタイルの好みは介入しないこと。
        - 曖昧なケースや判断に迷う場合は「修正不要」と判断すること。
        {special_word}
        - 修正する場合、必ず文法的に正しく、自然な文であること。
        - 修正は文法・語形・表記の客観的エラーに限る。
        - 原文に明らかな問題がない限り、修正を加えてはならない。
        - 表現の優劣に基づく改変や、「よりよい言い回し」は禁止。
        - 回答は50字以内に制限してください。
        - 送り仮名・常用外漢字・（）の全角／半角などチェック不要。

        **Proofreading Targets：**
        "{corrected}"

        {target_prompt}

        """
        yield common_result



def detect_hyogai_kanji(input_text, hyogaiKanjiList):
    corrected_map = {}
    for char in input_text:
        if char in hyogaiKanjiList:
            # 常用外漢字の読みまたは代替語をここでは仮に「？」としています。
            # 実際には、文脈に応じて適切な読みや代替語を設定する必要があります。
            replacement = f"<span style=\"color:red;\">?</span> (<span>修正理由: 常用外漢字の使用 <s style=\"background:yellow;color:red\">{char}</s> → ?</span>)"
            corrected_map[char] = replacement
            input_text = input_text.replace(char, replacement) # 逐次的に置換

    return input_text

@app.route('/api/opt_kanji', methods=['POST'])
def opt_kanji():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json
        input = data.get("full_text", "") # kanji api need full text
        input_list = data.get("input", "") # kanji api need full text

        pdf_base64 = data.get("pdf_bytes", "")
        excel_base64 = data.get("excel_bytes", "")
        resutlmap = data.get("original_text", "")

        fund_type = data.get("fund_type", "public")  #  'public'
        file_name_decoding = data.get("file_name", "")
        upload_type = data.get("upload_type", "")
        comment_type = data.get("comment_type", "")
        tenbrend = data.get("tenbrend", [])
        icon = data.get("icon", "")
        pageNumber = data.get('pageNumber',0)

        # URL Decoding
        file_name = urllib.parse.unquote(file_name_decoding)

        if not input:
            return jsonify({"success": False, "error": "No input provided"}), 400
        

        corrections = find_corrections_wording(input, pageNumber,tenbrend,fund_type,input_list)
        
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
            find_locations_in_pdf(pdf_bytes, corrections)

            return jsonify({
                "success": True,
                "corrections": corrections,
            })

        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
    
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    
# 2. PDF download endpoint
@app.route('/api/download_pdf/<token>', methods=['GET'])
def download_pdf(token):
    file_name = token if token.lower().endswith('.pdf') else f"{token}.pdf"
    temp_path = os.path.join("/tmp", file_name)

    # temp_path = os.path.join("/tmp", f"{token}.pdf")
    if not os.path.exists(temp_path):
        return jsonify({"error": "File not found1"}), 404
    return send_file(temp_path, mimetype='application/pdf', as_attachment=True, download_name=file_name)



def loop_in_ruru(input):
    ruru_all =[
        {
            "category": "表記の統一 (Standardized Notation)",
            "rule_id": "1.1",
            "description": "基準価額の騰落率に関する表現の統一および数値の四捨五入を行なうこと。指定された表現に厳密に従う。",
            "requirements": [
                {
                    "condition": "騰落率が 0.00％ / 0.0％ / 0％ の場合",
                    "correction": "騰落率は変わらずの代わりに、以下のいずれかに修正する:\n- 基準価額(分配金再投資)は前月末から変わらず\n- 前月末と同程度"
                },
                {
                    "condition": "騰落率の数値が小数第3位まである(例：0.546％)",
                    "correction": "小数第2位で四捨五入(round-half-up)し、0.55%のように修正する"
                },
                {
                    "condition": "ファンドとベンチマーク(参考指数)の騰落率を比較する場合",
                    "correction": "上記の四捨五入処理後の値で比較し、同じ場合は騰落率は同程度となりましたと記述する"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "input": "ファンドの騰落率は0.546％",
                    "output": "'original': '0.546％', 'correct': '0.55%', 'reason': '四捨五入'",
                },
                {
                    "input": "月間の基準価額(分配金再投資)の騰落率は+2.85％で、ベンチマークを0％を上回りました。",
                    "output": "'original': 'ベンチマークを0％を上回りました。', 'correct': 'ベンチマークは同程度となりました。', 'reason': '騰落率が同じ'",
                },
                {
                    "input": "月間の基準価額(分配金再投資)の騰落率は+2.85％で、ベンチマークを0.2ポイントを上回りました。",
                    "output": "'original': 'ベンチマークを0.2ポイントを上回りました。', 'correct': 'ベンチマークは同程度となりました。', 'reason': '騰落率が同じ'",
                },
                {
                    "input": "0.00％となりました",
                    "output": "'original': '0.00％となりました', 'correct': '前月末から変わらず', 'reason': '表記の修正'",
                }
            ]
        },
        {
        "category": "数値記号の統一(Numeric Sign Consistency)",
        "rule_id": "1.2",
        "description": "収益率・騰落率などにおいて、正の数値には明示的に「+」を付与して統一性を保つ。既に「+」「−」が付いているものや、比較的表現で増減が示されている場合は変更しない。",
        "requirements": [
            {
            "condition": "収益率、騰落率などで、正の数値に符号(+)が付いていない場合",
            "correction": "符号(+)を付与する (例：4.04％ → +4.04％)"
            },
            {
            "condition": "すでに「+」や「−」が付いている数値",
            "correction": "変更しない"
            },
            {
            "condition": "『下回った』『上回った』『減少』『増加』など、文脈で増減が明示されている場合",
            "correction": "符号は付けない（文脈により方向が明示されているため）"
            }
        ],
        "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「＋」「−」の明示的統一'",
        "Examples": [
            {
            "input": "○月間の基準価額の騰落率は4.04％",
            "output": "'original': '4.04％', 'correct': '+4.04％', 'reason': '「＋」「−」の明示的統一'"
            },
            {
            "input": "インフレ率は0.05ポイント下回っている",
            "output": "変更しない"
            }
        ],
        "notes": "対象数値は一般的に％ or ポイント が後ろに付く収益や成長値などに限定。整数・小数とも対象(例：5％、0.00％、1.234ポイントなど)。ただし「下回っている」「上回っている」「増加」「減少」など文脈的に方向が明示されている場合は記号不要。文章内に複数該当がある場合もすべて個別に対応する。"
        },
        {
        "category": "表現ルール：『大手』の語順と企業名の一般化",
        "rule_id": "CorrectOoteOrder_And_GeneralizeCompanyNames",
        "description": "このルールは、『大手』という語が文中に使われている場合にのみ適用されます。『○○大手』のように語順が逆転している場合は『大手○○企業』に修正し、かつ企業名が含まれる場合には業種・地域に一般化します。ただし、『大手』という語が含まれない場合は、このルールを適用しないでください。企業名を一律に削除・一般化することは禁止します。",
        "requirements": [
            {
                "condition": "『大手』という語が表現内に含まれており、かつ『○○大手』のように後置されている場合、語順を『大手○○』に修正する。",
                "correction": "例：『ゲーム大手企業』⇒『大手ゲーム企業』"
            },
            {
                "condition": "『大手』が含まれており、かつ特定企業名（例：クレディ・スイスなど）が記載されている場合、企業名を削除して地域や業種に一般化する。",
                "correction": "例：『スイス金融大手クレディ・スイス』⇒『スイスの大手金融グループ』"
            },
            {
                "condition": "『大手』という語が含まれていない場合、このルールは適用しない。企業名のみが記載されている場合（例：アマゾン、任天堂など）は原文のままとし、一般化・語順修正は行わない。",
                "correction": "例：『任天堂』⇒ 修正不要（大手という語がないため）"
            }
        ],
        "output_format": "'original': '誤りのある表現', 'correct': '修正後の表現', 'reason': '修正の理由'",
        "examples": [
            {
                "input": "通信大手が新サービスを発表しました。",
                "output": {
                    "original": "通信大手",
                    "correct": "大手通信会社",
                    "reason": "『大手』は業種（通信）の直前に置く必要があります。"
                }
            },
            {
                "input": "ゲーム大手企業の株価が上昇した。",
                "output": {
                    "original": "ゲーム大手企業",
                    "correct": "大手ゲーム企業",
                    "reason": "『大手』は『ゲーム』の直前に配置するのが適切です。"
                }
            },
            {
                "input": "スイス金融大手クレディ・スイスは経営破綻した。",
                "output": {
                    "original": "スイス金融大手クレディ・スイス",
                    "correct": "スイスの大手金融グループ",
                    "reason": "個別企業名は省略し、『大手』は業種の直前に置きます。"
                }
            },
            {
                "input": "任天堂は新作ゲームを発表した。",
                "output": {
                    "original": "任天堂",
                    "correct": "任天堂",
                    "reason": "『大手』という語が含まれていないため、修正の必要はありません。"
                }
            }
        ]
    },
        {
            "category": "YieldMovementdescription",
            "rule_id": "1.3",
            "description": "When describing the movement of yields (利回り), ensure that the inverse relationship with prices is properly reflected.",
            "requirements": [
                {
                    "condition": "If yields rise, it implies that prices fall.",
                    "correction": ""
                },
                {
                    "condition": "If yields fall, it implies that prices rise.",
                    "correction": ""
                },
                {
                    "condition": "If this inverse relationship is not mentioned where necessary, highlight and prompt for correction.",
                    "correction": ""
                },
                {
                    "condition": "利回りは「上昇(価格は下落)」または「低下(価格は上昇)」と表記。",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "利回りの低下",
                    "Output": "'original': '利回りの低下', 'correct': 'Corrected text', '利回りの低下(価格は上昇)','reason': '利回りと価格の逆相関関係を明記する必要あり'",
                },
                {
                    "Input": "日本10年国債利回りは、月間で上昇しました。",
                    "Output": "'original': '上昇', 'correct': 'Corrected text', '上昇(価格は低下)','reason': '利回りと価格の逆相関関係を明記する必要あり'",
                },
                {
                    "Input": "利回りの上昇",
                    "Output": "'original': '利回りの上昇', 'correct': 'Corrected text', '利回りの上昇(価格は低下)','reason': '利回りと価格の逆相関関係を明記する必要あり'",
                },
                {
                    "Input": "月間では債券利回りは上昇しました。",
                    "Output": "'original': '利回りは上昇', 'correct': 'Corrected text', '利回りは上昇(価格は低下)','reason': '利回りと価格の逆相関関係を明記する必要あり'",
                },
                {
                    "Input": "日本10年国債利回りは、月間で下落しました。",
                    "Output": "'original': '下落', 'correct': '下落(価格は上昇)', 'reason': '利回りと価格の逆相関関係を明記する必要あり'",
                }
            ]
        },
        {
            "category": "Correct Usage Of Teika And Geraku",
            "rule_id": "1.4",
            "description": "When describing changes in yields, prices, or interest rates, apply the following word choice rules strictly",
            "requirements": [
                {
                    "condition": "利回りについての数値変換の場合",
                    "correction": "use 低下 for decline, not 下落."
                },
                {
                    "condition": "価格についての数値変換の場合",
                    "correction": "use 下落 for decline, not 低下."
                },
                {
                    "condition": "金利についての数値変換の場合",
                    "correction": "use 低下 for decline, not 下落."
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "米国債利回りが下落しました。",
                    "Output": "'original': '米国債利回りが下落しました。', 'correct': '米国債利回りが低下しました。', 'reason': '利回りには「低下」を使用'",
                },
                {
                    "Input": "価額が低下しました。",
                    "Output": "'original': '価額が低下しました。', 'correct': '価額が下落しました', 'reason': '価格には「下落」を使用'",
                },
                {
                    "Input": "金利が下落しました。",
                    "Output": "'original': '金利が下落しました。', 'correct': '金利が低下しました。', 'reason': '金利には「低下」を使用'",
                }
            ]
        },
        {
            "category": "ZeroPercentCompositionNotation",
            "rule_id": "1.8",
            "description": "When describing a composition ratio of 0%, use either 「0％程度」 or 「ゼロ％程度」",
            "requirements": [
                {
                    "condition": "When describing a composition ratio of 0%",
                    "correction": "use either 「0％程度」 or 「ゼロ％程度」 direct expressions like just \"0%\" without 「程度」 should be corrected."
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '構成比0％の表記統一'",
            "Example": {
                "Input": "当ファンドの構成比は0％です。",
                "Output": "'original': '0％', 'correct': '0％程度', 'reason': '構成比0％の表記統一'",
            }
        },
        {
            "category": "TerminologyConsistency_Calm",
            "rule_id": "2.0",
            "description": "If the usage does not match the context, correct it according to the appropriate meaning.修正理由: 意味の誤用",
            "requirements": [
                {
                    "condition": "Use 「沈静」 when referring to natural calming down over time.",
                    "correction": ""
                },
                {
                    "condition": "Use 「鎮静」 when referring to intentional or artificial suppression (e.g., medical treatment, intervention).",
                    "correction": ""
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "市場は徐々に鎮静していった。",
                    "Output": "'original': '鎮静', 'correct': '沈静', 'reason': '意味の誤用'",
                },
                {
                    "Input": "医療チームは患者の暴動を沈静させた。",
                    "Output": "'original': '沈静', 'correct': '鎮静', 'reason': '意味の誤用'",
                }
            ]
        },
        {
            "category": "Prohibited_or_Cautioned_Expressions_Rise_and_Decline_Factors",
            "rule_id": "2.8",
            "description": "When '上昇' or '下落' appears, check the paragraph-level context for an explicit causal explanation. If no cause is provided, highlight the word and prompt the user. Do not modify the sentence itself—annotate only. If both rise and fall occurred in different courses, prioritize the one with the larger change. Describing both is also acceptable. (上昇・下落要因の記載漏れに対する警告)",
            "requirements": [
                {
                    "condition": "上昇",
                    "correction": "上昇の要因(背景や理由)を明記してください。"
                },
                {
                    "condition": "下落",
                    "correction": "下落の要因(背景や理由)を明記してください。"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '上昇・下落要因の記載漏れ、メッセージ提示'",
            "Examples": [
                {
                    "Input": "最近、米国経済において賃金上昇と物価高が続いており、景気過熱感が指摘されています。米国のインフレ懸念の高まりを背景に、株式市場は下落しました。金融当局は今後も利上げを続ける見通しです。",
                    "Output": "(エラーなし:前文に要因記載あり)"
                }
            ]
        },
        {
            "category": "Replacement Rules for Verb-Type Expressions(動詞・活用形を含む表現の置き換え)",
            "rule_id": "2.9",
            "description":"Certain terms or expressions require more than simple string or regex-based replacement. These are called dynamically varying expressions, which include but are not limited to: Register-sensitive expressions (e.g., polite/humble language variations) Compound phrases or abbreviations that appear in flexible forms When the term to be replaced is a verb, the system must detect and process all conjugated or inflected forms. Do not use rigid pattern matching. Ensure grammatical accuracy after replacement. In general, all such replacements must be done in a context-sensitive manner, ensuring the result remains grammatically and semantically correct",
            "requirements": [
                {
                    "condition": "～に賭ける to ～を予想して ,日本語の使い型変換を注意すべき",
                    "correction": "～を予想して"
                },
                {
                    "condition": "「横ばい」という表現は、期間中の価格・利回り等の値動きが非常に小さい場合に限定して使用すること。一方で、期間中に一定の変動があったものの、最終的に開始時点と同程度の水準に戻った場合には、「ほぼ変わらず」「同程度となる」などの表現を使用する。誤って「横ばい」と記述すると、値動きがなかったような誤認を与える可能性があるため、事実に基づいた正確な表現選択が求められる。",
                    "correction": "ほぼ変わらず"
                }
            ],
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason text'",
            "Examples": [
                {
                    "Input": "～に賭ける",
                    "Output": "'original': '～に賭ける', 'correct': '～を予想して', 'reason': 'Reason text'",
                },
                {
                    "Input": "当作成期を通してみると債券利回りは横ばいでした。",
                    "Output": "'original': '横ばい', 'correct': 'ほぼ変わらず', 'reason': '期間中に一定の変動幅が確認されており、「横ばい」という表現は実態と合わないため、「ほぼ変わらず」とするのが適切。'",
                }
            ]
        },
        {
            "category": "行って来い ⇒ 「上昇(下落)したのち下落(上昇)」等へ書き換える",
            "rule_id": "3.0",
            "description": "The expression “行って来い” is informal and vague. It must not be used in formal financial documents or reports intended for external audiences.Replace it with a precise description of the price movement, such as: “上昇したのち下落した” 下落したのち上昇した Always use fact-based, objective wording that clearly describes the market movement.",
            "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '「行って来い」は曖昧かつ口語的な表現であり、正式な金融文書では具体的な値動きを明記する必要があります。'",
            "Examples": [
                {
                    "Input": "相場は行って来いの展開となりました。",
                    "Output": "'original': '行って来い', 'correct': '一時上昇したものの、その後下落し、前日と同水準で終了しました。', 'reason': '「行って来い」は曖昧かつ口語的な表現であり、正式な金融文書では具体的な値動きを明記する必要があります。'",
                }
            ]
        }
        # ,{
        #     "category": "主語の欠落チェック",
        #     "rule_id": "4.0",
        #     "description": "日本語の文で、誰が・何がを示す主語が省略されると文が不自然になり、読み手に誤解を与える可能性があります。特に金融文書やレポートでは、主語を明確にすることで責任主体や動作主体が正確に伝わります。主語が欠けている場合は、文脈に基づき「同社」「市場」「決算発表」などを補う形で修正してください。",
        #     "output_format": "'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': '主語が欠落しており、誰が示唆したのかが不明確で不自然です。金融文書では動作の主体を明確にする必要があります。'",
        #     "Examples": [
        #         {
        #             "Input": "オランダ株は大手半導体製造装置メーカーの2025年4－6月期受注額は市場予想を上回ったものの、2026年の成長が期待できないと示唆したことなどを背景に下落しました。",
        #             "Output": "'original': '2026年の成長が期待できないと示唆した', 'correct': '同社が2026年の成長が期待できないことを示唆した', 'reason': '主語が欠落しており、「誰が示唆したのか」が分からないため不自然です。主語を補うことで文意が明確になります。'"
        #         }
        #     ]
        # }
    ]

    for ruru_split in ruru_all:
        result = f"""
        You are a professional Japanese business document proofreader specialized in financial and public disclosure materials. 
        Your task is to carefully and strictly proofread the provided Japanese report based on the detailed rules specified below.

        The proofreading targets include:

        Important:
        
        Each section must be strictly followed without omission.
        You are prohibited from making subjective judgments or skipping steps, even if an error seems minor.
        Always prioritize rule adherence over general readability or aesthetic preference.
        Final Output Requirements:
        Use the specified correction format for each detected error.
        Preserve the original sentence structure and paragraph formatting unless explicitly instructed otherwise.
        If no corrections are needed for a section, explicitly state "No errors detected" (検出された誤りなし).
        Follow all instructions strictly and proceed only according to the rules provided.:
        
        Do not correct or modify kana orthography variations (e.g., 「行なう」 vs 「行う」), unless explicitly instructed.
        Do not apply standardization unless listed in the rules.
        
        **Report Content to Proofread:**
        {input}

        **Proofreading Requirements:**
        {ruru_split}

        **Output Requirements:**
        1. **Return only structured correction results as a Python-style list of dictionaries:**
        - Format:
            'original': 'Incorrect text', 'correct': 'Corrected text', 'reason': 'Reason for correction'
        - Example:
            'original': '月間ではほぼ変わらずなりました。', 'correct': '月間ではほぼ変わらずとなりました。', 'reason': '誤字'
        
        2. **Each dictionary must include:**
            - 'original': the original incorrect text
            - 'correct': the corrected text
            - 'reason': a concise explanation for the correction
        3. **Do not include any explanation, HTML tags, or narrative. Only return the data in this dictionary format.**
        4. **Maintain the original document structure internally during processing, but the output should only contain corrections in the required format.**
    
        """
        yield result

@app.route('/api/opt_wording', methods=['POST'])
def opt_wording():
    try:
        token = token_cache.get_token()
        openai.api_key = token
        print("✅ Token Update SUCCESS")
        
        data = request.json

        def convert_fullwidth_to_halfwidth(text):
            return text.replace('（', '(').replace('）', ')')
        
        # input = data.get("input", "")
        input = convert_fullwidth_to_halfwidth(data.get("input", ""))

        pdf_base64 = data.get("pdf_bytes", "")

        fund_type = data.get("fund_type", "public")  #  'public'
        file_name_decoding = data.get("file_name", "")
        icon = data.get("icon", "")
        comment_type = data.get("comment_type", "")
        upload_type = data.get("upload_type", "")
        pageNumber = data.get('pageNumber',0)
        
        # URL Decoding
        file_name = urllib.parse.unquote(file_name_decoding)

        if not input:
            return jsonify({"success": False, "error": "No input provided"}), 400

        prompt_result = loop_in_ruru("\"" + input.replace('\n', '') + "\"")
        async def run_tasks():
            tasks = [handle_result(once) for once in prompt_result]
            return await asyncio.gather(*tasks)

        results = asyncio.run(run_tasks())
        sec_input = "\n".join(results)

        dt = [
            "以下の分析結果に基づき、原文中の誤りを抽出してください",
            "- 出力結果は毎回同じにしてください（**同じ入力に対して結果が変動しないように**してください）。",
            "出力は以下のJSON形式でお願いします:",
            "- [{'original': '[原文中の誤っている部分:]', 'correct': '[誤り部分を正しい部分のテキストに修正:]', 'reason': '[理由:]'}]",
            "- 分析結果が正しい場合は、空のリストを返します",
            "- 同じ入力には常に**同じJSON形式の出力**を返してください（推論の揺れを避けてください）。",
            f"原文:'{input}'\n分析結果:'{sec_input}'"
        ]
        sec_prompt = "\n".join(dt)

        _content = opt_common(input,sec_prompt,pdf_base64,pageNumber,False,False,False,False,False)
        
        return _content
    
        # loop = asyncio.new_event_loop()
        # asyncio.set_event_loop(loop)
        # _content = loop.run_until_complete(opt_common_wording(file_name,fund_type,input,prompt_result,excel_base64,pdf_base64,resutlmap,upload_type,comment_type,icon,pageNumber))
        
        # return _content

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# 820 ,pre-process
def get_words(converted_data, fund_type):
    filter_words = {
        " 「日本プロ": True,
        " 日本取引所": True,
        " 15日に東": True,
        " 日本経済は": True,
        "  (    ": True,
        " )    ": True,
        " 地方主要都": True,
        " )    ": True,
        " 相対的に利": True,
        " ポートフォ": True,
        " 米国と中国": True,
        " 米国の債券": True,
        " ・・・ 景": True,
        " ・・・ ド": True,
        " ・・・ 日": True,
        " ・・・ F": True,
        " ・・・ 長": True,
        " ・・・ ガ": True,
        " 為替 ・・": True,
        " ＜月間の基": True,
        " 長め（地域": True,
        " 現在 )": True,
        " 2025年": True,
        " 愛称：3県": True,
        " 当ファンド": True,
        " 2024年": True,
        " 新たなデジ": True,
        " 直近では、": True,
        " 2025年": True,
        "  (2025": True,
        "現在）": True,
        " ( ": True,
        "現在 ": True,
        "現在)": True,
        " 現在）": True,
        "１": True,
        "ＦＵＮＤＳ": True,
        "ＮＥＸＴ": True,
        "（適格機関": True,
        ")\n": True,
        "クスファンド\nファンドは、値": True,
        "#VALUE!\n野村": True,
        "）ので、基準価額は変動します。": True,
        "（USD）": True,
    }
    result_data = []
    for data in converted_data:
        afterChange = data["comment"].split("→")[-1].strip()
        beforeChange = data["original_text"].strip()
        if data["reason_type"] not in ["常用外漢字の使用", "新規銘柄", "不自然な空白", "同一表現", "異常な色"]:
            if afterChange == beforeChange:
                continue
        if "日付表記として不自然なため" in data["reason_type"]:
            continue
        #---821,fix the error disable

        if "正しい観点" in data["reason_type"]:
            continue
        if "修正不要" in data["reason_type"]:
            continue
        #---821,-----------------
        if beforeChange in ["先月の投資環境", "10", "先月の運用経過", "今後の運用方針", "必ず", "銘柄\n純資産比", "会社（以下「ＪＰＸ」という。", "（USD）"]:
            continue
        if re.search(r"^\d+/\d", beforeChange):
            continue
        if fund_type == "public" and filter_words.get(beforeChange):
            continue
        if re.search(r"現在|詳しくは、|（運用実績、分配金は、|4月のJ-|あります）。|当ファンド|この報告書は、ファンドの運用状|）。|パフォーマンス動向は|当月の投資配分|買い建てし|買い付けしなどをした|贅沢品株の買|などの", afterChange):
            continue
        # 827 fix
        if afterChange == "。" and beforeChange == "":
            continue
        # 903 fix
        if afterChange == "東京エレクトロンは社会効率化、":
            continue
        ignore_list = [
        "○。",
        "〇。",
        "3。",
        "銘柄。",
        "\n◆設定・運用は\n追加型投信／内外／株式\n6/10\n1\n2。",
        "員\n◆設定・運用は\n追加型投信／内外／株式\n6/10\n1\n2。",
        "○\nマンスリーレポート。",
        "○\n9 ARGENX SE-ADR\nアルジェンX。",
        "10 STRYKER CORPORATION\nストライカー。",
        "○\n1 ELI LILLY & CO.\nイーライリリー。",
        "○\n4 DANAHER CORPORATION\nダナハー。",
        "TICALS INC\nアルナイラム・ファーマシューティカルズ。",
        "BBOTT LABORATORIES\nアボットラボラトリーズ。",
        "EALTH GROUP INC\nユナイテッドヘルス・グループ。",
        "NSON & JOHNSON\nジョンソン・エンド・ジョンソン。",
        "CIENTIFIC CORP\nボストン・サイエンティフィック。"
        ]

        if afterChange in ignore_list:
            continue
    
        #---0901,fix the error disable
        if "不自然な空白" in data["reason_type"] and fund_type == "public":
            continue
        
        result_data.append(data)
    return result_data

# ruru_update_save_corrections
@app.route('/api/save_corrections', methods=['POST'])
def save_corrections():
    try:
        data = request.get_json()
        corrections = data.get('corrections','')
        fund_type = data.get("fund_type",'')
        pdf_base64 = data.get("pdf_base64",'')
        file_name_decoding = data.get('file_name','')
        icon = data.get('icon','')

        # URL Decoding
        file_name = urllib.parse.unquote(file_name_decoding)

        # match = re.search(r'(\d{0,}(?:-\d+)?_M\d{4})', file_name)
        # if match:
        #     file_id = match.group(1)
        # else:
        #     file_id = file_name

        if not file_name or not isinstance(corrections, list):
            return jsonify({"success": False, "error": "file_name 和 corrections(list)."}), 400
        
        # container name Setting
        container_name = f"{fund_type}_Fund"
        # 2. Cosmos DB 连接
        container = get_db_connection(container_name)

        existing_item = list(container.query_items(
            query="SELECT * FROM c WHERE c.id = @id",
            parameters=[{"name": "@id", "value": file_name}],
            enable_cross_partition_query=True
        ))

        # corrections
        existing_corrections = []
        if existing_item:
            result = existing_item[0].get("result", {})
            existing_corrections = result.get("corrections", [])

        corrections = get_words(corrections, fund_type)
        final_corrections  = existing_corrections + corrections

        item = {
            'id': file_name,
            'fileName': file_name,
            'icon': icon,
            "result": {
                "corrections": final_corrections
            },
            'updateTime': datetime.utcnow().isoformat(),
        }
        

        if not existing_item:
            container.create_item(body=item)
            logging.info(f"✅ Cosmos DB Update Success: {file_name}")
        else:
            existing_item[0].update(item)
            container.upsert_item(existing_item[0])

            logging.info(f"🔄 Cosmos DB update success: {file_name}")

        if not pdf_base64:
            return jsonify({"success": True, "message": "Data Update Success"}), 200
        
        try:
            pdf_bytes = base64.b64decode(pdf_base64)

            # Save temporarily (in memory or disk), generate a token or filename
            updated_pdf = add_comments_to_pdf(pdf_bytes, corrections)
           
            # store in blob then save in DB
            # temp_file = file_name+'_CHECKED.pdf'

            # rename file_name add suffix _checked
            root, ext = os.path.splitext(file_name)
            if ext.lower() == ".pdf":
                file_name = root + "_checked" + ext
            # ---------PDF -----------
            if pdf_base64:
                try:
                    pdf_bytes = base64.b64decode(pdf_base64)

                    response_data = {
                        "corrections": []
                    }

                    # Blob Upload
                    link_url = upload_checked_pdf_to_azure_storage(pdf_bytes, file_name, fund_type)
                    if not link_url:
                        return jsonify({"success": False, "error": "Blob upload failed"}), 500

                    # Cosmos DB Save
                    save_checked_pdf_cosmos(file_name, response_data, link_url, fund_type, upload_type, comment_type,icon)

                except ValueError as e:
                    return jsonify({"success": False, "error": str(e)}), 400
                except Exception as e:
                    return jsonify({"success": False, "error": str(e)}), 500

            

            # temp_path = os.path.join("/tmp", temp_filename)

            # with open(temp_path, "wb") as f:
            #     f.write(updated_pdf.read())
            #     updated_pdf.seek(0)


            return jsonify({
                "success": True,
                "corrections": corrections,
                "pdf_download_token": temp_filename
            })

        except ValueError as e:
            return jsonify({"success": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

        
    
    except CosmosHttpResponseError as e:
        logging.error(f"Cosmos DB Error: {str(e)}")
        return jsonify({"success": False, "error": "DB Error"}), 500
    except Exception as e:
        logging.error(f"Server error: {str(e)}")
        return jsonify({"success": False, "error": "Server error"}), 500
    
#th for test api

@app.route("/api/integrated_test", methods=["POST"])
async def integrated_test():
    data = request.json
    prompt = data.get("prompt", "")
    input_data = data.get("input_data", "")
    flag_type = data.get("flag_type", "")
    base64_img = data.get("base64_img", "")
    token = token_cache.get_token()
    openai.api_key = token

    if prompt and input_data:
        question = [
            {"role": "system", "content": "あなたは日本語文書の校正アシスタントです"},
            {"role": "user", "content": input_data}
        ]
        if flag_type == "picture":
            question.append({"role": "user", "content": [{"type": "text", "text": input_data},
                            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_img}"}}]})
        else:
            question.append({"role": "user", "content": input_data})

        response = await openai.ChatCompletion.acreate(
            deployment_id=deployment_id,  # Deploy Name
            messages=question,
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
            seed=SEED  # seed
        )
        answer = response['choices'][0]['message']['content'].strip()
        return jsonify({"response_ai": answer})

@app.after_request
def after_request(response):
    try:
        data = request.json
        error_text = ""
        file_name = data.get("file_name", "")
        log_controller = get_db_connection(LOG_RECORD_CONTAINER_NAME)
        response_json = json.loads(response.get_data(as_text=True))
        if response.status_code > 200:
            response_json = json.loads(response.get_data(as_text=True))
            error_text = response_json.get("error", "")
        if file_name:
            if re.search(r"save_corrections|write_upload_save", request.path) or (
                    "check_file" in request.path and response_json.get("success", False)):
                log_info = {
                    "id": str(uuid.uuid4()),
                    "fileName": file_name,
                    "path": request.path,
                    "ip_address": request.remote_addr,
                    "result": "NG" if response.status_code > 200 else "OK",
                    "error_text": error_text,
                    "created_at": datetime.now(UTC).isoformat(),
                }
                log_controller.upsert_item(log_info)

    finally:
        return response


#10铭柄新追加

# PDF 容器路径

def copy_row_style(ws, source_row_idx, target_row_idx):
    """
    将 source_row_idx 的样式复制到 target_row_idx 行（包括字体、边框、填充、对齐方式、数字格式等）
    """
    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row_idx, column=col_idx)
        target_cell = ws.cell(row=target_row_idx, column=col_idx)

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)


def write_wrapped_stock_cell(ws, row, col, stock_value):
    """
    写入 stock 到 Excel 单元格，自动在英日分界处换行并设置 wrap_text。
    """
    if not stock_value:
        return

    # ✅ 在英文(ASCII)和日文之间插入换行
    stock_value = re.sub(r'([a-zA-Z0-9]+)([^\x00-\x7F])', r'\1\n\2', stock_value)

    cell = ws.cell(row=row, column=col, value=stock_value)
    cell.alignment = Alignment(wrap_text=True)


def extract_pdf_table(pdf_path):
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "組入上位10銘柄の解説" in text or "組入上位銘柄の解説" in text:
                tables += page.extract_tables()
    return tables


def extract_pdf_table_special(pdf_path):
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "組入銘柄解説" in text:
                tables += page.extract_tables()
    return tables


# 清除样式
# def clean_text(text):
#     if not text:
#         return ""
#     return re.sub(r'\s+', '', text.replace('\n', '').strip())
def clean_text(text):
    if pd.isna(text):   # Excel 空单元格或 NaN 情况
        return ""
    
    text = str(text)    # 无条件转字符串，防止 float 报错

    # 全角转半角，并去掉换行、空白符（含全角空格）
    text = jaconv.z2h(text, kana=False, digit=True, ascii=True)
    return re.sub(r'[\s\u3000]+', '', text.strip())


# 获取决算月
def get_prev_month_str():
    today = datetime.today()
    prev_month_date = (today.replace(day=1) - timedelta(days=1))
    return prev_month_date.strftime("%Y%m")


# 往10铭柄的履历表里写
def insert_tenbrend_history(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "filename": record["filename"],
            "id": str(uuid.uuid4()),
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "分類": record["分類"],
            "no": record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)


def insert_tenbrend_history42(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元1": record["元1"],
            "新1": record["新1"],
            "元2": record["元2"],
            "新2": record["新2"],
            "元3": record["元3"],
            "新3": record["新3"],
            "元4": record["元4"],
            "新4": record["新4"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元ESG理由": record["元ESG理由"],
            "新ESG理由": record["新ESG理由"],
            "分類": record["分類"],
            "no": record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)


def insert_tenbrend_history41(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "fcode": record["fcode"],
            "filename": record["filename"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元ESG理由": record["元ESG理由"],
            "新ESG理由": record["新ESG理由"],
            "分類": record["分類"],
            "no": record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)


def insert_tenbrend_history5(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元解決すべき社会的課題": record["元解決すべき社会的課題"],
            "新解決すべき社会的課題": record["新解決すべき社会的課題"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元ESG理由": record["元ESG理由"],
            "新ESG理由": record["新ESG理由"],
            "分類": record["分類"],
            "no": record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)


# 🚩读取源文件并更新 diff_rows，往10铭柄的excel中写入
def update_excel_with_diff_rows(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_desc = row["新組入銘柄解説"]
        classify = row["分類"]
        # no = row["no"]
        try:
            no = int(row["no"])
        except (KeyError, TypeError, ValueError):
            no = 0
        months = row["months"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        headers = {cell.value: idx for idx, cell in enumerate(ws[3])}
        stock_col = headers.get("組入銘柄") or headers.get("銘柄")
        desc_col = headers.get("組入銘柄解説") or headers.get("銘柄解説")
        no_col = headers.get("No.")
        months_col = headers.get("決算月")

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        fcode_col = headers.get("Fコード")
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


def find_column_by_keyword(header_row, keywords):
    """
    在 header_row 中查找包含关键字的列索引
    """
    for idx, cell in enumerate(header_row):
        title = str(cell.value).strip() if cell.value else ""
        for key in keywords:
            if key in title:
                return idx
    return None


def update_excel_with_diff_rows4(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_desc = row["新組入銘柄解説"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_esg = row["新最高益更新回数"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]

        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["最高益更新回数"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


def update_excel_with_diff_rows42(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_1 = row["新1"]
        new_2 = row["新2"]
        new_3 = row["新3"]
        new_4 = row["新4"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_desc = row["新組入銘柄解説"]
        new_esg = row["新ESG理由"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[2]

        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["ESGへの取り組みが企業価値向上に資する理由"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)
            ws.cell(row=insert_idx, column=6, value=new_1)
            ws.cell(row=insert_idx, column=7, value=new_2)
            ws.cell(row=insert_idx, column=8, value=new_3)
            ws.cell(row=insert_idx, column=9, value=new_4)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)
            ws.cell(row=target_row_idx, column=6, value=new_1)
            ws.cell(row=target_row_idx, column=7, value=new_2)
            ws.cell(row=target_row_idx, column=8, value=new_3)
            ws.cell(row=target_row_idx, column=9, value=new_4)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


def update_excel_with_diff_rows41(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_desc = row["新組入銘柄解説"]
        new_esg = row["新ESG理由"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]

        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説", "組入発行体解説"])
        esg_col = find_column_by_keyword(header_row, ["ESGへの取り組みが企業価値向上に資する理由",
                                                      "脱炭素社会の実現への貢献と企業評価のポイント"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


def update_excel_with_diff_rows_shang(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        new_desc = row["新組入銘柄解説"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_esg = row["新上場年月"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]

        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["上場年月"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            # ws.cell(row=insert_idx, column=stock_col + 1, value=stock)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


def update_excel_with_diff_rows5(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:

        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_keti = row["新解決すべき社会的課題"]
        new_desc = row["新組入銘柄解説"]
        new_esg = row["新ESG理由"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]

        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄"])
        keti_col = find_column_by_keyword(header_row,
                                          ["解決すべき社会的課題", "業種", "投資分野", "分野", "目指すインパクト"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        esg_col = find_column_by_keyword(header_row, ["ESGへの取り組みが企業価値向上に資する理由",
                                                      "社会的課題の解決と利益成長を両立させるポイント"])
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx

                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


# 10铭柄的check
def check_tenbrend(filename, fund_type):
    try:
        fcode = os.path.basename(filename).split("_")[0]

        if fund_type == 'private':
            TENBREND_CONTAINER_NAME = 'tenbrend_private'
        else:
            TENBREND_CONTAINER_NAME = 'tenbrend'

        container = get_db_connection(TENBREND_CONTAINER_NAME)

        query = "SELECT c.sheetname FROM c WHERE CONTAINS(c.fcode, @fcode)"
        parameters = [{"name": "@fcode", "value": fcode}]
        result = list(container.query_items(query=query, parameters=parameters, enable_cross_partition_query=True))

        if not result:
            return "数据库没有查到数据,没有这个fcode的数据"

        sheetname = result[0]["sheetname"]
        pdf_url = f"{PDF_DIR}/{filename}"

        if sheetname == "過去分整理3列":
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return "PDF下载失败，没有找到pdf"
            if fcode in ['140193', '140386','140565-6','180295-8',"180291-2"]:
                # 将 .pdf 替换为 .xlsx 作为 Excel 文件路径
                excel_url = pdf_url.replace(".pdf", ".xlsx")

                # 下载 Excel 文件内容
                response = requests.get(excel_url)
                if response.status_code != 200:
                    return "Excel下载失败"

                # 转为 BytesIO 对象传给 extract_excel_table3
                excel_file = io.BytesIO(response.content)
                tables = extract_excel_table3(excel_file,fcode)
            elif fcode in ["140675", "140655-6", "140695-6"]:
                tables = extract_pdf_table_special(io.BytesIO(pdf_response.content))
            else:

                tables = extract_pdf_table(io.BytesIO(pdf_response.content))
            if not tables:
                return "PDF中未提取到表格"

            excel_url = f"{PDF_DIR}/10mingbing.xlsx"
            excel_response = requests.get(excel_url)
            if excel_response.status_code != 200:
                return "Excel文件下载失败，不能打开excel"

            wb = load_workbook(filename=io.BytesIO(excel_response.content))
            ws = wb.active

            seen_stocks = set()
            unique_rows = []
            if fcode in ['140193', '140386','140565-6','180295-8']:
                for row in tables:
                    stock = clean_text(row[0])
                    desc = clean_text(row[1])
                    seen_stocks.add(stock)
                    unique_rows.append([stock, desc])

                    if len(unique_rows) >= 10:
                        break                
            else:
                for table in tables:
                    for row in table:
                        if len(row) < 3:
                            continue
                        if (row[1] and ('組入銘柄' in row[1] or '銘柄' in row[1])) and \
                                ((row[2] and '銘柄解説' in row[2]) or (len(row) > 3 and row[3] and '銘柄解説' in row[3])):
                            continue
                        if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                            continue
                        if not row[1]:
                            pdf_stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row[2]), flags=re.IGNORECASE)
                        else:

                            pdf_stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row[1]), flags=re.IGNORECASE)
                        if not row[2]:
                            pdf_desc = clean_text(row[3])
                        else:
                            pdf_desc = clean_text(row[2])

                        if pdf_stock and not pdf_desc:
                            alt_desc = clean_text(row[3]) if len(row) > 3 else ""
                            if alt_desc:
                                pdf_desc = alt_desc
                            else:
                                continue

                        if not pdf_stock or pdf_stock in seen_stocks:
                            continue

                        seen_stocks.add(pdf_stock)
                        unique_rows.append([pdf_stock, pdf_desc])
                        if len(unique_rows) >= 10:
                            break
                    if len(unique_rows) >= 10:
                        break

            # ✅ 与 Cosmos DB 比对并插入必要记录
            diff_rows = []
            for stock, desc in unique_rows:
                # 查询当前项是否存在
                query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
                params = [
                    {"name": "@sheetname", "value": sheetname},
                    {"name": "@fcode", "value": fcode},
                    {"name": "@stock", "value": stock}
                ]
                matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

                # return matched[0]["組入銘柄解説"]

                if matched:
                    old_desc = clean_text(matched[0]["組入銘柄解説"])

                    if old_desc != desc:
                        # ✅ 差异更新
                        matched_item = matched[0]
                        matched_item["組入銘柄解説"] = desc
                        container.replace_item(item=matched_item["id"], body=matched_item)

                        diff_rows.append({
                            "filename": filename,
                            "fcode": fcode,
                            "sheetname": sheetname,
                            "no": 0,
                            "months": "",
                            "stocks": stock,
                            "元組入銘柄解説": old_desc,
                            "新組入銘柄解説": desc,
                            "分類": "銘柄解説更新あり"
                        })
                else:
                    # ✅ 新規插入
                    query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                    max_no = list(container.query_items(
                        query=query_max,
                        parameters=[{"name": "@fcode", "value": fcode}],
                        enable_cross_partition_query=True
                    ))[0] or 0

                    new_item = {
                        "id": str(uuid.uuid4()),
                        "filename": filename,
                        "fcode": fcode,
                        "months": get_prev_month_str(),  # 减1月
                        "no": max_no + 1,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "組入銘柄解説": desc,
                        "コメント": "",
                        "分類": "新規銘柄"
                    }
                    container.create_item(body=new_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": "",
                        "no": max_no + 1,
                        "months": get_prev_month_str(),
                        "新組入銘柄解説": desc,
                        "分類": "新規銘柄"
                    })
            insert_tenbrend_history(diff_rows)
            # update_excel_with_diff_rows(diff_rows, fund_type)

            return diff_rows or "全部一致，无需更新"

        elif sheetname == "過去分整理4列ESG一緒":
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return "PDF下载失败，没有找到pdf"

            tables = extract_pdf_table(io.BytesIO(pdf_response.content))
            if not tables:
                return "PDF中未提取到表格"

            excel_url = f"{PDF_DIR}/10mingbing.xlsx"
            excel_response = requests.get(excel_url)
            if excel_response.status_code != 200:
                return "Excel文件下载失败，不能打开excel"

            wb = load_workbook(filename=io.BytesIO(excel_response.content))
            ws = wb.active

            seen_stocks = set()
            unique_rows = []

            for table in tables:
                header_found = False
                for row in table:
                    if len(row) < 4:
                        continue

                    if (row[1] == "組入銘柄" and
                            "最高益更新回数" in row[2] and
                            "組入銘柄解説" in row[3]):
                        header_found = True
                        continue
                    if not header_found:
                        continue
                    if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                        continue
                    pdf_stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row[1]), flags=re.IGNORECASE)
                    pdf_esg = clean_text(row[2])
                    pdf_desc = clean_text(row[3])

                    if not pdf_stock or pdf_stock in seen_stocks:
                        continue

                    seen_stocks.add(pdf_stock)
                    unique_rows.append([pdf_stock, pdf_desc, pdf_esg])
                    if len(unique_rows) >= 10:
                        break
                if len(unique_rows) >= 10:
                    break

            # ✅ Excel 最后一行写入（调试用）
            for row in unique_rows:
                ws.append(row)

            output_stream = io.BytesIO()
            wb.save(output_stream)
            output_stream.seek(0)
            container_client = get_storage_container()
            blob_client = container_client.get_blob_client("10mingbing.xlsx")
            blob_client.upload_blob(output_stream, overwrite=True)

            # ✅ 比对逻辑
            diff_rows = []
            for stock, desc, esg in unique_rows:
                query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
                params = [
                    {"name": "@sheetname", "value": sheetname},
                    {"name": "@fcode", "value": fcode},
                    {"name": "@stock", "value": stock}
                ]
                matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

                if matched:
                    old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                    old_esg = clean_text(matched[0].get("最高益更新回数", ""))

                    classify = None
                    if old_desc != desc or old_esg != esg:
                        classify = "銘柄解説更新あり"

                    if classify:
                        matched_item = matched[0]
                        matched_item["組入銘柄解説"] = desc
                        matched_item["最高益更新回数"] = esg
                        container.replace_item(item=matched_item["id"], body=matched_item)

                        diff_rows.append({
                            "filename": filename,
                            "fcode": fcode,
                            "sheetname": sheetname,
                            "stocks": stock,
                            "元組入銘柄解説": old_desc,
                            "新組入銘柄解説": desc,
                            "元最高益更新回数": old_esg,
                            "新最高益更新回数": esg,
                            "分類": classify,
                            "no": matched_item.get("no", 0),
                            "months": matched_item.get("months", ""),
                            "分類": "銘柄解説更新あり"
                        })

                else:
                    query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                    max_no = list(container.query_items(
                        query=query_max,
                        parameters=[{"name": "@fcode", "value": fcode}],
                        enable_cross_partition_query=True
                    ))[0] or 0

                    new_item = {
                        "id": str(uuid.uuid4()),
                        "filename": filename,
                        "fcode": fcode,
                        "months": get_prev_month_str(),
                        "no": max_no + 1,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "組入銘柄解説": desc,
                        "最高益更新回数": esg,
                        "コメント": ""
                    }
                    container.create_item(body=new_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": "",
                        "新組入銘柄解説": desc,
                        "最高益更新回数": "",
                        "新最高益更新回数": esg,
                        "分類": "新規銘柄",
                        "no": max_no + 1,
                        "months": get_prev_month_str()
                    })

            insert_tenbrend_history(diff_rows)
            # update_excel_with_diff_rows4(diff_rows, fund_type)

            return diff_rows or "全部一致，无需更新"

        elif sheetname == "過去分整理4列+4列〇二行":
            return handle_sheet_plus42(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname == "過去分整理4列＆（4+1）二行":
            return handle_sheet_plus41(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname == "過去分整理4列上場年月":
            return handle_sheet_plus4(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname == "過去分整理5列＆（5+1）":
            return handle_sheet_plus5(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname in ["300355", "300469", "300481"]:
            return handle_sheet_plus_si4(pdf_url, fcode, sheetname, fund_type, container, filename)
        elif sheetname in ["300449", "300462", "300387"]:
            return handle_sheet_plus_si5(pdf_url, fcode, sheetname, fund_type, container, filename)

        else:
            return "找不到这个sheet页"

    except Exception as e:
        return f"❌ check_tenbrend error: {str(e)}"


def handle_sheet_plus42(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        pdf_response = requests.get(pdf_url)
        if pdf_response.status_code != 200:
            return "PDF下载失败"

        tables = extract_pdf_table(io.BytesIO(pdf_response.content))
        if not tables:
            return "PDF中未提取到表格"

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return "Excel文件下载失败"

        wb = load_workbook(filename=io.BytesIO(excel_response.content))
        ws = wb.active

        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = [row for table in tables for row in table]

        while i < len(all_rows) - 1:
            row1 = all_rows[i]
            row2 = all_rows[i + 1]

            if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                i += 1
                continue

            stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row1[1]), flags=re.IGNORECASE)
            if not stock or stock in seen_stocks:
                i += 1  # ❗ 这里是跳1行而不是2行
                continue

            v1 = clean_text(row1[3])
            v2 = clean_text(row1[4])
            v3 = clean_text(row1[5])
            v4 = clean_text(row1[6])
            desc = clean_text(row1[7]) if len(row1) > 7 else ""
            esg = clean_text(row2[7]) if len(row2) > 7 else ""

            seen_stocks.add(stock)
            unique_rows.append([stock, v1, v2, v3, v4, desc, esg])
            i += 2  # ✅ 只有追加成功才跳过2行

            if len(unique_rows) >= 10:
                break

        diff_rows = []
        for row in unique_rows:
            stock, v1, v2, v3, v4, desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_1 = clean_text(matched[0].get("1", ""))
                old_2 = clean_text(matched[0].get("2", ""))
                old_3 = clean_text(matched[0].get("3", ""))
                old_4 = clean_text(matched[0].get("4", ""))
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                old_esg = clean_text(matched[0].get("ESGへの取り組みが企業価値向上に資する理由", ""))
                if old_desc != desc or old_esg != esg or old_1 != v1 or old_2 != v2 or old_3 != v3 or old_4 != v4:
                    matched_item = matched[0]
                    matched_item.update({
                        "1": v1,
                        "2": v2,
                        "3": v3,
                        "4": v4,
                        "組入銘柄解説": desc,
                        "ESGへの取り組みが企業価値向上に資する理由": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新1": v1,
                        "元1": old_1,
                        "新2": v2,
                        "元2": old_2,
                        "新3": v3,
                        "元3": old_3,
                        "新4": v4,
                        "元4": old_4,
                        "新組入銘柄解説": desc,
                        "元組入銘柄解説": old_desc,
                        "新ESG理由": esg,
                        "元ESG理由": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "1": v1,
                    "2": v2,
                    "3": v3,
                    "4": v4,
                    "組入銘柄解説": desc,
                    "ESGへの取り組みが企業価値向上に資する理由": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                    "コメント": "",
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新1": v1,
                    "元1": "",
                    "新2": v2,
                    "元2": "",
                    "新3": v3,
                    "元3": "",
                    "新4": v4,
                    "元4": "",
                    "新組入銘柄解説": desc,
                    "元組入銘柄解説": "",
                    "新ESG理由": esg,
                    "元ESG理由": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history42(diff_rows)
        # update_excel_with_diff_rows42(diff_rows, fund_type)

        return diff_rows or "全部一致，无需更新"

    except Exception as e:
        return f"❌ handle_sheet_plus42 error: {str(e)}"


def extract_excel_table(file_like,fcode):
    try:
        # 支持传入 BytesIO 或本地路径
        if fcode == "180371-2":
            sheet_name = "PIC_24_S"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
        elif fcode == "140764-5":
            sheet_name = "銘柄紹介"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
        elif fcode == "140793-6":
            sheet_name = "組入銘柄(債券・1)"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
        else:
            sheet_name = "銘柄解説"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
    except Exception as e:
        print(f"❌ Excel 读取失败: {e}")
        return []

    df = df.reset_index(drop=True)
    results = []

    for i in range(len(df) - 1):
        index_val = str(df.iloc[i, 0]).strip()

        if index_val in [str(n) for n in range(1, 11)]:  # 只处理1~10
            stock = clean_text(df.iloc[i, 1])
            desc = clean_text(df.iloc[i, 2])
            esg = clean_text(df.iloc[i + 1, 2])
            if stock:
                results.append([stock, desc, esg])
                
    if fcode == "140793-6":
        sheet_name = "組入銘柄(債券・2)"
        df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
        df = df.reset_index(drop=True)
        for i in range(len(df) - 1):
            index_val = str(df.iloc[i, 0]).strip()

            if index_val in [str(n) for n in range(1, 11)]:  # 只处理1~10
                stock = clean_text(df.iloc[i, 1])
                desc = clean_text(df.iloc[i, 2])
                esg = clean_text(df.iloc[i + 1, 2])
                if stock:
                    results.append([stock, desc, esg])
    return results

def extract_excel_table3(file_like,fcode):
    try:
        # 支持传入 BytesIO 或本地路径
        if fcode == "140193":
            sheet_name = "140193"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:E", dtype=str)
        elif fcode == "140386":
            sheet_name = "140386 (3)"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:E", dtype=str)
        elif fcode == "140565-6":
            sheet_name = "銘柄解説入力ｼｰﾄ"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
        elif fcode == "180291-2":
            sheet_name = "上位10銘柄コメント"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
        else:
            sheet_name = "銘柄解説"
            df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:E", dtype=str)
    except Exception as e:
        print(f"❌ Excel 读取失败: {e}")
        return []

    df = df.reset_index(drop=True)
    results = []

    for i in range(len(df) - 1):
        index_val = str(df.iloc[i, 0]).strip()

        if index_val in [str(n) for n in range(1, 11)]:  # 只处理1~10
            stock = clean_text(df.iloc[i, 1])
            if fcode == "140193":
                desc = clean_text(df.iloc[i, 4])
            elif fcode == "140565-6":
                desc = clean_text(df.iloc[i, 3])
            else:
                desc = clean_text(df.iloc[i, 2])
            if stock:
                results.append([stock, desc])

    return results

def handle_sheet_plus41(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        if fcode in ['140752', '140302-3','180371-2','140389-90','140764-5','140793-6']:
            # 将 .pdf 替换为 .xlsx 作为 Excel 文件路径
            excel_url = pdf_url.replace(".pdf", ".xlsx")

            # 下载 Excel 文件内容
            response = requests.get(excel_url)
            if response.status_code != 200:
                return "Excel下载失败"

            # 转为 BytesIO 对象传给 extract_excel_table
            excel_file = io.BytesIO(response.content)
            tables = extract_excel_table(excel_file,fcode)
        else:
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return "PDF下载失败"

            tables = extract_structured_tables(io.BytesIO(pdf_response.content))

        if not tables:
            return "PDF中未提取到表格"

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return "Excel文件下载失败"

        wb = load_workbook(filename=io.BytesIO(excel_response.content))
        ws = wb.active

        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = tables

        if fcode in ['140752', '140302-3','180371-2','140389-90','140764-5','140793-6']:
            for row in all_rows:
                stock = clean_text(row[0])
                desc = clean_text(row[1])
                esg = clean_text(row[2])
                seen_stocks.add(stock)
                unique_rows.append([stock, desc, esg])

                if len(unique_rows) >= 10:
                    break
        else:

            while i < len(all_rows) - 1:
                row1 = all_rows[i]
                row2 = all_rows[i + 1]

                if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                    i += 1
                    continue

                stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row1[1]), flags=re.IGNORECASE)
                
                if not stock or stock in seen_stocks:
                    i += 1  # ❗ 这里是跳1行而不是2行
                    continue
                if fcode in ["140793-6", "140764-5"]:
                    desc = clean_text(row1[3]) if len(row1) > 2 else ""
                    esg = clean_text(row2[3]) if len(row2) > 2 else ""
                else:

                    desc = clean_text(row1[2]) if len(row1) > 2 else ""
                    esg = clean_text(row2[2]) if len(row2) > 2 else ""

                seen_stocks.add(stock)
                unique_rows.append([stock, desc, esg])
                i += 2  # ✅ 只有追加成功才跳过2行

                if len(unique_rows) >= 10:
                    break

        diff_rows = []
        for row in unique_rows:
            stock, desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                old_esg = clean_text(matched[0].get("ESGへの取り組みが企業価値向上に資する理由", ""))
                if old_desc != desc or old_esg != esg:
                    matched_item = matched[0]
                    matched_item.update({
                        "組入銘柄解説": desc,
                        "ESGへの取り組みが企業価値向上に資する理由": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新組入銘柄解説": desc,
                        "元組入銘柄解説": old_desc,
                        "新ESG理由": esg,
                        "元ESG理由": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "組入銘柄解説": desc,
                    "ESGへの取り組みが企業価値向上に資する理由": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                    "コメント": "",
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新組入銘柄解説": desc,
                    "元組入銘柄解説": "",
                    "新ESG理由": esg,
                    "元ESG理由": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history41(diff_rows)
        # update_excel_with_diff_rows41(diff_rows, fund_type)

        return diff_rows or "全部一致，无需更新"

    except Exception as e:
        return f"❌ handle_sheet_plus41 error: {str(e)}"


# 往10铭柄的履历表里写
def insert_tenbrend_history4(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "元上場年月": record["元上場年月"],
            "新上場年月": record["新上場年月"],
            "分類": record["分類"],
            "no": record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)
    
def format_date(value):
    try:
        if pd.isna(value):
            return ""

        # 尝试将纯数字字符串当作数字处理
        try:
            value_numeric = float(value)
            is_numeric = True
        except (ValueError, TypeError):
            is_numeric = False

        if is_numeric:
            base = datetime(1899, 12, 30)  # Excel序列号起点
            real_date = base + timedelta(days=value_numeric)
            return f"{real_date.year}年{real_date.month}月"

        # datetime 或 pd.Timestamp 类型
        elif isinstance(value, (datetime, pd.Timestamp)):
            return f"{value.year}年{value.month}月"

        # 其余字符串
        else:
            parsed = pd.to_datetime(str(value), errors='coerce')
            if pd.isna(parsed):
                return str(value)
            return f"{parsed.year}年{parsed.month}月"

    except Exception:
        return str(value)


def extract_excel_table4(file_like):
    try:
        # 支持传入 BytesIO 或本地路径
        sheet_name = "組入銘柄"
        df = pd.read_excel(file_like, sheet_name=sheet_name, header=1, usecols="A:D", dtype=str)
    except Exception as e:
        print(f"❌ Excel 读取失败: {e}")
        return []

    df = df.reset_index(drop=True)
    results = []

    for i in range(len(df)):
        index_val = str(df.iloc[i, 0]).strip()
        if index_val in [str(n) for n in range(1, 11)]:
            stock = clean_text(df.iloc[i, 1])
            desc = clean_text(df.iloc[i, 2])
            date_val = df.iloc[i, 3]
            date_str = format_date(date_val)
            if stock:
                results.append([stock, desc, date_str])

    return results


def handle_sheet_plus4(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        if fcode in ['140749']:
            # 将 .pdf 替换为 .xlsx 作为 Excel 文件路径
            excel_url = pdf_url.replace(".pdf", ".xlsx")

            # 下载 Excel 文件内容
            response = requests.get(excel_url)
            if response.status_code != 200:
                return "Excel下载失败"

            # 转为 BytesIO 对象传给 extract_excel_table4
            excel_file = io.BytesIO(response.content)
            tables = extract_excel_table4(excel_file)
            
        else:
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return "PDF下载失败，没有找到pdf"

            tables = extract_pdf_table(io.BytesIO(pdf_response.content))
            if not tables:
                return "PDF中未提取到表格"

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return "Excel文件下载失败，不能打开excel"

        wb = load_workbook(filename=io.BytesIO(excel_response.content))
        ws = wb.active

        seen_stocks = set()
        unique_rows = []
        if fcode in ['140749']:
            for row in tables:

                pdf_stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row[0]), flags=re.IGNORECASE)

                pdf_desc = clean_text(row[1])
                pdf_esg = re.sub(r"(\d{4})年(\d{1,2})月", lambda m: f"{m.group(1)}/{int(m.group(2))}/1", clean_text(row[2]))

                seen_stocks.add(pdf_stock)
                unique_rows.append([pdf_stock, pdf_desc, pdf_esg])
                if len(unique_rows) >= 10:
                    break
        else:

            for table in tables:
                header_found = False
                for row in table:
                    if len(row) < 4:
                        continue

                    if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                        continue
                    pdf_stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row[1]), flags=re.IGNORECASE)
                    pdf_desc = clean_text(row[2])
                    pdf_esg = re.sub(r"(\d{4})年(\d{1,2})月", lambda m: f"{m.group(1)}/{int(m.group(2))}/1",
                                     clean_text(row[3]))

                    if not pdf_stock or pdf_stock in seen_stocks:
                        continue

                    seen_stocks.add(pdf_stock)
                    unique_rows.append([pdf_stock, pdf_desc, pdf_esg])
                    if len(unique_rows) >= 10:
                        break
                if len(unique_rows) >= 10:
                    break

        # ✅ Excel 最后一行写入（调试用）
        for row in unique_rows:
            ws.append(row)

        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        container_client = get_storage_container()
        blob_client = container_client.get_blob_client("10mingbing.xlsx")
        blob_client.upload_blob(output_stream, overwrite=True)

        # ✅ 比对逻辑
        diff_rows = []
        for stock, desc, esg in unique_rows:
            query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                old_esg = clean_text(matched[0].get("上場年月", ""))

                classify = None
                if old_desc != desc or old_esg != esg:
                    classify = "銘柄解説更新あり"

                if classify:
                    matched_item = matched[0]
                    matched_item["組入銘柄解説"] = desc
                    matched_item["上場年月"] = esg
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": old_desc,
                        "新組入銘柄解説": desc,
                        "元上場年月": old_esg,
                        "新上場年月": esg,
                        "分類": classify,
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", ""),
                        "分類": "銘柄解説更新あり"
                    })

            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "months": get_prev_month_str(),
                    "no": max_no + 1,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "組入銘柄解説": desc,
                    "上場年月": esg,
                    "コメント": ""
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "元組入銘柄解説": "",
                    "新組入銘柄解説": desc,
                    "元上場年月": "",
                    "新上場年月": esg,
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history4(diff_rows)
        # update_excel_with_diff_rows_shang(diff_rows, fund_type)

        return diff_rows or "全部一致，无需更新"

    except Exception as e:
        return f"❌ handle_sheet_4plus41 error: {str(e)}"


def extract_excel_table5(excel_file,fcode):
    try:
        # 支持传入 BytesIO 或本地路径
        if fcode == "140312-3":
            sheet_name = "PIC_24_S"
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=1, usecols="A:G", dtype=str)
        # 支持传入 BytesIO 或本地路径
        else:
            sheet_name = "銘柄解説"
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=1, usecols="A:G", dtype=str)

    except Exception as e:
        print(f"❌ Excel 读取失败: {e}")
        return []

    df = df.reset_index(drop=True)
    results = []

    for i in range(len(df) - 1):
        no_val = str(df.iloc[i, 0]).strip()
        if not no_val.isdigit():
            continue  # 只处理数字编号行

        stock = clean_text(df.iloc[i, 1])  # 銘柄名
        category = clean_text(df.iloc[i, 2])  # 分野
        tmp_cat = clean_text(df.iloc[i, 4])
        if fcode == "140312-3":
            desc = clean_text(df.iloc[i, 3])  # 組入銘柄解説（G列，第1行）
            esg = clean_text(df.iloc[i + 1, 3])  # ESG理由（G列，第2行）
        elif category == tmp_cat or tmp_cat == '':
            desc = clean_text(df.iloc[i, 6])  # 組入銘柄解説（G列，第1行）
            esg = clean_text(df.iloc[i + 1, 6])  # ESG理由（G列，第2行）
        else:
            desc = clean_text(df.iloc[i, 4])  # 組入銘柄解説（G列，第1行）
            esg = clean_text(df.iloc[i + 1, 4])  # ESG理由（G列，第2行）

        if stock:
            results.append([stock, category, desc, esg])

    return results


def handle_sheet_plus5(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        if fcode in ['140787', '180342-3','140312-3']:
            # 将 .pdf 替换为 .xlsx 作为 Excel 文件路径
            excel_url = pdf_url.replace(".pdf", ".xlsx")

            # 下载 Excel 文件内容
            response = requests.get(excel_url)
            if response.status_code != 200:
                return "Excel下载失败"

            # 转为 BytesIO 对象传给 extract_excel_table5
            excel_file = io.BytesIO(response.content)
            tables = extract_excel_table5(excel_file,fcode)
        else:
            pdf_response = requests.get(pdf_url)
            if pdf_response.status_code != 200:
                return "PDF下载失败"

            tables = extract_structured_tables(io.BytesIO(pdf_response.content))
        if not tables:
            return "PDF中未提取到表格"

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return "Excel文件下载失败"

        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = tables

        if fcode in ['140787', '180342-3']:
            for row in all_rows:

                stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row[0]), flags=re.IGNORECASE)
                keti = clean_text(row[1])
                desc = clean_text(row[2])
                esg = clean_text(row[3])

                seen_stocks.add(stock)
                unique_rows.append([stock, keti, desc, esg])
                if len(unique_rows) >= 10:
                    break
        else:

            while i < len(all_rows) - 1:
                row1 = all_rows[i]
                row2 = all_rows[i + 1]

                if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                    i += 1
                    continue

                stock = clean_text(row1[1])
                stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row1[1]), flags=re.IGNORECASE)
                if not stock or stock in seen_stocks:
                    i += 1  # ❗ 这里是跳1行而不是2行
                    continue
                keti = clean_text(row1[2]) if len(row1) > 3 else ""
                if fcode in ["140793-6", "140406-7", "180340-1"]:
                    desc = clean_text(row1[3]) if len(row1) > 3 else ""
                    esg = clean_text(row2[3]) if len(row2) > 3 else ""
                else:
                    desc = clean_text(row1[4]) if len(row1) > 3 else ""
                    esg = clean_text(row2[4]) if len(row2) > 3 else ""

                seen_stocks.add(stock)
                unique_rows.append([stock, keti, desc, esg])
                i += 2  # ✅ 只有追加成功才跳过2行

                if len(unique_rows) >= 10:
                    break

        diff_rows = []
        for row in unique_rows:
            stock, keti, desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_keti = clean_text(matched[0].get("解決すべき社会的課題", ""))
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))
                old_esg = clean_text(matched[0].get("ESGへの取り組みが企業価値向上に資する理由", ""))
                if old_desc != desc or old_esg != esg:
                    matched_item = matched[0]
                    matched_item.update({
                        "解決すべき社会的課題": keti,
                        "組入銘柄解説": desc,
                        "ESGへの取り組みが企業価値向上に資する理由": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新解決すべき社会的課題": keti,
                        "元解決すべき社会的課題": old_keti,
                        "新組入銘柄解説": desc,
                        "元組入銘柄解説": old_desc,
                        "新ESG理由": esg,
                        "元ESG理由": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "解決すべき社会的課題": keti,
                    "組入銘柄解説": desc,
                    "ESGへの取り組みが企業価値向上に資する理由": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新解決すべき社会的課題": keti,
                    "元解決すべき社会的課題": "",
                    "新組入銘柄解説": desc,
                    "元組入銘柄解説": "",
                    "新ESG理由": esg,
                    "元ESG理由": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history5(diff_rows)
        # update_excel_with_diff_rows5(diff_rows, fund_type)

        return diff_rows or "全部一致，无需更新"

    except Exception as e:
        return f"❌ handle_sheet_plus5 error: {str(e)}"


# 私募相关的处理
def insert_tenbrend_history_si4(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "元組入銘柄解説": record["元組入銘柄解説"],
            "新組入銘柄解説": record["新組入銘柄解説"],
            "分類": record["分類"],
            "no": record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)


def update_excel_with_diff_si4(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_desc = row["新組入銘柄解説"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]

        stock_col = find_column_by_keyword(header_row, ["銘柄"])
        desc_col = find_column_by_keyword(header_row, ["組入銘柄解説", "銘柄解説"])
        no_col = find_column_by_keyword(header_row, ["No", "NO"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


def handle_sheet_plus_si4(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        pdf_response = requests.get(pdf_url)
        if pdf_response.status_code != 200:
            return "PDF下载失败，没有找到pdf"

        tables = extract_pdf_table(io.BytesIO(pdf_response.content))
        if not tables:
            return "PDF中未提取到表格"

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return "Excel文件下载失败，不能打开excel"

        seen_stocks = set()
        unique_rows = []

        for table in tables:
            for row in table:
                if len(row) < 3:
                    continue

                if not row or str(row[0]).strip() not in [str(i) for i in range(1, 11)]:
                    continue
                
                pdf_stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row[1]), flags=re.IGNORECASE)
                if not row[2]:
                    pdf_desc = clean_text(row[3]) if len(row) > 3 else ""
                else:
                    pdf_desc = clean_text(row[2])

                if not pdf_stock or pdf_stock in seen_stocks:
                    continue

                seen_stocks.add(pdf_stock)
                unique_rows.append([pdf_stock, pdf_desc])
                if len(unique_rows) >= 10:
                    break
            if len(unique_rows) >= 10:
                break

        # ✅ 比对逻辑
        diff_rows = []
        for stock, desc in unique_rows:
            query = """
                    SELECT * FROM c
                    WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
                """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_desc = clean_text(matched[0].get("組入銘柄解説", ""))

                classify = None
                if old_desc != desc:
                    classify = "銘柄解説更新あり"

                if classify:
                    matched_item = matched[0]
                    matched_item["組入銘柄解説"] = desc
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "元組入銘柄解説": old_desc,
                        "新組入銘柄解説": desc,
                        "分類": classify,
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", ""),
                        "分類": "銘柄解説更新あり"
                    })

            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "months": get_prev_month_str(),
                    "no": max_no + 1,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "組入銘柄解説": desc,
                    "コメント": ""
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "元組入銘柄解説": "",
                    "新組入銘柄解説": desc,
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history_si4(diff_rows)
        # update_excel_with_diff_si4(diff_rows, fund_type)

        return diff_rows or "全部一致，无需更新"

    except Exception as e:
        return f"❌ handle_sheet_4plus41 error: {str(e)}"


def insert_tenbrend_history_si5(diff_rows):
    container = get_db_connection(TENBREND_CONTAINER_NAME)

    for record in diff_rows:
        history_item = {
            "id": str(uuid.uuid4()),
            "filename": record["filename"],
            "fcode": record["fcode"],
            "sheetname": record["sheetname"],
            "stocks": record["stocks"],
            "新社会的課題": record["新社会的課題"],
            "元社会的課題": record["元社会的課題"],
            "新コメント": record["新コメント"],
            "元コメント": record["元コメント"],
            "新ESGコメント": record["新ESGコメント"],
            "元ESGコメント": record["元ESGコメント"],
            "分類": record["分類"],
            "no": record["no"],
            "created_at": datetime.now(UTC).isoformat()  # ✅ 当前时间
        }
        container.create_item(body=history_item)


def update_excel_with_diff_si5(diff_rows, fund_type):
    if not diff_rows:
        return

    if fund_type == "public":
        target_excel_name = "10銘柄マスタ管理_公募.xlsx"
    else:
        # 如需私募逻辑可扩展
        target_excel_name = "10銘柄マスタ管理_私募.xlsx"

    # 下载原始 Excel
    excel_url = f"{PDF_DIR}/{target_excel_name}"
    excel_response = requests.get(excel_url)
    if excel_response.status_code != 200:
        raise Exception("Excel文件下载失败")

    wb = load_workbook(filename=io.BytesIO(excel_response.content))

    for row in diff_rows:
        sheetname = row["sheetname"]
        fcode = row["fcode"]
        stock = row["stocks"]
        classify = row["分類"]
        no = row["no"]
        months = row["months"]
        new_keti = row["新社会的課題"]
        new_desc = row["新コメント"]
        new_esg = row["新ESGコメント"]

        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]

        # 获取表头位置
        header_row = ws[3]

        stock_col = find_column_by_keyword(header_row, ["組入銘柄", "銘柄", "銘柄名"])
        keti_col = find_column_by_keyword(header_row, ["社会的課題", "目指すインパクト"])
        desc_col = find_column_by_keyword(header_row, ["コメント"])
        esg_col = find_column_by_keyword(header_row, ["ESGコメント"])  # 仅当你处理ESG表时需要
        no_col = find_column_by_keyword(header_row, ["No"])
        months_col = find_column_by_keyword(header_row, ["決算月"])
        fcode_col = find_column_by_keyword(header_row, ["Fコード"])

        if stock_col is None or desc_col is None:
            continue

        # 查找 fcode 所属块的范围
        if fcode_col is None:
            continue

        last_row = ws.max_row
        target_row_idx = None
        fcode_block_end = None

        for row_idx in range(2, last_row + 1):
            if str(ws.cell(row=row_idx, column=fcode_col + 1).value).strip() == fcode:
                if fcode_block_end is None or row_idx > fcode_block_end:
                    fcode_block_end = row_idx
                # 查找是否已存在该 stock
                current_stock = str(ws.cell(row=row_idx, column=stock_col + 1).value).strip()
                current_stock = clean_text(current_stock)
                if current_stock == stock:
                    target_row_idx = row_idx
                    break

        if classify == "新規銘柄" and fcode_block_end:
            # 插入新规到 fcode 组最后一行的下一行
            insert_idx = fcode_block_end + 1
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.insert_rows(insert_idx)
            copy_row_style(ws, insert_idx - 1, insert_idx)
            ws.cell(row=insert_idx, column=fcode_col + 1, value=fcode)
            write_wrapped_stock_cell(ws, insert_idx, stock_col + 1, stock)
            ws.cell(row=insert_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=insert_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=insert_idx, column=no_col + 1, value=no)
            ws.cell(row=insert_idx, column=months_col + 1, value=months)
            ws.cell(row=insert_idx, column=esg_col + 1, value=new_esg)

        elif classify == "銘柄解説更新あり" and target_row_idx:
            # 直接更新原值
            ws.cell(row=target_row_idx, column=keti_col + 1, value=new_keti)
            ws.cell(row=target_row_idx, column=desc_col + 1, value=new_desc)
            ws.cell(row=target_row_idx, column=esg_col + 1, value=new_esg)

    # 上传到原始 Blob 路径（覆盖）
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    container_client = get_storage_container()
    blob_client = container_client.get_blob_client(target_excel_name)
    blob_client.upload_blob(output_stream, overwrite=True)


def clean_text_si(text):
    return text.replace("\n", "").replace(" ", " ").strip()


def split_by_numbered_blocks(text_block):
    parts = re.split(r'\n?(?=\s*([1-9]|10)[^\d])', text_block)
    combined_parts = []
    i = 1
    while i < len(parts):
        num = parts[i].strip()
        content = parts[i + 1].strip() if i + 1 < len(parts) else ""
        combined_parts.append(f"{num} {content}")
        i += 2

    results = []
    for part in combined_parts:
        match = re.match(r"^([1-9]|10)[\s ]*([^\s\d]{2,30})[\s ]*([\s\S]+)", part)
        if match:
            no = match.group(1)
            company = clean_text_si(match.group(2))
            description = clean_text_si(match.group(3))
            # 如果三项都不为空，再加入
            if no and company and description:
                results.append([no, company, description])
    return results


def extract_structured_tables(pdf_input):
    all_rows = []
    with pdfplumber.open(pdf_input) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "組入上位10銘柄の解説" not in text:
                continue

            tables = page.extract_tables()
            if not tables:
                continue

            for table in tables:
                if len(table) == 1 and len(table[0]) == 1:
                    # 说明是合并文本块（单元格中是段落文字）
                    text_block = table[0][0]
                    rows = split_by_numbered_blocks(text_block)
                    all_rows.extend(rows)
                else:
                    # 普通表格结构
                    for row in table:
                        cleaned_row = [clean_text_si(cell) if cell else "" for cell in row]
                        # ✅ 过滤掉字段数量少于3的行
                        if len(cleaned_row) >= 3:
                            all_rows.append(cleaned_row)
    return all_rows


def handle_sheet_plus_si5(pdf_url, fcode, sheetname, fund_type, container, filename):
    try:
        pdf_response = requests.get(pdf_url)
        if pdf_response.status_code != 200:
            return "PDF下载失败"

        tables = extract_structured_tables(io.BytesIO(pdf_response.content))
        if not tables:
            return "PDF中未提取到表格"

        excel_url = f"{PDF_DIR}/10mingbing.xlsx"
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            return "Excel文件下载失败"

        seen_stocks = set()
        unique_rows = []

        i = 0

        # 合并所有表格行为一个大列表
        all_rows = tables

        while i < len(all_rows) - 1:
            row1 = all_rows[i]
            row2 = all_rows[i + 1]

            if len(row1) < 2 or str(row1[0]).strip() not in [str(n) for n in range(1, 11)]:
                i += 1
                continue

            stock = re.sub(r'^(NEW\s*|new\s*)|(\s*NEW|\s*new)$', '', clean_text(row1[1]), flags=re.IGNORECASE)
            if not stock or stock in seen_stocks:
                i += 1  # ❗ 这里是跳1行而不是2行
                continue
            keti = clean_text(row1[2]) if len(row1) > 3 else ""
            desc = clean_text(row1[3]) if len(row1) > 3 else ""
            esg = clean_text(row2[3]) if len(row2) > 3 else ""

            seen_stocks.add(stock)
            unique_rows.append([stock, keti, desc, esg])
            i += 2  # ✅ 只有追加成功才跳过2行

            if len(unique_rows) >= 10:
                break

        diff_rows = []
        for row in unique_rows:
            stock, keti, desc, esg = row
            query = """
                SELECT * FROM c
                WHERE c.sheetname = @sheetname AND c.fcode = @fcode AND c.stocks = @stock
            """
            params = [
                {"name": "@sheetname", "value": sheetname},
                {"name": "@fcode", "value": fcode},
                {"name": "@stock", "value": stock}
            ]
            matched = list(container.query_items(query=query, parameters=params, enable_cross_partition_query=True))

            if matched:
                old_keti = clean_text(matched[0].get("社会的課題", ""))
                old_desc = clean_text(matched[0].get("コメント", ""))
                old_esg = clean_text(matched[0].get("ESGコメント", ""))
                if old_desc != desc or old_esg != esg:
                    matched_item = matched[0]
                    matched_item.update({
                        "社会的課題": keti,
                        "コメント": desc,
                        "ESGコメント": esg
                    })
                    container.replace_item(item=matched_item["id"], body=matched_item)

                    diff_rows.append({
                        "filename": filename,
                        "fcode": fcode,
                        "sheetname": sheetname,
                        "stocks": stock,
                        "新社会的課題": keti,
                        "元社会的課題": old_keti,
                        "新コメント": desc,
                        "元コメント": old_desc,
                        "新ESGコメント": esg,
                        "元ESGコメント": old_esg,
                        "分類": "銘柄解説更新あり",
                        "no": matched_item.get("no", 0),
                        "months": matched_item.get("months", "")
                    })
            else:
                query_max = "SELECT VALUE MAX(c.no) FROM c WHERE c.fcode = @fcode"
                max_no = list(container.query_items(
                    query=query_max,
                    parameters=[{"name": "@fcode", "value": fcode}],
                    enable_cross_partition_query=True
                ))[0] or 0

                new_item = {
                    "id": str(uuid.uuid4()),
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "社会的課題": keti,
                    "コメント": desc,
                    "ESGコメント": esg,
                    "no": max_no + 1,
                    "months": get_prev_month_str(),
                }
                container.create_item(body=new_item)

                diff_rows.append({
                    "filename": filename,
                    "fcode": fcode,
                    "sheetname": sheetname,
                    "stocks": stock,
                    "新社会的課題": keti,
                    "元社会的課題": "",
                    "新コメント": desc,
                    "元コメント": "",
                    "新ESGコメント": esg,
                    "元ESGコメント": "",
                    "分類": "新規銘柄",
                    "no": max_no + 1,
                    "months": get_prev_month_str()
                })

        insert_tenbrend_history_si5(diff_rows)
        # update_excel_with_diff_si5(diff_rows, fund_type)

        return diff_rows or "全部一致，无需更新"

    except Exception as e:
        return f"❌ handle_sheet_plussi5 error: {str(e)}"




app = WsgiToAsgi(app)

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True) # 启用HTTPS, ssl_context='adhoc'
